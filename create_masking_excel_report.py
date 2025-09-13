#!/usr/bin/env python3
"""
Gold3 Database Masking Coverage Report - Excel Generator
Creates a professional Excel report with multiple formatted worksheets
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def create_excel_report():
    """Create comprehensive Excel report with multiple worksheets"""

    # Create Excel writer
    excel_file = "Gold3_Masking_Coverage_Report.xlsx"
    writer = pd.ExcelWriter(excel_file, engine="openpyxl")

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12, color="000000")
    warning_font = Font(color="FF6B35", bold=True)
    success_font = Font(color="2E8B57", bold=True)
    critical_font = Font(color="DC143C", bold=True)

    # Color fills
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    critical_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    # Border styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ===== WORKSHEET 1: EXECUTIVE SUMMARY =====
    summary_data = {
        "Metric": [
            "Report Date",
            "Database",
            "Total Tables",
            "Total Columns",
            "Masked Columns",
            "Coverage Rate",
            "Risk Level",
        ],
        "Value": [
            "September 12, 2025",
            "gchub_dev",
            "122",
            "1,052",
            "70",
            "6.65%",
            "MEDIUM-HIGH",
        ],
        "Status": ["📅", "🗄️", "📋", "📊", "✅", "⚠️", "🟡"],
    }

    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name="Executive Summary", index=False, startrow=1)

    worksheet = writer.sheets["Executive Summary"]

    # Title
    worksheet["A1"] = "🔒 Gold3 Database - Data Masking Coverage Report"
    worksheet["A1"].font = Font(bold=True, size=16, color="2E8B57")
    worksheet.merge_cells("A1:C1")

    # Format headers
    for col in range(1, 4):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(3, 10):
        for col in range(1, 4):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 3:  # Status column
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in range(1, 4):
        worksheet.column_dimensions[get_column_letter(col)].width = 20

    # ===== WORKSHEET 2: MASKED TABLES =====
    masked_tables_data = {
        "Table": ["auth_user", "workflow_job", "workflow_jobaddress", "workflow_item"],
        "Total Columns": [11, 52, 15, 112],
        "Masked Columns": [4, 11, 8, 6],
        "Coverage %": ["36.4%", "21.2%", "53.3%", "5.4%"],
        "Status": ["✅ Protected", "✅ Protected", "✅ Protected", "✅ Protected"],
    }

    df_masked = pd.DataFrame(masked_tables_data)
    df_masked.to_excel(writer, sheet_name="Protected Tables", index=False, startrow=1)

    worksheet = writer.sheets["Protected Tables"]

    # Title
    worksheet["A1"] = "✅ FULLY PROTECTED TABLES"
    worksheet["A1"].font = Font(bold=True, size=14, color="2E8B57")
    worksheet.merge_cells("A1:E1")

    # Format headers
    for col in range(1, 6):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(3, 7):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 5:  # Status column
                cell.font = success_font

    # Auto-adjust column widths
    for col in range(1, 6):
        worksheet.column_dimensions[get_column_letter(col)].width = 18

    # ===== WORKSHEET 3: CRITICAL RISKS =====
    critical_data = {
        "Table": [
            "accounts_userprofile",
            "accounts_userprofile_backup",
            "address_contact",
            "workflow_salesservicerep",
            "address_contact",
            "art_req_artreq",
            "art_req_artreq",
            "workflow_salesservicerep",
        ],
        "Column": [
            "phone_number",
            "phone_number",
            "phone",
            "phone",
            "email",
            "contact_email",
            "ship_to_email",
            "email",
        ],
        "Data Type": ["character varying"] * 8,
        "Risk Level": [
            "CRITICAL",
            "CRITICAL",
            "CRITICAL",
            "HIGH",
            "HIGH",
            "HIGH",
            "HIGH",
            "HIGH",
        ],
        "Impact": [
            "Direct contact info exposure",
            "Backup data vulnerability",
            "Contact database exposure",
            "Sales contact exposure",
            "Contact database exposure",
            "Art request contact exposure",
            "Shipping contact exposure",
            "Sales contact exposure",
        ],
    }

    df_critical = pd.DataFrame(critical_data)
    df_critical.to_excel(writer, sheet_name="Critical Risks", index=False, startrow=1)

    worksheet = writer.sheets["Critical Risks"]

    # Title
    worksheet["A1"] = "🚨 CRITICAL SECURITY GAPS (Immediate Action Required)"
    worksheet["A1"].font = Font(bold=True, size=14, color="DC143C")
    worksheet.merge_cells("A1:E1")

    # Format headers
    for col in range(1, 6):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = critical_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(3, 11):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 4:  # Risk Level column
                risk_level = cell.value
                if risk_level == "CRITICAL":
                    cell.font = critical_font
                    cell.fill = critical_fill
                elif risk_level == "HIGH":
                    cell.font = warning_font
                    cell.fill = warning_fill

    # Auto-adjust column widths
    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 30

    # ===== WORKSHEET 4: COVERAGE ANALYSIS =====
    coverage_data = {
        "Category": [
            "Phone Numbers",
            "Email Addresses",
            "Personal Names",
            "Addresses & Locations",
        ],
        "Count": [4, 4, 4, 3],
        "Risk Level": ["CRITICAL", "HIGH", "MEDIUM", "LOW"],
        "Priority": ["IMMEDIATE", "HIGH", "MEDIUM", "LOW"],
    }

    df_coverage = pd.DataFrame(coverage_data)
    df_coverage.to_excel(writer, sheet_name="Coverage Analysis", index=False, startrow=1)

    worksheet = writer.sheets["Coverage Analysis"]

    # Title
    worksheet["A1"] = "📈 COVERAGE ANALYSIS DASHBOARD"
    worksheet["A1"].font = Font(bold=True, size=14, color="2E8B57")
    worksheet.merge_cells("A1:D1")

    # Overall stats
    worksheet["A3"] = "Overall Statistics:"
    worksheet["A3"].font = subheader_font
    worksheet["A4"] = "Total Columns:"
    worksheet["B4"] = "1,052"
    worksheet["A5"] = "Masked Columns:"
    worksheet["B5"] = "70"
    worksheet["A6"] = "Coverage Rate:"
    worksheet["B6"] = "6.65%"
    worksheet["B6"].font = warning_font

    # Format headers
    for col in range(1, 5):
        cell = worksheet.cell(row=8, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(9, 13):
        for col in range(1, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 3:  # Risk Level column
                risk_level = cell.value
                if risk_level == "CRITICAL":
                    cell.font = critical_font
                    cell.fill = critical_fill
                elif risk_level == "HIGH":
                    cell.font = warning_font
                    cell.fill = warning_fill

    # Auto-adjust column widths
    for col in range(1, 5):
        worksheet.column_dimensions[get_column_letter(col)].width = 18

    # ===== WORKSHEET 5: ACTION PLAN =====
    action_data = {
        "Phase": [
            "Phase 1: Critical Fixes",
            "Phase 2: Business Data",
            "Phase 3: Extended Data",
        ],
        "Timeline": ["Week 1", "Week 2", "Week 3"],
        "Priority": ["IMMEDIATE", "HIGH", "MEDIUM"],
        "Status": ["Not Started", "Not Started", "Not Started"],
        "Description": [
            "User profiles, contact info, phone numbers",
            "Sales reps, art requests, business contacts",
            "Comments, additional sensitive fields",
        ],
    }

    df_action = pd.DataFrame(action_data)
    df_action.to_excel(writer, sheet_name="Action Plan", index=False, startrow=1)

    worksheet = writer.sheets["Action Plan"]

    # Title
    worksheet["A1"] = "🛠️ RECOMMENDED ACTION PLAN"
    worksheet["A1"].font = Font(bold=True, size=14, color="2E8B57")
    worksheet.merge_cells("A1:E1")

    # Format headers
    for col in range(1, 6):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(3, 6):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 3:  # Priority column
                priority = cell.value
                if priority == "IMMEDIATE":
                    cell.font = critical_font
                    cell.fill = critical_fill
                elif priority == "HIGH":
                    cell.font = warning_font
                    cell.fill = warning_fill

    # Auto-adjust column widths
    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 15
    worksheet.column_dimensions["E"].width = 35

    # ===== WORKSHEET 6: SUCCESS METRICS =====
    metrics_data = {
        "Metric": [
            "Overall Coverage",
            "Critical Data Protected",
            "High-Risk Fields",
            "Business Impact",
        ],
        "Current": ["6.65%", "25%", "7 exposed", "High Risk"],
        "Target": ["85%", "100%", "0 exposed", "Low Risk"],
        "Status": ["🔴 Behind", "🟡 Progress", "🔴 Action Needed", "🟡 Improving"],
    }

    df_metrics = pd.DataFrame(metrics_data)
    df_metrics.to_excel(writer, sheet_name="Success Metrics", index=False, startrow=1)

    worksheet = writer.sheets["Success Metrics"]

    # Title
    worksheet["A1"] = "🎯 SUCCESS METRICS"
    worksheet["A1"].font = Font(bold=True, size=14, color="2E8B57")
    worksheet.merge_cells("A1:D1")

    # Format headers
    for col in range(1, 5):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format data rows
    for row in range(3, 7):
        for col in range(1, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 4:  # Status column
                status = cell.value
                if "🔴" in status:
                    cell.font = critical_font
                    cell.fill = critical_fill
                elif "🟡" in status:
                    cell.font = warning_font
                    cell.fill = warning_fill

    # Auto-adjust column widths
    for col in range(1, 5):
        worksheet.column_dimensions[get_column_letter(col)].width = 20

    # Save the Excel file
    writer.close()

    print(f"✅ Excel report created successfully: {excel_file}")
    print("📊 Worksheets created:")
    print("  1. Executive Summary")
    print("  2. Protected Tables")
    print("  3. Critical Risks")
    print("  4. Coverage Analysis")
    print("  5. Action Plan")
    print("  6. Success Metrics")

    return excel_file


if __name__ == "__main__":
    try:
        excel_file = create_excel_report()
        print(f"\n🎉 Report saved as: {os.path.abspath(excel_file)}")
    except ImportError as e:
        print(f"❌ Missing required package: {e}")
        print("Please install required packages:")
        print("pip install pandas openpyxl")
    except Exception as e:
        print(f"❌ Error creating Excel report: {e}")
