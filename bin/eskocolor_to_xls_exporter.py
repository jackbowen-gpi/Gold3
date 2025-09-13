"""Export Esko color book data into an XLS spreadsheet for analysis."""

# Setup the Django environment
import bin_functions
import openpyxl
from colormath.color_objects import LabColor

bin_functions.setup_paths()
import django

django.setup()
# Back to the ordinary imports

from esko_color.color_reader import InvalidColor
from esko_color.colorbook_reader import EskoColorBook

from gchub_db.apps.color_mgt.models import ColorDefinition

# Esko color book name.
cb = EskoColorBook("FSB_Pantone_Coated")
# Type of book (C or U).... used for comparison purposes.
book_type = "C"

print(cb)
color_list = cb.get_color_name_list()

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

docSheet1 = workBookDocument.active
docSheet1.title = "%s" % cb


# Label column headings
docSheet1.cell(row=1, column=1).value = "Color"
docSheet1.cell(row=1, column=2).value = "L"
docSheet1.cell(row=1, column=3).value = "a"
docSheet1.cell(row=1, column=4).value = "b"

for i in range(len(color_list)):
    try:
        color = cb.get_color(color_list[i])
        print(color)
        # Increment rows, write charge data.
        # docSheet1.write(row, column, value)creation_date
        docSheet1.cell(row=i + 2, column=1).value = color_list[i]
        docSheet1.cell(row=i + 2, column=2).value = color.get_lab_color_obj().lab_l
        docSheet1.cell(row=i + 2, column=3).value = color.get_lab_color_obj().lab_a
        docSheet1.cell(row=i + 2, column=4).value = color.get_lab_color_obj().lab_b
        # Look up color standard and run comparison.
        try:
            db_color = ColorDefinition.objects.get(name=color_list[i], coating=book_type)
            db_color_lab = LabColor(lab_l=db_color.lab_l, lab_a=db_color.lab_a, lab_b=db_color.lab_b)
            # Calculate dE between color from Esko library and database library.
            de2000 = color.get_lab_color_obj().delta_e(db_color_lab)
            docSheet1.cell(row=i + 2, column=5).value = de2000
        except ColorDefinition.DoesNotExist:
            print("No database match found. Skipping dE.")

    except InvalidColor:
        print("Invalid Color. Skipping:", color_list[i])
    except Exception:
        print("Some Error. Skipping:", color_list[i])

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("%s.xls" % cb)

print("Exported.")
