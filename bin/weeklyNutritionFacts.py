#!/usr/bin/env python
"""Cleans up the Scans folder and deletes everything that is older than one month.

Should run once per day.
"""

# Setup the Django environment
import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from django.conf import settings
from django.contrib.auth.models import User
from django.core.mail import EmailMessage

from gchub_db.apps.workflow.models import ItemTracker

# all trackers that have tracker.item for new nutrition labels
trackers = ItemTracker.objects.filter(
    item__workflow__name="Beverage", type__category__name="Beverage Nutrition"
).order_by("item__job__customer_name")
print(len(trackers))

workBookDocument = openpyxl.Workbook()
docSheet1 = workBookDocument.active
docSheet1.title = "New Nutrition"

docSheet1.cell(row=1, column=1).value = "Job-Item"
docSheet1.cell(row=1, column=2).value = "Customer"
docSheet1.cell(row=1, column=3).value = "Size"
docSheet1.cell(row=1, column=4).value = "Item Designation"
docSheet1.cell(row=1, column=5).value = "Brand"
docSheet1.cell(row=1, column=6).value = "Description"

docSheet1.cell(row=1, column=7).value = "Completed"
docSheet1.cell(row=1, column=8).value = "Analyst"
docSheet1.cell(row=1, column=9).value = "Plate Code(s)"

counter = 0
for i in range(len(trackers)):
    if trackers[i].item.final_file_date():
        docSheet1.cell(row=counter + 2, column=1).value = (
            str(trackers[i].item.job.id) + "-" + str(trackers[i].item.num_in_job)
        )
        docSheet1.cell(row=counter + 2, column=2).value = str(
            trackers[i].item.job.customer_name
        )
        docSheet1.cell(row=counter + 2, column=3).value = str(
            trackers[i].item.size.size
        )
        docSheet1.cell(row=counter + 2, column=4).value = str(
            trackers[i].item.get_item_designation()
        )
        try:
            docSheet1.cell(row=counter + 2, column=5).value = (
                str(trackers[i].item.bev_brand_code.name)
                + " - "
                + str(trackers[i].item.bev_brand_code.code)
            )
        except Exception:
            docSheet1.cell(row=counter + 2, column=5).value = str(
                trackers[i].item.bev_brand_code
            )
        docSheet1.cell(row=counter + 2, column=6).value = str(
            trackers[i].item.description
        )

        if trackers[i].item.final_file_date():
            docSheet1.cell(row=counter + 2, column=7).value = str(
                trackers[i].item.final_file_date()
            )[0:10]
        else:
            docSheet1.cell(row=counter + 2, column=7).value = str("None")
        docSheet1.cell(row=counter + 2, column=8).value = str(
            trackers[i].item.job.salesperson
        )

        plate_codes = ""
        itemcolors = trackers[i].item.itemcolor_set.all()
        for itemcolor in itemcolors:
            if itemcolor.plate_code:
                plate_codes += itemcolor.plate_code + "/"
        docSheet1.cell(row=counter + 2, column=9).value = str(plate_codes)[0:-1]

        counter = counter + 1

docSheet1.panes_frozen = docSheet1["B2"]
workBookDocument.save("New_Nutrition_Items.xlsx")
print("Exported.")

filename = "New_Nutrition_Items.xlsx"
# Create an email message for attaching the invoice to.
mail_list = []
group_members = User.objects.filter(groups__name="EmailNewNutrition", is_active=True)
for user in group_members:
    mail_list.append(user.email)

email = EmailMessage(
    "New Nutrition Facts",
    "Here is the New Nutrition Facts report for this week.",
    settings.EMAIL_FROM_ADDRESS,
    mail_list,
)
# Attach the file.
email.attach_file(filename)
# Poof goes the mail.
email.send(fail_silently=False)

# Email the resulting document to Bobbi Wagner and Laura Rosenthal at Evergreen Packaging
# and Shelly Congdon in Beverage
