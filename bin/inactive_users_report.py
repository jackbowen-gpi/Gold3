#!/usr/bin/python
"""Create a spreadsheet of inactive users and email the report to admins."""

# openpyxl is now used instead of pyExcelerator because it supports
# python3 and pyexcelerator is deprecated
import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import timedelta

from django.conf import settings
from django.contrib.auth.models import User
from django.core.mail import EmailMessage
from django.utils import timezone

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

now = timezone.now()

# search criteria
thirty_days = now - timedelta(days=30)
sixty_days = now - timedelta(days=60)
ninety_days = now - timedelta(days=90)
never = None

"""
Add a new search query here and then to the tuple below to make a new page of users in the excel spreadsheet
We want the less than search here because we want people whose last login is older than X days. Greater than would
give us those within that X -> now time window.
"""
never_login = User.objects.filter(last_login=never, is_active=True).order_by("username")
ninety_days_login = User.objects.filter(
    last_login__lt=ninety_days, is_active=True
).order_by("username")

"""
Logging
"""
print("Never logged: %d" % len(never_login))
print("90 days logged: %d" % len(ninety_days_login))

"""
Add a new page label and new search criteria to this tuple to add new pages to the excel document with data
"""
tuple = (
    ("Never Login", never_login),
    ("Ninety Days", ninety_days_login),
)

sheetCounter = 0
for criteria in tuple:
    label, search_results = criteria
    if sheetCounter == 0:
        docSheet1 = workBookDocument.active
        docSheet1.title = "%s" % label
    else:
        # Create a new sheet for each plant.
        docSheet1 = workBookDocument.create_sheet("%s" % label)

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Username"
    docSheet1.cell(row=1, column=2).value = "Email"
    docSheet1.cell(row=1, column=3).value = "Last Login"

    sheetCounter = sheetCounter + 1

    if search_results:
        for i in range(len(search_results)):
            docSheet1.cell(row=i + 2, column=1).value = search_results[i].username
            docSheet1.cell(row=i + 2, column=2).value = search_results[i].email
            if search_results[i].last_login:
                docSheet1.cell(row=i + 2, column=3).value = search_results[
                    i
                ].last_login.strftime("%m/%d/%Y")
            else:
                docSheet1.cell(row=i + 2, column=3).value = "Never"

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("inactive_users_report.xlsx")

mail_list = []
cc_mail_list = []
group_members = User.objects.filter(groups__name="EmailGCHubManager", is_active=True)
for user in group_members:
    mail_list.append(user.email)
group_members = User.objects.filter(groups__name="IT Staff", is_active=True)
try:
    mail_list.append(group_members[0].email)
    cc_mail_list.append(group_members[1].email)
except Exception:
    for user in group_members:
        mail_list.append(user.email)

email = EmailMessage(
    "Inactive Users Report",
    "This is a report of inactive users in GOLD. Each page of the excel document represents "
    "a different time period starting with those who have never logged in.",
    settings.EMAIL_FROM_ADDRESS,
    mail_list,
    cc_mail_list,
)
# Attach the file and specify type.
email.attach_file("inactive_users_report.xls")
# Poof goes the mail.
email.send(fail_silently=False)

print("Exported.")
