#!/usr/bin/python
"""
Create volume trend reports across workflows and years.

Produces an Excel spreadsheet summarizing item counts per workflow.
"""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.workflow.models import Item

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

start_date = date(2000, 1, 1)
item_set = Item.objects.filter(
    creation_date__gte=start_date,
    job__prepress_supplier__in=("", None, "OPT", "Optihue"),
).exclude(job__status="Cancelled")

print("Total items:", item_set.count())

# Create a new sheet for each plant.
docSheet1 = workBookDocument.active
docSheet1.title = "Items In"
# Label column headings

docSheet1.cell(row=1, column=1).value = "Year"
docSheet1.cell(row=1, column=2).value = "FSB"
docSheet1.cell(row=1, column=3).value = "Bev"
docSheet1.cell(row=1, column=4).value = "Cont"

year = start_date.year
for x in range(11):
    docSheet1.cell(row=x + 2, column=1).value = year
    items = item_set.filter(creation_date__year=year, job__workflow__name="Foodservice")
    docSheet1.cell(row=x + 2, column=2).value = items.count()
    items = item_set.filter(creation_date__year=year, job__workflow__name="Beverage")
    docSheet1.cell(row=x + 2, column=3).value = items.count()
    items = item_set.filter(creation_date__year=year, job__workflow__name="Container")
    docSheet1.cell(row=x + 2, column=4).value = items.count()
    year += 1

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("xls_output/Hub_Volumes.xls")

print("Exported.")
