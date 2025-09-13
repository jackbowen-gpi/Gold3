"""On time results for end of the year."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from gchub_db.apps.workflow.models import Item, Job

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

# Year to run results on
year = 2010

jobs_active = (
    Job.objects.filter(due_date__year=year, workflow__name__in=("Beverage", "Foodservice"))
    .exclude(status__in=("Hold", "Cancelled", "Hold for Art"))
    .exclude(prepress_supplier__in=("PHT", "SGS", "SHK"))
)
print("Total Jobs %d" % jobs_active.count())

jobs_on_hold = Job.objects.filter(
    due_date__year=year,
    status__in=("Hold", "Hold for Art"),
    workflow__name__in=("Beverage", "Foodservice"),
)
print("Total Jobs on Hold %d" % jobs_on_hold.count())

jobs_on_cancelled = Job.objects.filter(
    due_date__year=year,
    status__in=("Cancelled"),
    workflow__name__in=("Beverage", "Foodservice"),
)
print("Total Jobs Cancelled %d" % jobs_on_cancelled.count())

items = Item.objects.filter(job__in=jobs_active).exclude(overdue_exempt=True).order_by("job__artist__username", "job__id")
print("Total Items %d" % items.count())

items_exempted = Item.objects.filter(job__in=jobs_active, overdue_exempt=True).order_by("job__artist__username", "job__id")

# Set up blank lists.
on_time = []
overdue = []
never_sent = []
filed_out_no_proof = []

# Sort eligable items to see how they are.
for i in items:
    proof = i.first_proof_date()
    if proof:
        if proof.date() <= i.job.real_due_date:
            on_time.append(i)
        else:
            # Assume items added after the due date were fine.
            if i.creation_date.date() > i.job.due_date:
                on_time.append(i)
            else:
                overdue.append(i)
    else:
        # Never proofed, never filed out, must have died :(
        if not i.final_file_date():
            never_sent.append(i)
        else:
            # Likely a press change or bad data from import.
            filed_out_no_proof.append(i)

print("Items On Time: %d" % len(on_time))
print("Items Overdue: %d" % len(overdue))
print("Items Not Proofed: %d" % len(never_sent))
print("Items Filed Out Not Proofed: %d" % len(filed_out_no_proof))

# First sheet provides item summary.
docSheet1 = workBookDocument.active
docSheet1.title = "%d Summary" % year

# Label column headings
docSheet1.cell(row=1, column=1).value = "Total Eligible Jobs"
docSheet1.cell(row=1, column=2).value = str(jobs_active.count())
docSheet1.cell(row=2, column=1).value = "Total Eligible Items"
docSheet1.cell(row=2, column=2).value = str(items.count())
docSheet1.cell(row=3, column=1).value = "Items on Time"
docSheet1.cell(row=3, column=2).value = str(len(on_time))
docSheet1.cell(row=4, column=1).value = "Items Overdue"
docSheet1.cell(row=4, column=2).value = str(len(overdue))
docSheet1.cell(row=5, column=1).value = "Items Never Proofed"
docSheet1.cell(row=5, column=2).value = str(len(never_sent))
docSheet1.cell(row=6, column=1).value = "Filed Out No Proof"
docSheet1.cell(row=6, column=2).value = str(len(filed_out_no_proof))

# Create 2nd sheet to display all items marked as overdue.
docSheet2 = workBookDocument.create_sheet("Overdue Items")

docSheet2.cell(row=1, column=1).value = "Item Creation"
docSheet2.cell(row=1, column=2).value = "Artist"
docSheet2.cell(row=1, column=3).value = "Job Num."
docSheet2.cell(row=1, column=4).value = "Job Name"
docSheet2.cell(row=1, column=5).value = "Size"
docSheet2.cell(row=1, column=6).value = "Num."
docSheet2.cell(row=1, column=7).value = "Proof Out Date"
docSheet2.cell(row=1, column=8).value = "Due Date"

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

for i in range(len(overdue)):
    # Increment rows, write data.
    # docSheet1.write(row, column, value)creation_date
    docSheet2.cell(row=i + 2, column=1).value = str(overdue[i].creation_date.date())
    try:
        docSheet2.cell(row=i + 2, column=2).value = str(overdue[i].job.artist.username)
    except Exception:
        docSheet2.cell(row=i + 2, column=2).value = "No Artist"
    docSheet2.cell(row=i + 2, column=3).value = str(overdue[i].job.id)
    docSheet2.cell(row=i + 2, column=4).value = str(overdue[i].job.name)
    docSheet2.cell(row=i + 2, column=5).value = str(overdue[i].size)
    docSheet2.cell(row=i + 2, column=6).value = str(overdue[i].num_in_job)
    docSheet2.cell(row=i + 2, column=7).value = str(overdue[i].first_proof_date().date())
    docSheet2.cell(row=i + 2, column=8).value = str(overdue[i].job.due_date)

# Create 3rd sheet to display all items exempted.
docSheet3 = workBookDocument.create_sheet("Exempted Items")

docSheet3.cell(row=1, column=1).value = "Item Creation"
docSheet3.cell(row=1, column=2).value = "Artist"
docSheet3.cell(row=1, column=3).value = "Job Num."
docSheet3.cell(row=1, column=4).value = "Job Name"
docSheet3.cell(row=1, column=5).value = "Size"
docSheet3.cell(row=1, column=6).value = "Num."
docSheet3.cell(row=1, column=7).value = "Proof Out Date"
docSheet3.cell(row=1, column=8).value = "Due Date"

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

for i in range(len(items_exempted)):
    # Increment rows, write data.
    # docSheet1.write(row, column, value)creation_date
    docSheet3.cell(row=i + 2, column=1).value = str(items_exempted[i].creation_date.date())
    try:
        docSheet3.cell(row=i + 2, column=2).value = str(items_exempted[i].job.artist.username)
    except Exception:
        docSheet3.cell(row=i + 2, column=2).value = "No Artist"
    docSheet3.cell(row=i + 2, column=3).value = str(items_exempted[i].job.id)
    docSheet3.cell(row=i + 2, column=4).value = str(items_exempted[i].job.name)
    docSheet3.cell(row=i + 2, column=5).value = str(items_exempted[i].size)
    docSheet3.cell(row=i + 2, column=6).value = str(items_exempted[i].num_in_job)
    if items_exempted[i].first_proof_date():
        docSheet3.cell(row=i + 2, column=7).value = str(items_exempted[i].first_proof_date().date())
    docSheet3.cell(row=i + 2, column=8).value = str(items_exempted[i].job.due_date)

# Create 3rd sheet to display all items exempted.
docSheet4 = workBookDocument.create_sheet("Items Never Proofed")

docSheet4.cell(row=1, column=1).value = "Item Creation"
docSheet4.cell(row=1, column=2).value = "Artist"
docSheet4.cell(row=1, column=3).value = "Job Num."
docSheet4.cell(row=1, column=4).value = "Job Name"
docSheet4.cell(row=1, column=5).value = "Size"
docSheet4.cell(row=1, column=6).value = "Num."
docSheet4.cell(row=1, column=7).value = "Proof Out Date"
docSheet4.cell(row=1, column=8).value = "Due Date"
docSheet4.cell(row=1, column=9).value = "Status"

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

for i in range(len(never_sent)):
    # Increment rows, write data.
    # docSheet1.write(row, column, value)creation_date
    docSheet4.cell(row=i + 2, column=1).value = str(never_sent[i].creation_date.date())
    try:
        docSheet4.cell(row=i + 2, column=2).value = str(never_sent[i].job.artist.username)
    except Exception:
        docSheet4.cell(row=i + 2, column=2).value = "No Artist"
    docSheet4.cell(row=i + 2, column=3).value = str(never_sent[i].job.id)
    docSheet4.cell(row=i + 2, column=4).value = str(never_sent[i].job.name)
    docSheet4.cell(row=i + 2, column=5).value = str(never_sent[i])
    docSheet4.cell(row=i + 2, column=6).value = str(never_sent[i].num_in_job)
    if never_sent[i].first_proof_date():
        docSheet4.cell(row=i + 2, column=7).value = str(never_sent[i].first_proof_date().date())
    docSheet4.cell(row=i + 2, column=8).value = str(never_sent[i].job.due_date)
    docSheet4.cell(row=i + 2, column=9).value = str(never_sent[i].job.status)


fsb_items = Item.objects.filter(job__workflow__name="Foodservice", job__creation_date__year=year)


# Save XLS document
workBookDocument.save("%dResults.xls" % year)
