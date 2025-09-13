#!/usr/bin/python
"""Export beverage volume data by plant into a spreadsheet."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.workflow.models import Item, ItemCatalog, Plant

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

workflow = "Beverage"

start_date = date(2010, 1, 1)
end_date = date(2011, 1, 30)
item_set = Item.objects.filter(
    job__workflow__name=workflow,
    creation_date__gte=start_date,
    creation_date__lte=end_date,
    job__prepress_supplier__in=("", None, "OPT", "Optihue"),
).exclude(job__status="Cancelled")

print("Total items:", item_set.count())

# Create a new sheet for each plant.
docSheet1 = workBookDocument.active
docSheet1.title = "Evergreen Volumes"

# Label column headings
docSheet1.cell(row=1, column=1).value = "Size"

evergreen_plants = Plant.objects.filter(workflow__name=workflow).order_by("name")
col_no = 1
for plant in evergreen_plants:
    docSheet1.write(0, col_no, plant.name)
    col_no += 1


bev_items = ItemCatalog.objects.filter(workflow__name=workflow).order_by("size")
row_no = 1
for product in bev_items:
    docSheet1.cell(row=row_no + 1, column=1).value = product.size
    col_no = 1
    for plant in evergreen_plants:
        num_items = item_set.filter(size=product, job__temp_printlocation__plant=plant).count()
        docSheet1.cell(row=row_no + 1, column=col_no + 1).value = num_items
        col_no += 1
    row_no += 1

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("xls_output/Evergreen_Volume_By_plant.xls")

print("Exported.")
