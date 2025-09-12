#!/usr/bin/python
"""
Generate weekly Foodservice billing spreadsheets and optionally send emails.

Produces Excel files summarizing billable charges for the Foodservice workflow.
"""

import calendar

import bin_functions

# openpyxl is now used instead of pyExcelerator because it supports
# python3 and pyexcelerator is deprecated
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import timedelta

from django.conf import settings
from django.contrib.auth.models import User
from django.core.mail import EmailMessage

from gchub_db.apps.budget import billing_funcs
from gchub_db.includes import general_funcs

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

# Set to False for initial run, then to True to update the billing charges
# as being invoiced.
# UPDATE_INVOICING = False

# Set your month and year here and only here: (as a number)
# To be passed as a variable eventually.
# FORCE_BILLING = 59390

# Set the timeframe for the billing cycle FORMAT = "5/10/2016"
# Searches should be from monday to monday with the thinking being that on the current monday morning
# no billing will be captured as it is too early for things to be filed out for that day, and the
# previous monday will be captured as everything should have been filed out after the last weekly
# billing was run.
today = general_funcs._utcnow_naive()
week = timedelta(days=7)
end = today - week

endDate = today.strftime("%m/%d/%Y")
startDate = end.strftime("%m/%d/%Y")

print(startDate)
print(endDate)

print("Begin Weekly Foodservice Invoicing.")

dateStartArr = startDate.split("/")
dateEndArr = endDate.split("/")
workflow = "Foodservice"

# Set the month name for use in the file name.
month_name = calendar.month_abbr[int(dateStartArr[0])]

# Get billable charge qset for workflow.
# Dates in format "5/10/2016"
billable_charges = billing_funcs.get_billable_timeframe(startDate, endDate, workflow)["charges"]

print(("Charges to invoice: %s" % str(billable_charges.count())))

# memphis cost center variables
memphisALL = "Memphis All 760279"
memphisFSB = "Memphis FSB 117409"

QADClark = "QAD Migration - Clarksville"
QADPitt = "QAD Migration - Pittston"

# Get list of plants to be invoiced
plants = []
items = []
for charge in billable_charges:
    # Add to the list of items for billing checks.
    if charge.item not in items:
        items.append(charge.item)

    try:
        if charge.item.printlocation.plant.name not in plants:
            # adding in memphis all and fsb tabs with the purchase center number, but only want to add them once
            if charge.item.printlocation.plant.name == "Memphis":
                if memphisALL not in plants:
                    plants.append(memphisALL)
                    plants.append(memphisFSB)
            elif "QAD/Avante" in charge.item.job.name:
                # Add plants for QAD/Avante changeover so they get their own tabs
                if charge.item.printlocation.plant.name == "Clarksville":
                    if QADClark not in plants:
                        plants.append(QADClark)
                elif charge.item.printlocation.plant.name == "Pittston":
                    if QADPitt not in plants:
                        plants.append(QADPitt)
            else:
                plants.append(charge.item.printlocation.plant.name)
    except AttributeError:
        # Use Other to catch items not assigned to a plant.
        if "Other" not in plants:
            plants.append("Other")

print("---> BEGIN ITEM BILLING WARNINGS.")

for item in items:
    information = item.num_in_job, item, item.job, item.job.artist
    if item.check_too_few_charges():
        print(("Warning: only one charge for item", information))
    if item.check_too_few_revision_charges():
        print(("Warning: missing revision charges for item", information))
    if item.check_fileout_post_production():
        print(("Warning: no post production charge for item", information))
    if item.check_prepress_charges():
        print(("Warning: no prepress charge for item", information))
    if item.check_color_keys():
        print(("Warning: no color keys for item", information))

plants.append("Avante-QAD")

sheetCounter = 0
for plant in plants:
    # Other is the category for items not assigned to a plant.
    if plant == "Other":
        charge_set = billable_charges.filter(item__printlocation__isnull=True)
    elif plant == "Avante-QAD":
        charge_set = billable_charges.filter(item__job__name__startswith="Letica - QAD/Avante")
    elif plant == memphisALL:
        charge_set = billable_charges.filter(
            item__printlocation__plant__name="Memphis",
            item__printlocation__press__name="All",
        )
    elif plant == memphisFSB:
        charge_set = billable_charges.filter(
            item__printlocation__plant__name="Memphis",
            item__printlocation__press__name="FSB",
        )
    elif plant == QADClark:
        charge_set = billable_charges.filter(
            item__job__name__contains="QAD/Avante",
            item__printlocation__plant__name="Clarksville",
            item__printlocation__press__name="Other",
        )
    elif plant == QADPitt:
        charge_set = billable_charges.filter(
            item__job__name__contains="QAD/Avante",
            item__printlocation__plant__name="Pittston",
            item__printlocation__press__name="Other",
        )
    else:
        charge_set = billable_charges.filter(item__printlocation__plant__name=plant).exclude(
            item__job__name__startswith="Letica - QAD/Avante"
        )

    if len(charge_set) == 0:
        continue

    # Some feedback as to how many charges per plant there are.
    print((plant, charge_set.count()))

    if sheetCounter == 0:
        docSheet1 = workBookDocument.active
        docSheet1.title = "%s Billing" % plant
    else:
        # Create a new sheet for each plant.
        docSheet1 = workBookDocument.create_sheet("%s Billing" % plant)

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Date Billed"
    docSheet1.cell(row=1, column=2).value = "Salesperson"
    docSheet1.cell(row=1, column=3).value = "Job No."
    docSheet1.cell(row=1, column=4).value = "Job Name"
    docSheet1.cell(row=1, column=5).value = "Size"
    docSheet1.cell(row=1, column=6).value = "Filed Out Date"
    docSheet1.cell(row=1, column=7).value = "Charge Type"
    docSheet1.cell(row=1, column=8).value = "Rush Days"
    docSheet1.cell(row=1, column=9).value = "Plant"
    docSheet1.cell(row=1, column=10).value = "Press"
    docSheet1.cell(row=1, column=11).value = "Amount"

    sheetCounter = sheetCounter + 1
    running_total = 0

    for i in range(len(charge_set)):
        # Increment rows, write charge data.
        # docSheet1.write(row, column, value)creation_date
        docSheet1.cell(row=i + 2, column=1).value = charge_set[i].creation_date.strftime("%m/%d/%y")
        try:
            docSheet1.cell(row=i + 2, column=2).value = str(charge_set[i].item.job.salesperson.username)
        except Exception:
            pass
        name = charge_set[i].item.job.name.encode("utf8", "replace")
        docSheet1.cell(row=i + 2, column=3).value = str(charge_set[i].item.job.id)
        docSheet1.cell(row=i + 2, column=4).value = str(name.decode("utf-8"))
        docSheet1.cell(row=i + 2, column=5).value = str(charge_set[i].item.size)

        docSheet1.cell(row=i + 2, column=6).value = charge_set[i].item.final_file_date().strftime("%m/%d/%y")
        docSheet1.cell(row=i + 2, column=7).value = str(charge_set[i].description)
        docSheet1.cell(row=i + 2, column=8).value = str(charge_set[i].rush_days)
        try:
            docSheet1.cell(row=i + 2, column=9).value = str(charge_set[i].item.printlocation.plant.name)
        except AttributeError:
            docSheet1.cell(row=i + 2, column=9).value = "----"
        try:
            docSheet1.cell(row=i + 2, column=10).value = str(charge_set[i].item.printlocation.press.name)
        except AttributeError:
            docSheet1.cell(row=i + 2, column=10).value = "----"
        docSheet1.cell(row=i + 2, column=11).value = charge_set[i].amount
        running_total += charge_set[i].amount

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    # Put the total at the bottom of the sheet.
    docSheet1.cell(row=len(charge_set) + 3, column=10).value = "Total:"
    docSheet1.cell(row=len(charge_set) + 3, column=11).value = running_total

# Save XLS document
workBookDocument.save("FSB_Billing_WEEKLY.xlsx")

print("Exported.")

filename = "FSB_Billing_WEEKLY.xlsx"
# Create an email message for attaching the invoice to.
mail_list = []
group_members = User.objects.filter(groups__name="EmailWeeklyBilling", is_active=True)
for user in group_members:
    mail_list.append(user.email)

email = EmailMessage(
    "Weekly Billing",
    "Here is the weekly billing for Clemson GCH Foodservice.",
    settings.EMAIL_FROM_ADDRESS,
    mail_list,
)
# Attach the file.
email.attach_file(filename)
# Poof goes the mail.
email.send(fail_silently=False)
