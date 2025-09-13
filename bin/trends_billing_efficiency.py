#!/usr/bin/python
"""
Generate billing efficiency and volume trend Excel reports.

Reports include yearly, monthly, and breakdowns by artist, salesperson,
quality, and plant for Foodservice workflow data.
"""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from django.contrib.auth.models import Permission, User
from django.db.models import Q, Sum

from gchub_db.apps.joblog import app_defs as joblog_defs
from gchub_db.apps.joblog.models import JobLog
from gchub_db.apps.workflow.app_defs import ITEM_TYPES
from gchub_db.apps.workflow.models import Charge, Item, Job

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

# Setup date ranges.
# end_year = datetime.today().year
end_year = 2010
print("End range", end_year)
year_set = (
    end_year,
    end_year - 1,
    end_year - 2,
    end_year - 3,
    end_year - 4,
    end_year - 5,
)
start_year = end_year - 5

start_date = date(start_year, 1, 1)
end_date = date(end_year, 12, 31)

# Define months
month_set = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)

# Define Foodservice permission.
FOODSERVICE_PERMISSION = Permission.objects.get(codename="foodservice_access")

# Define artist set:
ARTIST_PERMISSION = Permission.objects.get(codename="in_artist_pulldown")
artists = (
    User.objects.filter(groups__in=ARTIST_PERMISSION.group_set.all(), is_active=True)
    .filter(groups__in=FOODSERVICE_PERMISSION.group_set.all())
    .order_by("last_name")
)
# print "Foodservice Artists:", artists

# Definte salesperson set:
SALES_PERMISSION = Permission.objects.get(codename="salesperson")
salespeople = (
    User.objects.filter(groups__in=SALES_PERMISSION.group_set.all(), is_active=True)
    .filter(groups__in=FOODSERVICE_PERMISSION.group_set.all())
    .order_by("last_name")
)
# print "Foodservice Salespeople:", salespeople

# Retrieve actions for jobs, which will be used to sort items by progress.
file_outs = (
    JobLog.objects.filter(
        type=joblog_defs.JOBLOG_TYPE_ITEM_FILED_OUT,
        item__job__workflow__name="Foodservice",
    )
    .values("item__id")
    .query
)

proof_outs = (
    JobLog.objects.filter(
        type=joblog_defs.JOBLOG_TYPE_ITEM_PROOFED_OUT,
        item__job__workflow__name="Foodservice",
    )
    .values("item__id")
    .query
)

approved = (
    JobLog.objects.filter(
        type=joblog_defs.JOBLOG_TYPE_ITEM_APPROVED,
        item__job__workflow__name="Foodservice",
    )
    .values("item__id")
    .query
)

revised = (
    JobLog.objects.filter(
        type=joblog_defs.JOBLOG_TYPE_ITEM_REVISION,
        item__job__workflow__name="Foodservice",
    )
    .values("item__id")
    .query
)

revision_charges = Charge.objects.filter(description__category__name="Revision").values("item__id").query

# Some master data that needs to be established early.
# Heavy filtering here to obtain the most accurate customer-related data.
"""
master_job_list = Job.objects.filter(workflow__name="Foodservice",
                                     creation_date__range=(start_date, end_date))

"""
master_job_list = (
    Job.objects.filter(
        workflow__name="Foodservice",
        creation_date__range=(start_date, end_date),
        duplicated_from__isnull=True,
        plantpress_change=False,
    )
    .exclude(
        Q(name__icontains="Infinity")
        | Q(name__icontains="Photography")
        | Q(name__icontains="KD")
        | Q(name__icontains="Corrugate")
        | Q(name__icontains="Sell_Sheet")
        | Q(name__icontains="Sell Sheet")
        | Q(name__icontains="Video")
        | Q(name__icontains="Catalog")
        | Q(name__icontains="Soho")
        | Q(name__icontains="test")
        | Q(name__icontains="fingerprint")
        | Q(name__icontains="IPFSB")
        | Q(name__icontains="Its_Showtime")
        | Q(name__icontains="Milliken")
        | Q(name__icontains="Avanti")
        | Q(name__icontains="IP_")
    )
    .exclude(
        Q(salesperson__username="Sharon_Ault")
        | Q(salesperson__username="Debra_Gregg")
        | Q(salesperson__username="Beth_McKeithen")
        | Q(salesperson__username="Kristin_Newman")
        | Q(salesperson__username="Matthew_Alloway")
        | Q(salesperson__username="Donna_Storm")
        | Q(salesperson__username="Ian_Ruff")
        | Q(salesperson__username="Angela_Morris")
        | Q(salesperson__username="Eric_Wiley")
        | Q(salesperson__username="Todd_Fish")
        | Q(salesperson__username="Chad_Guckes")
        | Q(salesperson__username="Cathy_Routt")
        | Q(salesperson__username="Paul_Mamczur")
    )
    .order_by("id")
)

print("Master job count:", master_job_list.count())

master_item_list = Item.objects.filter(job__in=master_job_list)
print("Master item count:", master_item_list.count())

# Establish sets for all data, to be picked apart later.
all_filedout_items = master_item_list.filter(id__in=file_outs)
all_unfiledout_items = master_item_list.exclude(id__in=file_outs)


def _calculate_billing_data(item_set):
    """Return billing information for a given item set."""
    print("----Billing")
    item_ids = item_set.values("id").query
    # Calculate charges of billed items.
    charges = Charge.objects.filter(item__in=item_ids)
    if charges:
        total_charges = round(charges.aggregate(total=Sum("amount"))["total"], 2)
        billed_charges = charges.filter(invoice_date__isnull=False)
        total_billed_charges = round(billed_charges.aggregate(total=Sum("amount"))["total"], 2)
        unbilled_charges = round((total_charges - total_billed_charges), 2)
        unbilled_percentage = round((float(unbilled_charges / total_charges)) * 100, 2)
    else:
        total_charges = 0
        total_billed_charges = 0
        unbilled_charges = 0
        unbilled_percentage = 0

    billing_dict = {}
    billing_dict["total_charges"] = total_charges
    billing_dict["billed_charges"] = total_billed_charges
    billing_dict["unbilled_charges"] = unbilled_charges
    billing_dict["unbilled_percentage"] = unbilled_percentage
    return billing_dict


def _do_list_applicable_jobs():
    """Create sheet of all applicable jobs, so that they can be refined."""
    print("List Jobs")
    # Yearly Breakdown Worksheet
    docSheetJobs = workBookDocument.create_sheet("Applicable Jobs")
    # Label column headings
    i = 1
    for job in master_job_list:
        docSheetJobs.cell(row=i, column=1).value = str(job)
        i += 1


def _do_yearly_breakdown():
    """
    Iterate through the given years and fill in the information on docSheet1.

    TODO: Predictive billing for items that would be finaled out (Post-Production
    and Color Keys/Extra Proof).
    """
    print("Begin Yearly Breakdown")
    # Yearly Breakdown Worksheet
    docSheet1 = workBookDocument.create_sheet("Yearly")
    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Year"
    docSheet1.cell(row=1, column=2).value = "Total Jobs"
    docSheet1.cell(row=1, column=3).value = "Total Items"
    docSheet1.cell(row=1, column=4).value = "Items Filed Out"
    docSheet1.cell(row=1, column=5).value = "Hit Ratio"
    docSheet1.cell(row=1, column=6).value = "Charges Applied"
    docSheet1.cell(row=1, column=7).value = "Charges Billed"
    docSheet1.cell(row=1, column=8).value = "Charges Unbilled"
    docSheet1.cell(row=1, column=9).value = "Unbilled Portion"

    i = 1
    for year in year_set:
        # print year
        docSheet1.cell(row=i + 1, column=1).value = year

        # Get jobs & items for given year.
        # Do not exlcude Cancelled items, as we want to look at those, too.
        all_jobs = master_job_list.filter(creation_date__year=year)

        job_count = all_jobs.count()
        # print "No. Jobs:", job_count
        docSheet1.cell(row=i + 1, column=2).value = job_count

        all_items = Item.objects.filter(job__in=all_jobs)
        all_item_count = all_items.count()
        # print "No. Items:", all_item_count
        docSheet1.cell(row=i + 1, column=3).value = all_item_count

        # Determine which of the items in the given year were filed out.
        filed_out_items = all_items.filter(id__in=file_outs)
        filed_out_item_count = filed_out_items.count()
        # print "No. Filed Out:", filed_out_item_count
        docSheet1.cell(row=i + 1, column=4).value = filed_out_item_count

        # Calculate the hit ratio (ratio between all items and those filed out.)
        hit_ratio = round((float(filed_out_item_count) / float(all_item_count)) * 100, 2)
        # print "Hit Ratio:", hit_ratio
        docSheet1.cell(row=i + 1, column=5).value = hit_ratio

        # Calculate charges of billed items.
        all_charges = Charge.objects.filter(item__in=all_items)
        total_charges = round(all_charges.aggregate(total=Sum("amount"))["total"], 2)
        all_billed_charges = all_charges.filter(invoice_date__isnull=False)
        total_billed_charges = round(all_billed_charges.aggregate(total=Sum("amount"))["total"], 2)

        # All charges applied to items in the given year.
        # print "Charges applied: $", total_charges
        docSheet1.cell(row=i + 1, column=6).value = total_charges

        # All charges applied to items in the given year that were transferred to the plant.
        # print "Charges billed: $", total_billed_charges
        docSheet1.cell(row=i + 1, column=7).value = total_billed_charges

        unbilled_charges = round((total_charges - total_billed_charges), 2)
        unbilled_percentage = round((float(unbilled_charges / total_charges)) * 100, 2)
        # print "Unbilled charges: $", unbilled_charges, unbilled_percentage, "%"
        docSheet1.cell(row=i + 1, column=8).value = unbilled_charges
        docSheet1.cell(row=i + 1, column=9).value = unbilled_percentage

        # print "==========="
        i += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]


def _do_monthly_breakdown():
    """
    Do breakdown of hit ratios on a month-to-month basis.

    NOTE: This one takes a while to run. Look to speed up, or disable if needing
    a quicker report.
    """
    print("Begin Monthly Breakdown")
    # Monthly Breakdown
    docSheet2 = workBookDocument.create_sheet("Monthly")
    # Label column headings
    docSheet2.cell(row=1, column=1).value = "Month"
    docSheet2.cell(row=1, column=2).value = "Total Jobs"
    docSheet2.cell(row=1, column=3).value = "Total Items"
    docSheet2.cell(row=1, column=4).value = "Items Filed Out"
    docSheet2.cell(row=1, column=5).value = "Hit Ratio"
    docSheet2.cell(row=1, column=6).value = "Charges Applied"
    docSheet2.cell(row=1, column=7).value = "Charges Billed"
    docSheet2.cell(row=1, column=8).value = "Charges Unbilled"
    docSheet2.cell(row=1, column=9).value = "Unbilled Portion"

    i = 1
    for month in month_set:
        print("Month", month)
        docSheet2.cell(row=i + 1, column=1).value = str(month)
        month_jobs = master_job_list.filter(creation_date__month=month)
        month_jobs_count = month_jobs.count()
        docSheet2.cell(row=i + 1, column=2).value = month_jobs_count
        # Determine # of items in month.
        month_job_ids = month_jobs.values("id").query
        month_items = master_item_list.filter(job__id__in=month_job_ids)
        month_items_count = month_items.count()
        docSheet2.cell(row=i + 1, column=3).value = month_items_count
        # Determine # filed out.
        month_filed_items = month_items.filter(id__in=file_outs)
        month_filed_items_count = month_filed_items.count()
        docSheet2.cell(row=i + 1, column=4).value = month_filed_items_count
        # Calc hit ratio for month.
        month_hit_ratio = round((float(month_filed_items_count) / float(month_items_count)) * 100, 2)
        docSheet2.cell(row=i + 1, column=5).value = month_hit_ratio

        i += 1
    # Freeze the top row of column headings.
    docSheet2.panes_frozen = docSheet2["B2"]


def _do_artist_breakdown():
    """Iterate through artists and fill in docSheet3."""
    print("Begin Artist Breakdown")
    docSheet3 = workBookDocument.create_sheet("Artist")
    # Label column headings
    docSheet3.cell(row=1, column=1).value = "Artist"
    docSheet3.cell(row=1, column=2).value = "Total Items"
    docSheet3.cell(row=1, column=3).value = "Items Filed Out"
    docSheet3.cell(row=1, column=4).value = "Hit Ratio"
    docSheet3.cell(row=1, column=5).value = "Total Charged"
    docSheet3.cell(row=1, column=6).value = "Total Unbilled"
    docSheet3.cell(row=1, column=7).value = "Unbilled Percentage"
    i = 1
    for artist in artists:
        artist_items = master_item_list.filter(job__artist=artist)
        # If there are no items for the artist, skip to the next one.
        if not artist_items:
            continue
        docSheet3.write(i, 0, str(artist))
        # Filter by artist.
        artist_item_count = artist_items.count()
        docSheet3.write(i, 1, artist_item_count)
        # Filter artist by filed out.
        artist_filed_items_count = artist_items.filter(id__in=file_outs).count()
        docSheet3.write(i, 2, artist_filed_items_count)
        # Calc hit ratio.
        if artist_item_count > 0:
            artist_hit_ratio = round((float(artist_filed_items_count) / float(artist_item_count)) * 100, 2)
        else:
            artist_hit_ratio = 0
        docSheet3.write(i, 3, artist_hit_ratio)
        billing = _calculate_billing_data(artist_items)
        docSheet3.write(i, 4, billing["total_charges"])
        docSheet3.write(i, 5, billing["unbilled_charges"])
        docSheet3.write(i, 6, billing["unbilled_percentage"])
        i += 1

    # Freeze the top row of column headings.
    docSheet3.panes_frozen = docSheet3["B2"]


def _do_salesperson_breakdown():
    """Iterate through salespeople and fill in docSheet4."""
    print("Begin Sales Breakdown")
    docSheet4 = workBookDocument.create_sheet("Salesperson")
    # Label column headings
    docSheet4.cell(row=1, column=1).value = "Salesperson"
    docSheet4.cell(row=1, column=2).value = "Total Items"
    docSheet4.cell(row=1, column=3).value = "Items Filed Out"
    docSheet4.cell(row=1, column=4).value = "Hit Ratio"
    docSheet4.cell(row=1, column=5).value = "Total Charged"
    docSheet4.cell(row=1, column=6).value = "Total Unbilled"
    docSheet4.cell(row=1, column=7).value = "Unbilled Percentage"

    i = 1
    for salesperson in salespeople:
        salesperson_items = master_item_list.filter(job__salesperson=salesperson)
        if not salesperson_items:
            continue
        docSheet4.write(i, 0, str(salesperson))
        # Filter by salesperson.
        salesperson_item_count = salesperson_items.count()
        docSheet4.write(i, 1, salesperson_item_count)
        # Filter salesperson by filed out.
        salesperson_filed_items_count = salesperson_items.filter(id__in=file_outs).count()
        docSheet4.write(i, 2, salesperson_filed_items_count)
        # Calc hit ratio.
        if salesperson_item_count > 0:
            salesperson_hit_ratio = round(
                (float(salesperson_filed_items_count) / float(salesperson_item_count)) * 100,
                2,
            )
        else:
            salesperson_hit_ratio = 0
        docSheet4.write(i, 3, salesperson_hit_ratio)
        billing = _calculate_billing_data(salesperson_items)
        docSheet4.write(i, 4, billing["total_charges"])
        docSheet4.write(i, 5, billing["unbilled_charges"])
        docSheet4.write(i, 6, billing["unbilled_percentage"])
        i += 1

    # Freeze the top row of column headings.
    docSheet4.panes_frozen = docSheet4["B2"]


def do_progress_breakdown():
    """
    Look at progress of all items, determine hit ratio for each stage.

    Proofed, Approved, Revisions, etc...
    """
    print("Begin Progress Breakdown")
    docSheet5 = workBookDocument.create_sheet("Timeline Progress")
    # Label column headings
    docSheet5.cell(row=1, column=1).value = "Stage"
    docSheet5.cell(row=1, column=2).value = "Items"
    docSheet5.cell(row=1, column=3).value = "Filed Out Items"
    docSheet5.cell(row=1, column=4).value = "Hit Ratio"
    docSheet5.cell(row=1, column=5).value = "Total Charged"
    docSheet5.cell(row=1, column=6).value = "Total Unbilled"
    docSheet5.cell(row=1, column=7).value = "Unbilled Percentage"

    # Totals for all items in data set.
    docSheet5.write(1, 0, "Totals")
    print("--Totals")
    # Count
    all_item_count = master_item_list.count()
    docSheet5.write(1, 1, all_item_count)
    # Filed out count
    all_filed_item_count = all_filedout_items.count()
    docSheet5.write(1, 2, all_filed_item_count)
    # Calc hit ratio
    total_hit_ratio = round((float(all_filed_item_count) / float(all_item_count)) * 100, 2)
    docSheet5.write(1, 3, total_hit_ratio)
    billing = _calculate_billing_data(master_item_list)
    docSheet5.write(1, 4, billing["total_charges"])
    docSheet5.write(1, 5, billing["unbilled_charges"])
    docSheet5.write(1, 6, billing["unbilled_percentage"])

    # Totals for proofed items in data set.
    docSheet5.write(2, 0, "Proofed")
    print("--Proofed")
    # Count
    proofed_items = master_item_list.filter(id__in=proof_outs)
    proofed_item_count = proofed_items.count()
    print("Proofed items:", proofed_item_count)
    docSheet5.write(2, 1, proofed_item_count)
    # Proof and filed out count
    filed_proofed_items = proofed_items.filter(id__in=file_outs)
    proofed_filed_items_count = filed_proofed_items.count()
    print("Proofed and filed out items:", proofed_filed_items_count)
    docSheet5.write(2, 2, proofed_filed_items_count)
    # Calc hit ratio
    total_hit_ratio = round((float(proofed_filed_items_count) / float(proofed_item_count)) * 100, 2)
    docSheet5.write(2, 3, total_hit_ratio)
    billing = _calculate_billing_data(proofed_items)
    docSheet5.write(2, 4, billing["total_charges"])
    docSheet5.write(2, 5, billing["unbilled_charges"])
    docSheet5.write(2, 6, billing["unbilled_percentage"])

    # Items with revisions/multiple proof outs.
    docSheet5.write(3, 0, "Revised")
    print("--Revised")
    # NOTE: Items imported into GOLD2 may not have any Revisions attached. They
    # would however have multiple proofs.
    # Count
    revised_items = master_item_list.filter(id__in=revision_charges)
    revised_item_count = revised_items.count()
    docSheet5.write(3, 1, revised_item_count)
    # Filed out count
    revised_filed_items = revised_items.filter(id__in=file_outs)
    revised_filed_item_count = revised_filed_items.count()
    docSheet5.write(3, 2, revised_filed_item_count)
    # Calc hit ratio
    revised_hit_ratio = round((float(revised_filed_item_count) / float(revised_item_count)) * 100, 2)
    docSheet5.write(3, 3, revised_hit_ratio)
    billing = _calculate_billing_data(revised_items)
    docSheet5.write(3, 4, billing["total_charges"])
    docSheet5.write(3, 5, billing["unbilled_charges"])
    docSheet5.write(3, 6, billing["unbilled_percentage"])

    # Totals for approved items in data set.
    docSheet5.write(4, 0, "Approved")
    print("--Approved")
    # Count
    approved_items = master_item_list.filter(id__in=approved)
    approved_item_count = approved_items.count()
    docSheet5.write(4, 1, approved_item_count)
    # Filed out count
    approved_filed_items = approved_items.filter(id__in=file_outs)
    approved_filed_item_count = approved_filed_items.count()
    docSheet5.write(4, 2, approved_filed_item_count)
    # Calc hit ratio
    total_hit_ratio = round((float(approved_filed_item_count) / float(approved_item_count)) * 100, 2)
    docSheet5.write(4, 3, total_hit_ratio)
    billing = _calculate_billing_data(approved_items)
    docSheet5.write(4, 4, billing["total_charges"])
    docSheet5.write(4, 5, billing["unbilled_charges"])
    docSheet5.write(4, 6, billing["unbilled_percentage"])

    # Freeze the top row of column headings.
    docSheet5.panes_frozen = docSheet5["B2"]


def _do_general_stats():
    """Just some general statistics from the data."""
    print("Begin General Statistics")
    docSheet6 = workBookDocument.create_sheet("General Stats")
    # Label column headings
    docSheet6.cell(row=2, column=1).value = "Total Items"
    all_item_count = master_item_list.count()
    docSheet6.cell(row=2, column=2).value = all_item_count
    docSheet6.cell(row=1, column=3).value = "Percentage of Unfiled"
    # Unfiled out items
    docSheet6.cell(row=3, column=1).value = "Unfiled Items"
    all_unfiledout_items_count = all_unfiledout_items.count()
    docSheet6.cell(row=3, column=2).value = all_unfiledout_items_count
    # Cancelled items
    docSheet6.cell(row=4, column=1).value = "Canceled Items"
    canceled_items_count = all_unfiledout_items.filter(job__status="Cancelled").count()
    docSheet6.cell(row=4, column=2).value = canceled_items_count
    docSheet6.cell(row=4, column=3).value = round((float(canceled_items_count) / float(all_unfiledout_items_count)) * 100, 2)
    # Hold Items
    docSheet6.cell(row=5, column=1).value = "Hold Items"
    hold_items_count = all_unfiledout_items.filter(job__status="Hold").count()
    docSheet6.cell(row=5, column=2).value = hold_items_count
    docSheet6.cell(row=5, column=3).value = round((float(hold_items_count) / float(all_unfiledout_items_count)) * 100, 2)
    # Before and after proof reminder email.
    docSheet6.cell(row=6, column=1).value = "Pre-Reminder Email"
    pre_email_items = master_item_list.filter(creation_date__lte=date(2009, 7, 1))
    pre_email_items_count = pre_email_items.count()
    pre_email_items_filedout = pre_email_items.filter(id__in=all_filedout_items)
    pre_email_items_filedout_count = pre_email_items_filedout.count()
    docSheet6.cell(row=6, column=2).value = pre_email_items_count
    docSheet6.cell(row=6, column=3).value = round((float(pre_email_items_filedout_count) / float(pre_email_items_count)) * 100, 2)
    # Before and after proof reminder email.
    docSheet6.cell(row=7, column=1).value = "Post-Reminder Email"
    post_email_items = master_item_list.filter(creation_date__gte=date(2009, 7, 1))
    post_email_items_count = post_email_items.count()
    post_email_items_filedout = post_email_items.filter(id__in=all_filedout_items)
    post_email_items_filedout_count = post_email_items_filedout.count()
    docSheet6.cell(row=7, column=2).value = post_email_items_count
    docSheet6.cell(row=7, column=3).value = round(
        (float(post_email_items_filedout_count) / float(post_email_items_count)) * 100,
        2,
    )


def _do_quality_breakdown():
    """
    Determine hit ratios based on item quality.

    TODO: and number of colors?
    """
    print("Begin Quality Breakdown")
    docSheet7 = workBookDocument.create_sheet("Quality")
    # Label column headings
    docSheet7.cell(row=1, column=1).value = "Quality"
    docSheet7.cell(row=1, column=2).value = "Total Items"
    docSheet7.cell(row=1, column=3).value = "Filed Out Items"
    docSheet7.cell(row=1, column=4).value = "Hit Ratio"
    docSheet7.cell(row=1, column=5).value = "Total Charged"
    docSheet7.cell(row=1, column=6).value = "Total Unbilled"
    docSheet7.cell(row=1, column=7).value = "Unbilled Percentage"

    quality_set = ("A", "B", "C")

    i = 1
    for year in year_set:
        docSheet7.write(i, 0, year)
        for qual in quality_set:
            docSheet7.write(i, 7, qual)
            # Total of given quality.
            quality_items = master_item_list.filter(quality=qual, creation_date__year=year)
            quality_items_count = quality_items.count()
            docSheet7.write(i, 1, quality_items_count)
            # Num. filed out
            quality_items_filedout = quality_items.filter(id__in=file_outs)
            quality_items_filedout_count = quality_items_filedout.count()
            docSheet7.write(i, 2, quality_items_filedout_count)
            # Calc. hit ratio.
            quality_ratio = round(
                (float(quality_items_filedout_count) / float(quality_items_count)) * 100,
                2,
            )
            docSheet7.write(i, 3, quality_ratio)
            billing = _calculate_billing_data(quality_items)
            docSheet7.write(i, 4, billing["total_charges"])
            docSheet7.write(i, 5, billing["unbilled_charges"])
            docSheet7.write(i, 6, billing["unbilled_percentage"])
            i += 1


def _do_plant_breakdown():
    """Filter total item list by plant, then look at filed vs no filed out."""
    print("Begin Plant Breakdown")
    docSheet8 = workBookDocument.create_sheet("Plant")
    # Label column headings
    docSheet8.cell(row=1, column=1).value = "Plant"
    docSheet8.cell(row=1, column=2).value = "Quality"
    docSheet8.cell(row=1, column=3).value = "Total Items"
    docSheet8.cell(row=1, column=4).value = "Filed Out Items"
    docSheet8.cell(row=1, column=5).value = "Hit Ratio"
    docSheet8.cell(row=1, column=6).value = "Total Charged"
    docSheet8.cell(row=1, column=7).value = "Total Unbilled"
    docSheet8.cell(row=1, column=8).value = "Unbilled Percentage"

    plant_set = [
        "Kenton",
        "Shelbyville",
        "Visalia",
    ]

    quality_set = [
        "A",
        "B",
        "C",
    ]

    i = 1
    for plant in plant_set:
        docSheet8.write(i, 0, plant)

        for quality in quality_set:
            docSheet8.write(i, 7, quality)
            # Total of given plant.
            plant_items = master_item_list.filter(printlocation__plant__name=plant, quality=quality)
            plant_items_count = plant_items.count()
            docSheet8.write(i, 1, plant_items_count)
            # Num. filed out
            plant_items_filedout = plant_items.filter(id__in=file_outs)
            plant_items_filedout_count = plant_items_filedout.count()
            docSheet8.write(i, 2, plant_items_filedout_count)
            # Calc. hit ratio.
            plant_ratio = round((float(plant_items_filedout_count) / float(plant_items_count)) * 100, 2)
            docSheet8.write(i, 3, plant_ratio)
            billing = _calculate_billing_data(plant_items)
            docSheet8.write(i, 4, billing["total_charges"])
            docSheet8.write(i, 5, billing["unbilled_charges"])
            docSheet8.write(i, 6, billing["unbilled_percentage"])
            i += 1


def _do_itemtype_breakdown():
    """Analyze by substrate type."""
    print("Begin Item Type Breakdown")
    docSheet9 = workBookDocument.create_sheet("Item Type")
    # Label column headings
    docSheet9.cell(row=1, column=1).value = "Item Type"
    docSheet9.cell(row=1, column=2).value = "Total Items"
    docSheet9.cell(row=1, column=3).value = "Filed Out Items"
    docSheet9.cell(row=1, column=4).value = "Hit Ratio"
    docSheet9.cell(row=1, column=5).value = "Total Charged"
    docSheet9.cell(row=1, column=6).value = "Total Unbilled"
    docSheet9.cell(row=1, column=7).value = "Unbilled Percentage"

    type_set = ITEM_TYPES

    i = 1
    for item_type in type_set:
        type = item_type[0]
        docSheet9.write(i, 0, item_type[1])
        type_items = master_item_list.filter(size__item_type=type)
        type_items_count = type_items.count()
        docSheet9.write(i, 1, type_items_count)
        # Num. filed out
        type_items_filedout = type_items.filter(id__in=file_outs)
        type_items_filedout_count = type_items_filedout.count()
        docSheet9.write(i, 2, type_items_filedout_count)
        # Calc. hit ratio.
        if type_items_count > 0:
            type_ratio = round((float(type_items_filedout_count) / float(type_items_count)) * 100, 2)
        else:
            type_ratio = 0
        docSheet9.write(i, 3, type_ratio)
        billing = _calculate_billing_data(type_items)
        docSheet9.write(i, 4, billing["total_charges"])
        docSheet9.write(i, 5, billing["unbilled_charges"])
        docSheet9.write(i, 6, billing["unbilled_percentage"])
        i += 1


def _do_monthly_itemtype():
    """Monthly volume by item type"""
    print("Begin Monthly Type Breakdown")
    # Monthly Breakdown
    docSheet10 = workBookDocument.create_sheet("Monthly Type")
    # Label column headings
    docSheet10.cell(row=1, column=1).value = "Month"
    docSheet10.cell(row=1, column=2).value = "Hot Cup"
    docSheet10.cell(row=1, column=3).value = "Cold Cup"

    i = 1
    for month in month_set:
        print("Month", month)
        docSheet10.write(i, 0, str(month))
        month_items = master_item_list.filter(job__creation_date__month=month, size__item_type="HC")
        month_items_count = month_items.count()
        docSheet10.write(i, 1, month_items_count)
        month_items = master_item_list.filter(job__creation_date__month=month, size__item_type="CC")
        month_items_count = month_items.count()
        docSheet10.write(i, 2, month_items_count)

        i += 1

    # Freeze the top row of column headings.
    docSheet10.panes_frozen = docSheet10["B2"]


def _do_case_estimates():
    """Calculate total number of cases involved for items that did not file out."""
    case_total = 0
    # Garbage_items are items with case counts of < 10
    garbage_items = 0
    case_set = all_unfiledout_items.filter(id__in=approved).exclude(annual_use=None).order_by("-annual_use")

    for item in case_set:
        if item.annual_use < 10:
            garbage_items += 1
        else:
            case_total += item.annual_use

    print("Applied Items: ", case_set.count())
    print(case_set[0].job)
    print("All Unfinished Items: ", all_unfiledout_items.count())
    print("Cases: ", case_total)
    print("Garbage Items: ", garbage_items)


def _do_progress_averages():
    """Calculate average times between creation and approval, file out."""
    """
    # Time to file out.
    filed_out_count = all_filedout_items.count()
    total_days = 0
    for item in all_filedout_items:
        elapsed_fileout = item.final_file_date().date() - item.creation_date.date()
        total_days += elapsed_fileout.days

    average_elapsed_fileout = total_days / filed_out_count
    print("Average Time to File Out: ", average_elapsed_fileout)

    # Time to approval.
    approved_items = master_item_list.filter(id__in=approved)
    approved_items_count = approved_items.count()
    total_days = 0
    for item in approved_items:
        elapsed_approval = item.approval_date().date() - item.creation_date.date()
        total_days += elapsed_approval.days

    average_elapsed_approval = total_days / approved_items_count
    print("Average Time to Approval: ", average_elapsed_approval)
    """
    # Time to approval for non-filed out items.
    approved_items_incomplete = all_unfiledout_items.filter(id__in=approved)
    approved_items_incomplete_count = approved_items_incomplete.count()
    total_days = 0
    for item in approved_items_incomplete:
        elapsed_approval = item.approval_date().date() - item.creation_date.date()
        total_days += elapsed_approval.days

    average_elapsed_approval = total_days / approved_items_incomplete_count
    print("Num. Items approved, not filed out: ", approved_items_incomplete_count)
    print("Average Time to Approval, Not Filed Out: ", average_elapsed_approval)


def _do_approval_time_chances():
    """Calculate chances of file out based on time to approval."""
    approved_items = master_item_list.filter(id__in=approved)
    # 0-10
    set1 = []
    # 11-20
    set2 = []
    # 21-30
    set3 = []
    # 31-40
    set4 = []
    # 41-50
    set5 = []
    # 51-60
    set6 = []
    # 60+
    set7 = []
    for item in approved_items:
        elapsed_approval = (item.approval_date().date() - item.creation_date.date()).days
        if elapsed_approval <= 10:
            set1.append(item)
        elif 11 <= elapsed_approval <= 20:
            set2.append(item)
        elif 21 <= elapsed_approval <= 30:
            set3.append(item)
        elif 31 <= elapsed_approval <= 40:
            set4.append(item)
        elif 41 <= elapsed_approval <= 50:
            set5.append(item)
        elif 51 <= elapsed_approval <= 60:
            set6.append(item)
        else:
            set7.append(item)

    print("1 :", len(set1))
    print("2 :", len(set2))
    print("3 :", len(set3))
    print("4 :", len(set4))
    print("5 :", len(set5))
    print("6 :", len(set6))
    print("7 :", len(set7))

    ffoset1 = 0
    ffoset2 = 0
    ffoset3 = 0
    ffoset4 = 0
    ffoset5 = 0
    ffoset6 = 0
    ffoset7 = 0

    for item in set1:
        if item.is_filed_out():
            ffoset1 += 1
    for item in set2:
        if item.is_filed_out():
            ffoset2 += 1
    for item in set3:
        if item.is_filed_out():
            ffoset3 += 1
    for item in set4:
        if item.is_filed_out():
            ffoset4 += 1
    for item in set5:
        if item.is_filed_out():
            ffoset5 += 1
    for item in set6:
        if item.is_filed_out():
            ffoset6 += 1
    for item in set7:
        if item.is_filed_out():
            ffoset7 += 1

    print("1 :", ffoset1)
    print("2 :", ffoset2)
    print("3 :", ffoset3)
    print("4 :", ffoset4)
    print("5 :", ffoset5)
    print("6 :", ffoset6)
    print("7 :", ffoset7)


def _do_approved_not_filed_detail():
    """Details on jobs where items have been approved, but not filed out."""
    approved_items_incomplete = all_unfiledout_items.filter(id__in=approved)
    print("Begin Approval Detail")
    # Monthly Breakdown
    docSheet12 = workBookDocument.create_sheet("Approval Detail")
    # Label column headings
    docSheet12.cell(row=1, column=1).value = "Job"
    docSheet12.cell(row=1, column=2).value = "Item"
    docSheet12.cell(row=1, column=3).value = "Approval Date"
    docSheet12.cell(row=1, column=4).value = "Status"
    docSheet12.cell(row=1, column=5).value = "Salesperson"
    docSheet12.cell(row=1, column=6).value = "Any Items Complete?"

    jobs_incomplete = []

    i = 1
    for item in approved_items_incomplete:
        docSheet12.write(i, 0, str(item.job))
        docSheet12.write(i, 1, str(item))
        docSheet12.write(i, 2, str(item.approval_date()))
        docSheet12.write(i, 3, str(item.job.status))
        docSheet12.write(i, 4, str(item.job.salesperson))
        other_items_complete = False
        for x in Item.objects.filter(job=item.job):
            if x.is_filed_out():
                other_items_complete = True
        if other_items_complete:
            docSheet12.write(i, 5, "Yes")
        else:
            if item.job not in jobs_incomplete:
                jobs_incomplete.append(item.job)
        i += 1

    jobs_cancelled = 0
    for job in jobs_incomplete:
        if job.status == "Cancelled":
            jobs_cancelled += 1

    print("Num. jobs cancelled:", jobs_cancelled)

    print("Number of jobs with approved items, no filed out items:", len(jobs_incomplete))


# Execute each breakdown.
# _do_list_applicable_jobs()
_do_yearly_breakdown()
# Monthly takes fooorr-ev-eerrr.
# _do_monthly_breakdown()
# _do_artist_breakdown()
# _do_salesperson_breakdown()
# Progress not working at the moment. Stalls on Proofed item report. Weird.
# do_progress_breakdown()
# _do_general_stats()
_do_quality_breakdown()
# _do_plant_breakdown()
# _do_itemtype_breakdown()
# _do_monthly_itemtype()
# _do_case_estimates()
# _do_progress_averages()
# _do_approval_time_chances()
# _do_approved_not_filed_detail()

# Save XLS document
workBookDocument.save("xls_output/TrendData.xls")

print("Exported.")
