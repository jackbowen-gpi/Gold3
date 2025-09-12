#!/usr/bin/python
"""Export Ecotainer items and their color usage to a spreadsheet."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from gchub_db.apps.workflow.models import Item

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

# Grab all ecotainer items.
item_set = (
    Item.objects.filter(job__workflow__name="Foodservice", size__size__icontains="e-", is_deleted=False)
    .exclude(size__size__istartswith="b")
    .exclude(size__size__iendswith="kd")
)

print("Total items:", item_set.count())

# Create a new sheet for each plant.
docSheet1 = workBookDocument.active
docSheet1.title = "Ecotainer"

# Label column headings
docSheet1.cell(row=1, column=1).value = "Job"
docSheet1.cell(row=1, column=2).value = "Item"
docSheet1.cell(row=1, column=3).value = "Part#"
docSheet1.cell(row=1, column=4).value = "Color Use"

all_colors = []

for i in range(len(item_set)):
    if item_set[i].is_filed_out():
        docSheet1.cell(row=i + 2, column=1).value = str(item_set[i].job)
        docSheet1.cell(row=i + 2, column=2).value = str(item_set[i].size)
        docSheet1.cell(row=i + 2, column=3).value = str(item_set[i].fsb_nine_digit)
        colors_used = item_set[i].itemcolor_set.all()
        colorset = []
        for color in colors_used:
            name = color.color
            if color.definition:
                name += color.definition.coating
            else:
                if not name.endswith(("U", "C")):
                    if color.item.size.size.startswith(
                        (
                            "S",
                            "s",
                            "r",
                            "R",
                        )
                    ):
                        name += " U"
                    else:
                        name += " C"
            colorset.append(name)
            # Add to master list.
            if name not in all_colors:
                all_colors.append(name)
        docSheet1.cell(row=i + 2, column=4).value = str(colorset)
    else:
        # Item not filed out, don't need to include it.
        pass

docSheet2 = workBookDocument.create_sheet("Color Usage")

# Label column headings
docSheet2.cell(row=1, column=1).value = "Colors"
for i in range(len(all_colors)):
    docSheet2.cell(row=i + 2, column=1).value = all_colors[i]

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("xls_output/Ecotainer_items.xls")

print("Exported.")
