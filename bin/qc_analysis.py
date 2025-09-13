#!/usr/bin/python
"""
Generate QC analysis spreadsheets for artist error tracking.

Produces an Excel file summarizing QC catches by artist and category.
"""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from django.contrib.auth.models import Permission, User

from gchub_db.apps.qc.models import QCCategory, QCWhoops

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

# Define artist set:
ARTIST_PERMISSION = Permission.objects.get(codename="in_artist_pulldown")
artists = User.objects.filter(groups__in=ARTIST_PERMISSION.group_set.all(), is_active=True).order_by("last_name")


def _do_qc_breakdown():
    """
    Look at catches caught on QCs and break down by artists.

    Populates an Excel worksheet with counts of errors made and caught
    per artist and category.
    """
    print("Begin QC Breakdown")
    # QC Breakdown Worksheet

    docSheet1 = workBookDocument.active
    docSheet1.title = "QC Catches"
    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Artist"
    docSheet1.cell(row=1, column=2).value = "Errors Made"
    docSheet1.cell(row=1, column=3).value = "Errors Caught"
    n = 4
    qc_category = QCCategory.objects.all()
    for cat in qc_category:
        docSheet1.cell(row=1, column=n).value = "Made: %s" % str(cat)
        n += 1
    for cat in qc_category:
        docSheet1.cell(row=1, column=n).value = "Caught: %s" % str(cat)
        n += 1

    i = 1
    for artist in artists:
        all_made = QCWhoops.objects.filter(qc_response__qcdoc__job__artist=artist)
        all_caught = QCWhoops.objects.filter(qc_response__qcdoc__reviewer=artist)
        if all_made and all_caught:
            docSheet1.cell(row=i + 1, column=1).value = str(artist)
            docSheet1.cell(row=i + 1, column=2).value = all_made.count()
            docSheet1.cell(row=i + 1, column=3).value = all_caught.count()
            n = 3
            for cat in qc_category:
                docSheet1.cell(row=i + 1, column=n + 1).value = all_made.filter(qc_response__category=cat).count()
                n += 1
            for cat in qc_category:
                docSheet1.cell(row=i + 1, column=n + 1).value = all_caught.filter(qc_response__category=cat).count()
                n += 1
            i += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]


# Execute each breakdown.
_do_qc_breakdown()

# Save XLS document
workBookDocument.save("xls_output/QC_Analysis.xls")

print("Exported.")
