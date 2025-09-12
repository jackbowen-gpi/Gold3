#!/usr/bin/python
"""
Produce lead time reports for salespeople in Foodservice workflow.

Exports an XLS with average lead times per salesperson.
"""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from django.contrib.auth.models import Permission, User
from django.db.models import Q

from gchub_db.apps.workflow.models import Job

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

workflow = "Foodservice"

start_date = date(2007, 1, 1)
job_set = Job.objects.filter(creation_date__gte=start_date, workflow__name=workflow).exclude(status="Cancelled")

print("Total jobs:", job_set.count())

SALES_PERMISSION = Permission.objects.get(codename="salesperson")
sales = User.objects.filter(groups__in=SALES_PERMISSION.group_set.all(), is_active=True)

workflow_permission = Permission.objects.get(codename="foodservice_access")
qset_fsb = sales.filter(Q(groups__in=workflow_permission.group_set.all())).values("id").query

fsb_sales = User.objects.filter(Q(id__in=qset_fsb)).order_by("username")

# Create a new sheet for each plant.

docSheet1 = workBookDocument.active
docSheet1.title = "Sales Lead Times"

# Label column headings
docSheet1.cell(row=1, column=1).value = "Salesperson"
docSheet1.cell(row=1, column=2).value = "Number of Jobs"
docSheet1.cell(row=1, column=3).value = "Avg Lead Time"
docSheet1.cell(row=1, column=4).value = "Jobs Under 2 Day Turn"

i = 1
for salesperson in fsb_sales:
    # Increment rows, write data.
    # docSheet1.write(row, column, value)
    docSheet1.cell(row=i + 1, column=1).value = salesperson.username
    # No. of jobs for salesperson.
    sales_jobs = job_set.filter(salesperson=salesperson)
    docSheet1.cell(row=i + 1, column=2).value = sales_jobs.count()
    # Calculate lead times.
    total_lead_time = 0
    total_sub_one_days = 0
    jobs_applied = 0
    if sales_jobs.count() != 0:
        for job in sales_jobs:
            lead_time = float((job.due_date - job.creation_date.date()).days)
            if not lead_time < -1 and not lead_time > 20:
                total_lead_time += lead_time
                jobs_applied += 1
                if lead_time <= 2 and lead_time > -5:
                    total_sub_one_days += 1
        avg_lead_time = total_lead_time / jobs_applied
    else:
        avg_lead_time = "N/A"

    docSheet1.cell(row=i + 1, column=3).value = avg_lead_time
    docSheet1.cell(row=i + 1, column=4).value = total_sub_one_days
    i += 1

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("FSB_Leadtime.xls")

print("Exported.")
