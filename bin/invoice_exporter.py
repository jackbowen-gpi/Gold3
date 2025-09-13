#!/usr/bin/python
"""Export invoice items into spreadsheet reports."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.workflow.models import Charge

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet
docSheet1 = workBookDocument.active
docSheet1.title = "Monthly Charges"
# Take Month and Workflow from request

# Conduct the master search
start_date = date(2008, 7, 1)
end_date = date(2008, 11, 1)
workflow = "Foodservice"
invoice_items = Charge.objects.filter(
    invoice_date__isnull=True,
    creation_date__range=(start_date, end_date),
    item__job__workflow__name=workflow,
).order_by("item__job__id", "item__id")

# Get list of plants to be invoiced
plant_list = []
for charge in invoice_items:
    plant = charge.item.printlocation.plant
    if plant not in plant_list:
        plant_list.append(plant)

# Create sheet for each plant.
for plant in plant_list:
    docSheet1 = workBookDocument.create_sheet(plant.name)

    # Label column headings
    docSheet1.cell(row=1, column=1).value = plant.name
    docSheet1.cell(row=1, column=2).value = "Job"
    docSheet1.cell(row=1, column=3).value = "Item"
    docSheet1.cell(row=1, column=4).value = "Charge Type"
    docSheet1.cell(row=1, column=5).value = "Amount"
    docSheet1.cell(row=1, column=6).value = "Charge Added"
    docSheet1.cell(row=1, column=7).value = "Salesperson"

    # Set column widths and other properties
    docSheet1.column_dimensions[0].width = 4200
    docSheet1.column_dimensions[1].width = 1600
    docSheet1.column_dimensions[2].width = 2400

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    plant_invoices = invoice_items.filter(item__printlocation__plant=plant)
    for i in range(len(plant_invoices)):
        # docSheet1.write(row, column, value)
        docSheet1.cell(row=i + 2, column=1).value = str(plant_invoices[i].item.job)
        docSheet1.cell(row=i + 2, column=2).value = str(plant_invoices[i].item.num_in_job) + "-  " + str(invoice_items[i].item.size)
        docSheet1.cell(row=i + 2, column=3).value = str(plant_invoices[i].description)
        docSheet1.cell(row=i + 2, column=4).value = plant_invoices[i].amount
        docSheet1.cell(row=i + 2, column=5).value = plant_invoices[i].creation_date
        docSheet1.cell(row=i + 2, column=6).value = str(plant_invoices[i].item.job.salesperson)

# Save XLS document
workBookDocument.save("invoicetest.xls")

print("Invoice Exported.")
