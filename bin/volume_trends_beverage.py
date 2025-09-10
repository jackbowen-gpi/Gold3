#!/usr/bin/python
"""Create beverage workflow volume trend reports.

Exports monthly item counts for the Beverage workflow.
"""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.workflow.models import Item

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

workflow = "Beverage"

start_date = date(2007, 1, 1)
item_set = Item.objects.filter(
    creation_date__gte=start_date,
    workflow__name=workflow,
    job__prepress_supplier__in=("", None, "OPT", "Optihue"),
).exclude(job__status="Cancelled")

print("Total items:", item_set.count())

# Create a new sheet for each plant.
docSheet1 = workBookDocument.active
docSheet1.title = "Data"

# Label column headings
docSheet1.cell(row=1, column=1).value = "Month"
docSheet1.cell(row=1, column=2).value = "Items In"

year = start_date.year
month = start_date.month
for x in range(39):
    # Increment rows, write data.
    str_date = str(month) + "-" + str(year)
    print(str_date)
    docSheet1.cell(row=x + 2, column=1).value = str_date
    monthly_items = item_set.filter(
        creation_date__year=year, creation_date__month=month
    )
    docSheet1.cell(row=x + 1, column=2).value = monthly_items.count()
    month += 1
    if month == 13:
        month = 1
        year += 1

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("xls_output/Beverage_Volumes.xls")

print("Exported.")
