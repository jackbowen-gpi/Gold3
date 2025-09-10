#!/usr/bin/python
"""Generate custom FSB billing analysis spreadsheets."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()

from datetime import date

from django.db.models import Sum

from gchub_db.apps.workflow.models import Charge, Job

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

# Setup date ranges.
# end_year = datetime.today().year
end_year = 2011
start_year = 2006

start_date = date(start_year, 11, 1)
end_date = date(end_year, 1, 31)

# Define months
year_set = (2006, 2007, 2008, 2009, 2010, 2011)
month_set = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)


# Some master data that needs to be established early.
# Heavy filtering here to obtain the most accurate customer-related data.
"""
master_job_list = Job.objects.filter(workflow__name="Foodservice",
                                     creation_date__range=(start_date, end_date))

"""
master_job_list = Job.objects.filter(
    workflow__name="Foodservice", name__icontains="infinity"
)

print("Job count:", master_job_list.count())


def _do_billing_analysis():
    """Filter total item list by plant, then look at filed vs no filed out."""
    print("Begin Plant Breakdown")
    # Add a page in the spreadsheet
    docSheet8 = workBookDocument.active
    docSheet8.title = "Plant"
    # Label column headings
    docSheet8.cell(row=1, column=1).value = ""
    docSheet8.cell(row=1, column=2).value = "Kenton"
    docSheet8.cell(row=1, column=3).value = "Shelbyville"
    docSheet8.cell(row=1, column=4).value = "Visalia"
    docSheet8.cell(row=1, column=5).value = "Marketing"
    docSheet8.cell(row=1, column=6).value = "Unknown"

    plant_set = ["Kenton", "Shelbyville", "Visalia", "Marketing", "Unknown"]

    # i is the row iterator
    i = 1
    for year in year_set:
        for month in month_set:
            docSheet8.cell(row=i + 1, column=1).value = "%s-%s" % (
                str(month),
                str(year),
            )
            charges = Charge.objects.filter(
                invoice_date__month=month,
                invoice_date__year=year,
                item__job__in=master_job_list,
            )
            plant_col = 1
            for plant in plant_set:
                plant_charges = charges.filter(item__printlocation__plant__name=plant)
                if plant_charges:
                    docSheet8.cell(
                        row=i + 1, column=plant_col + 1
                    ).value = plant_charges.aggregate(total=Sum("amount"))["total"]
                plant_col += 1
            # Iterate row for new month/year
            i += 1


# Execute each breakdown.
_do_billing_analysis()


# Save XLS document
workBookDocument.save("xls_output/InfinityBilling.xls")

print("Exported.")
