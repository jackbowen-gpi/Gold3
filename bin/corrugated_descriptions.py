#!/usr/bin/env python
"""Output items and descriptions for the automated corrugated system."""

import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from gchub_db.apps.auto_corrugated.models import BoxItem

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

# Add a page in the spreadsheet
docSheet1 = workBookDocument.active
docSheet1.title = "ACS Data"

docSheet1.cell(row=1, column=1).value = "Item"
docSheet1.cell(row=1, column=2).value = "English"
docSheet1.cell(row=1, column=3).value = "Spanish"
docSheet1.cell(row=1, column=4).value = "French"
docSheet1.cell(row=1, column=5).value = "English Lid"
docSheet1.cell(row=1, column=6).value = "Spanish Lid"
docSheet1.cell(row=1, column=7).value = "French Lid"

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

i = 1
for item in BoxItem.objects.filter(active=True):
    docSheet1.cell(row=i + 1, column=1).value = item.item_name
    docSheet1.cell(row=i + 1, column=2).value = item.english_description
    docSheet1.cell(row=i + 1, column=3).value = item.spanish_description
    if item.french_description:
        docSheet1.cell(row=i + 1, column=4).value = item.french_description
    docSheet1.cell(row=i + 1, column=5).value = item.english_lid_description
    docSheet1.cell(row=i + 1, column=6).value = item.spanish_lid_description
    if item.french_lid_description:
        docSheet1.cell(row=i + 1, column=7).value = item.french_lid_description
    i += 1

# Save XLS document
workBookDocument.save("xls_output/ACS Data.xls")
