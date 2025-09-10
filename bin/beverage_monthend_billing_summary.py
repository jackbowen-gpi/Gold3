#!/usr/bin/python
"""Generate month-end beverage billing summary spreadsheet."""

# openpyxl is now used instead of pyExcelerator because it supports
# python3 and pyexcelerator is deprecated
import bin_functions
import openpyxl

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.bev_billing.models import BevInvoice
from gchub_db.apps.workflow.models import Charge

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

# Set your month here and only here: (as a number)
# To be passed as a variable eventually.
month_num = 4
year_num = 2021

# Set to False when running first to verify charges, then set to True
# when running final.
# UPDATE_INVOICING = True
# INCLUDE_PLATES = True
# BILL_NOW_LIST = ()

# start_date = date(2008, 12, 21)
end_date = date(year_num, month_num, 21)
workflow = "Beverage"

month_name = end_date.strftime("%B")

# invoice entry date grabs invoices entered into QAD in the given month.
# This would indicate the handoff between GOLD and the invoice# generating system.
# Make sure all pending invoices to be entered into QAD are complete before
# running this report.
invoices = BevInvoice.objects.filter(
    qad_entry_date__month=month_num, qad_entry_date__year=year_num
)
print("Invoices", invoices.count())

billable_charges = Charge.objects.filter(bev_invoice__in=invoices)
print("Charges", billable_charges.count())

docSheet1 = workBookDocument.active
docSheet1.title = "%s Billing" % month_name

# Label column headings
docSheet1.cell(row=1, column=1).value = "Inv. Date"
docSheet1.cell(row=1, column=2).value = "Job No."
docSheet1.cell(row=1, column=3).value = "Plant"
docSheet1.cell(row=1, column=4).value = "PO Number"
docSheet1.cell(row=1, column=5).value = "Item"
docSheet1.cell(row=1, column=6).value = "Description"
docSheet1.cell(row=1, column=7).value = "Customer"
docSheet1.cell(row=1, column=8).value = "Brand"
docSheet1.cell(row=1, column=9).value = "PO Date"
docSheet1.cell(row=1, column=10).value = "Filed Out Date"
docSheet1.cell(row=1, column=11).value = "Invoice No."
docSheet1.cell(row=1, column=12).value = "Plate Supplier"
docSheet1.cell(row=1, column=13).value = "Charge To"
docSheet1.cell(row=1, column=14).value = "Bus. Type"
docSheet1.cell(row=1, column=15).value = "Charge Type"
docSheet1.cell(row=1, column=16).value = "BTC Amount"
docSheet1.cell(row=1, column=17).value = "AMO Amount"
docSheet1.cell(row=1, column=18).value = "ASN Amount"
docSheet1.cell(row=1, column=19).value = "Total Amount"

for i in range(len(billable_charges)):
    # Increment rows, write charge data.
    docSheet1.cell(row=i + 2, column=1).value = billable_charges[
        i
    ].bev_invoice.creation_date.strftime("%m/%d/%y")
    docSheet1.cell(row=i + 2, column=2).value = billable_charges[i].item.job.id
    docSheet1.cell(row=i + 2, column=3).value = str(
        billable_charges[i].item.printlocation.plant.name
    )
    docSheet1.cell(row=i + 2, column=4).value = str(
        billable_charges[i].item.job.po_number
    )
    docSheet1.cell(row=i + 2, column=5).value = str(
        billable_charges[i].item.bev_nomenclature()
    )
    docSheet1.cell(row=i + 2, column=6).value = str(
        billable_charges[i].item.description
    )
    #    print("Got to: %s" % billable_charges[i].item.job.id)
    docSheet1.cell(row=i + 2, column=7).value = str(
        billable_charges[i].item.job.customer_name
    )
    docSheet1.cell(row=i + 2, column=8).value = str(
        billable_charges[i].item.job.brand_name
    )
    docSheet1.cell(row=i + 2, column=9).value = billable_charges[
        i
    ].item.job.creation_date.strftime("%m/%d/%y")
    try:
        docSheet1.cell(row=i + 2, column=10).value = (
            billable_charges[i].item.final_file_date().strftime("%m/%d/%y")
        )
    except Exception:
        pass
    docSheet1.cell(row=i + 2, column=11).value = billable_charges[
        i
    ].bev_invoice.invoice_number
    try:
        docSheet1.cell(row=i + 2, column=12).value = str(
            billable_charges[i].item.job.temp_platepackage.platemaker.name
        )
    except AttributeError:
        # Handle NoneType errors.
        pass
    if billable_charges[i].description.type == "Revision (Evergreen Absorbs)":
        docSheet1.cell(row=i + 2, column=13).value = "AMO"
    else:
        docSheet1.cell(row=i + 2, column=13).value = str(
            billable_charges[i].item.job.bill_to_type
        )

    docSheet1.cell(row=i + 2, column=14).value = str(
        billable_charges[i].item.job.business_type
    )
    docSheet1.cell(row=i + 2, column=15).value = str(billable_charges[i].description)
    if billable_charges[i].description.type == "Revision (Evergreen Absorbs)":
        docSheet1.cell(row=i + 2, column=16).value = 0
        docSheet1.cell(row=i + 2, column=17).value = billable_charges[i].amount
        docSheet1.cell(row=i + 2, column=18).value = 0
    else:
        if billable_charges[i].item.job.bill_to_type in (
            "BTC",
            "Bill To Customer",
        ):
            docSheet1.cell(row=i + 2, column=16).value = billable_charges[i].amount
        else:
            docSheet1.cell(row=i + 2, column=16).value = 0
        if billable_charges[i].item.job.bill_to_type in (
            "AMO",
            "Absorbed Mfg Operations",
        ):
            docSheet1.cell(row=i + 2, column=17).value = billable_charges[i].amount
        else:
            docSheet1.cell(row=i + 2, column=17).value = 0
        if billable_charges[i].item.job.bill_to_type in (
            "ASN",
            "Absorbed Sales New",
        ):
            docSheet1.cell(row=i + 2, column=18).value = billable_charges[i].amount
        else:
            docSheet1.cell(row=i + 2, column=18).value = 0
    docSheet1.cell(row=i + 2, column=19).value = billable_charges[i].amount

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("EvergreenGraphics_%sBilling.xlsx" % month_name)

print("Exported.")
