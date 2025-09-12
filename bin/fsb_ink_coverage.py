#!/usr/bin/env python
"""Export ink coverage data for FSB into a spreadsheet."""

import bin_functions

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

import openpyxl

from django.db import models

from gchub_db.apps.joblog.app_defs import JOBLOG_TYPE_ITEM_FILED_OUT
from gchub_db.apps.joblog.models import JobLog
from gchub_db.apps.workflow.models import Item, ItemColor

# Set the plant and start date.
plant = "Shelbyville"
start_date = date(2020, 8, 1)
end_date = date.today()

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()

# Get items that have filed out for the given plant.
fileout_logs = JobLog.objects.filter(
    type=JOBLOG_TYPE_ITEM_FILED_OUT,
    event_time__range=(start_date, end_date),
    item__printlocation__plant__name=plant,
).values("item")
fileout_items = Item.objects.filter(id__in=fileout_logs).exclude(job__id=59300)

# Here's how we'll remove duplicates
# Make a list/dictionary of each unique bev_item_name in our filed-out items
# and note the most recent item that uses that name.
unique_names = fileout_items.values("id").order_by().annotate(max_id=models.Max("id"))
# Make a list of just the item IDs.
unique_items = [str(item["max_id"]) for item in unique_names]
# Filter filed-out items leaving only the most recent item to use that name.
ready_items = fileout_items.filter(id__in=unique_items)

# Gather the colors for the items specified.
qset = ItemColor.objects.filter(item__in=ready_items).order_by("-item__job__id")

print(len(qset))

title = "FSB Ink Coverage %s" % plant

# Add a page in the spreadsheet
docSheet1 = workBookDocument.active
docSheet1.title = plant

# Label column headings
docSheet1.cell(row=1, column=1).value = "Job"
docSheet1.cell(row=1, column=2).value = "Item Num"
docSheet1.cell(row=1, column=3).value = "Nine Digit"
docSheet1.cell(row=1, column=4).value = "Color"
docSheet1.cell(row=1, column=5).value = "Coverage (sq in)"

# Used to track which row we're on in the spreadsheet.
row = 1

# Write a row in the spreadsheet for each object in the query set.
for color in qset:
    docSheet1.cell(row=row + 1, column=1).value = int(color.item.job.id)
    docSheet1.cell(row=row + 1, column=2).value = int(color.item.num_in_job)
    docSheet1.cell(row=row + 1, column=3).value = str(color.item.fsb_nine_digit)
    docSheet1.cell(row=row + 1, column=4).value = str(color.fsb_display_name())
    # Leave this cell blank if the coverage is 'None'
    if color.coverage_sqin:
        docSheet1.cell(row=row + 1, column=5).value = float(color.coverage_sqin)

    # Move to the next row
    row += 1

# Freeze the top row of column headings.
docSheet1.panes_frozen = docSheet1["B2"]

# Save XLS document
workBookDocument.save("FSB_Ink_Coverage_Shelbyville.xlsx")
