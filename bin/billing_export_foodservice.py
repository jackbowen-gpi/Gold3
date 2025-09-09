#!/usr/bin/python
"""Export billing data for the Foodservice workflow into spreadsheet(s)."""
# openpyxl is now used instead of pyExcelerator because it supports python3 and pyexcelerator is deprecated
import calendar

import bin_functions
import openpyxl
from openpyxl.styles import NamedStyle

bin_functions.setup_paths()
import django

django.setup()
from datetime import date

from gchub_db.apps.budget import billing_funcs

# Setup the Worksheet
workBookDocument = openpyxl.Workbook()
# Setup the first sheet to be the summary sheet

# Set to False for initial run, then to True to update the billing charges
# as being invoiced.
UPDATE_INVOICING = False

# Set your month and year here and only here: (as a number)
# To be passed as a variable eventually.
# FORCE_BILLING = 59390
month_num = 6
year_num = 2020

print("Begin Foodservice Invoicing.")
print("Update as Invoiced: %s" % UPDATE_INVOICING)

# Set the month name for use in the file name.
month_name = calendar.month_abbr[month_num]

date_style = NamedStyle(name="datetime", number_format="DD/MM/YYYY")

if month_num == 12:
    next_month = 1
    next_year = year_num + 1
else:
    next_month = month_num + 1
    next_year = year_num

# This is needed for the invoice updating.
# Set to the first of the next month.
end_date = date(next_year, next_month, 1)
workflow = "Foodservice"

# Get billable charge qset for workflow.
billable_charges = billing_funcs.get_billable_data(year_num, month_num, workflow)[
    "charges"
]
# Use this qset if the charges have already been marked as invoiced, and the
# spreadsheet needs to be recreated.
# billable_charges = billing_funcs.get_invoiced_data(year_num, month_num, workflow)['charges']

print("Charges to invoice: %s" % str(billable_charges.count()))

# memphis cost center variables
memphisALL = "Memphis All 760279"
memphisFSB = "Memphis FSB 117409"

# QADClark = "QAD Migration - Clarksville"
# QADPitt = "QAD Migration - Pittston"

# Get list of plants to be invoiced
plants = []
items = []
for charge in billable_charges:
    # Add to the list of items for billing checks.
    if charge.item not in items:
        items.append(charge.item)
    try:
        if charge.item.printlocation.plant.name not in plants:
            if charge.item.printlocation.plant.name == "Memphis":
                if memphisALL not in plants:
                    plants.append(memphisALL)
                    plants.append(memphisFSB)
            #             elif "QAD/Avante" in charge.item.job.name:
            #                 #Add plants for QAD/Avante changeover so they get their own tabs
            #                 if charge.item.printlocation.plant.name == "Clarksville":
            #                     if QADClark not in plants:
            #                         plants.append(QADClark)
            #                 elif charge.item.printlocation.plant.name == "Pittston":
            #                     if QADPitt not in plants:
            #                         plants.append(QADPitt)
            else:
                plants.append(charge.item.printlocation.plant.name)
    except AttributeError:
        # Use Other to catch items not assigned to a plant.
        if "Other" not in plants:
            plants.append("Other")

print("---> BEGIN ITEM BILLING WARNINGS.")

for item in items:
    information = item.num_in_job, item, item.job, item.job.artist
    if item.check_too_few_charges():
        print("Warning: only one charge for item", information)
    if item.check_too_few_revision_charges():
        print("Warning: missing revision charges for item", information)
    if item.check_fileout_post_production():
        print("Warning: no post production charge for item", information)
    if item.check_prepress_charges():
        print("Warning: no prepress charge for item", information)
    if item.check_color_keys():
        print("Warning: no color keys for item", information)

plants.append("Avante-QAD")

sheetCounter = 0
for plant in plants:
    # Other is the category for items not assigned to a plant.
    if plant == "Other":
        charge_set = billable_charges.filter(item__printlocation__isnull=True)
    elif plant == memphisALL:
        charge_set = billable_charges.filter(
            item__printlocation__plant__name="Memphis",
            item__printlocation__press__name="All",
        )
    elif plant == memphisFSB:
        charge_set = billable_charges.filter(
            item__printlocation__plant__name="Memphis",
            item__printlocation__press__name="FSB",
        )
    #    elif plant == QADClark:
    #        charge_set = billable_charges.filter(item__job__name__contains="QAD/Avante", item__printlocation__plant__name="Clarksville", item__printlocation__press__name="Other")
    #    elif plant == QADPitt:
    #        charge_set = billable_charges.filter(item__job__name__contains="QAD/Avante", item__printlocation__plant__name="Pittston", item__printlocation__press__name="Other")
    elif plant == "Avante-QAD":
        charge_set = billable_charges.filter(
            item__job__name__startswith="Letica - QAD/Avante"
        )
    else:
        charge_set = billable_charges.filter(
            item__printlocation__plant__name=plant
        ).exclude(item__job__name__startswith="Letica - QAD/Avante")

    if len(charge_set) == 0:
        continue
    # Some feedback as to how many charges per plant there are.
    print(plant, charge_set.count())

    if sheetCounter == 0:
        docSheet1 = workBookDocument.active
        docSheet1.title = "%s Billing" % plant
    else:
        # Create a new sheet for each plant.
        docSheet1 = workBookDocument.create_sheet("%s Billing" % plant)

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Date Billed"
    docSheet1.cell(row=1, column=2).value = "Salesperson"
    docSheet1.cell(row=1, column=3).value = "Job No."
    docSheet1.cell(row=1, column=4).value = "Job Name"
    docSheet1.cell(row=1, column=5).value = "Size"
    docSheet1.cell(row=1, column=6).value = "Filed Out Date"
    docSheet1.cell(row=1, column=7).value = "Charge Type"
    docSheet1.cell(row=1, column=8).value = "Rush Days"
    docSheet1.cell(row=1, column=9).value = "Plant"
    docSheet1.cell(row=1, column=10).value = "Press"
    docSheet1.cell(row=1, column=11).value = "Amount"

    sheetCounter = sheetCounter + 1

    for i in range(len(charge_set)):
        # Increment rows, write charge data.
        # docSheet1.write(row, column, value)creation_date
        docSheet1.cell(row=i + 2, column=1).value = charge_set[
            i
        ].creation_date.strftime("%m/%d/%y")
        try:
            docSheet1.cell(row=i + 2, column=2).value = str(
                charge_set[i].item.job.salesperson.username
            )
        except Exception:
            pass
        docSheet1.cell(row=i + 2, column=3).value = str(charge_set[i].item.job.id)
        docSheet1.cell(row=i + 2, column=4).value = str(
            charge_set[i].item.job.name.encode("utf8", "replace")
        )
        docSheet1.cell(row=i + 2, column=5).value = str(charge_set[i].item.size)

        docSheet1.cell(row=i + 2, column=6).value = (
            charge_set[i].item.final_file_date().strftime("%m/%d/%y")
        )
        docSheet1.cell(row=i + 2, column=7).value = str(charge_set[i].description)
        docSheet1.cell(row=i + 2, column=8).value = str(charge_set[i].rush_days)
        try:
            docSheet1.cell(row=i + 2, column=9).value = str(
                charge_set[i].item.printlocation.plant.name
            )
        except AttributeError:
            docSheet1.cell(row=i + 2, column=9).value = "----"
        try:
            docSheet1.cell(row=i + 2, column=10).value = str(
                charge_set[i].item.printlocation.press.name
            )
        except AttributeError:
            docSheet1.cell(row=i + 2, column=10).value = "----"
        docSheet1.cell(row=i + 2, column=11).value = charge_set[i].amount

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

if UPDATE_INVOICING:
    # Mark all charges as invoiced today. Set Invoice number to job number?
    for charge in billable_charges:
        # today = datetime.date.today()
        charge.invoice_date = end_date
        charge.invoice_number = str(charge.item.job.id)
        charge.save()

# Save XLS document
workBookDocument.save("FSB_%s_Billing.xlsx" % month_name)

print("Exported.")
