"""Manager Tool Views"""

import calendar
import os
from datetime import date, timedelta
from io import BytesIO
from tempfile import NamedTemporaryFile

# openpyxl is now used instead of pyExcelerator because it supports python3 and pyexcelerator is deprecated
import openpyxl
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission, User
from django.contrib.contenttypes.models import ContentType
from django.contrib.sites.models import Site
from django.db.models import Q, Sum
from django.forms import DateField
from django.http import HttpResponse
from django.shortcuts import render

from gchub_db.apps.budget import billing_funcs
from gchub_db.apps.joblog.app_defs import *
from gchub_db.apps.joblog.models import JobLog
from gchub_db.apps.manager_tools.manager_tool_funcs import (
    get_item_average_hours,
    get_job_average_hours,
)
from gchub_db.apps.qc.models import QCCategory, QCResponseDoc, QCWhoops
from gchub_db.apps.timesheet.models import TimeSheet, TimeSheetCategory
from gchub_db.apps.workflow.app_defs import *
from gchub_db.apps.workflow.models import (
    Charge,
    Item,
    ItemTracker,
    Job,
    JobComplexity,
    Plant,
    ProofTracker,
    Revision,
)
from gchub_db.includes import general_funcs


def home(request):
    """Main page for Manager's Tools."""
    pagevars = {
        "page_title": "Manager Tools Home",
    }

    return render(request, "manager_tools/home.html", context=pagevars)


class HourByPlantForm(forms.Form):
    """Form used to display the Hours by Plant manager tool."""

    # Month choice range.
    month_choices = []
    for month_num in range(1, 13):
        month_choices.append((month_num, calendar.month_name[month_num]))

    # Year choice range.
    year_choices = []
    year_range = list(range(2018, general_funcs._utcnow_naive().date().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    month = forms.ChoiceField(
        choices=month_choices, initial=general_funcs._utcnow_naive().date().month
    )
    year = forms.ChoiceField(
        choices=year_choices, initial=general_funcs._utcnow_naive().date().year
    )
    plant = forms.ModelChoiceField(queryset=Plant.objects.all().order_by("name"))


@login_required
def hoursbyplant(request):
    """How many timesheet hours we've logged for a given plant in a given month."""
    # Process a submitted form.
    if request.POST:
        form = HourByPlantForm(request.POST)
        if form.is_valid():
            # Read choices from the form.
            month = form.cleaned_data.get("month", None)
            year = form.cleaned_data.get("year", None)
            month = int(month)
            year = int(year)
            plant = form.cleaned_data.get("plant", None)

            # Gather the data.
            timesheets = TimeSheet.objects.filter(date__month=month, date__year=year)
            hours = 0

            # Exclude the timesheets not associated with the selected plant.
            for sheet in timesheets:
                try:
                    items = Item.objects.filter(
                        job=sheet.job, printlocation__plant__name=plant
                    )
                except Exception:
                    items = None
                if items:
                    hours += sheet.hours

            welcome_message = None

    # Show a blank page without data at first.
    else:
        form = HourByPlantForm()
        hours = None
        welcome_message = (
            "Set the plant and month then hit the green refresh arrow to see data."
        )

    pagevars = {
        "page_title": "Hours by Plant Report",
        "welcome_message": welcome_message,
        "form": form,
        "hours": hours,
    }
    return render(request, "manager_tools/hoursbyplant.html", context=pagevars)


def sick(request):
    """Sick Manager's Tool."""
    if request.POST:
        user = User.objects.get(id=request.POST["user_id"])
        profile = user.profile
        profile.total_sick = request.POST["new_total"]
        profile.save()

    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        try:
            if (
                user.has_perms(["accounts.clemson_employee"])
                and user.username != "James_Baxter"
            ):
                clemson_employee_list.append(user)
        except Exception:
            # if profile/perms lookup fails for a user, skip them
            pass

    pagevars = {
        "page_title": "Sick Tool",
        "clemson_employee_list": clemson_employee_list,
    }


@login_required
def stalecharges(request):
    """Page for totaling up un-invoiced charges by sales person and print group.
    We don't load the page with default data so the page will load quickly and
    the user can select the year before they incur the longer page load. Most
    of the logic is handled by stalecharges_by_date().
    """
    # Submitted form. Calculate data. (Slow)
    if request.POST:
        staleform = StaleChargesForm(request.POST)
        if staleform.is_valid():
            # Read choices from the form.
            date_from = staleform.cleaned_data.get("date_from", None)
            date_to = staleform.cleaned_data.get("date_to", None)

            # Only show the welcome message when the page first loads.
            welcome_message = None

            # Calculate the un-invoiced charges for that year.
            data = stalecharges_by_date(date_from, date_to)

            # Format the dates for the excel download URL.
            date_from_url = "%s/%s/%s" % (
                date_from.year,
                date_from.month,
                date_from.day,
            )
            date_to_url = "%s/%s/%s" % (date_to.year, date_to.month, date_to.day)

    # Show a blank page without data at first.
    else:
        staleform = StaleChargesForm()
        welcome_message = "Set the start and end dates then hit the green refresh arrow to see data. Be patient, it can take 30 seconds or more."
        data = None
        date_from_url = None
        date_to_url = None

    pagevars = {
        "page_title": "Stale Charges",
        "welcome_message": welcome_message,
        "staleform": staleform,
        "data": data,
        "date_from_url": date_from_url,
        "date_to_url": date_to_url,
    }

    return render(request, "manager_tools/stalecharges.html", context=pagevars)


def stalecharges_excel(
    request, year_from, month_from, day_from, year_to, month_to, day_to
):
    """Download raw metrics data as an excel spreadsheet."""
    # Set the time span.
    date_from = date(int(year_from), int(month_from), int(day_from))
    date_to = date(int(year_to), int(month_to), int(day_to))

    # Calculate the un-invoiced charges for that time span.
    data = stalecharges_by_date(date_from, date_to)

    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = "Charges"

    # Used to track which row we're on in the spreadsheet.
    row = 0

    # Write the title at the top of the sheet.
    title_text = "Un-Invoiced Charges by Sales Person and Print Group (%s to %s)" % (
        date_from,
        date_to,
    )
    docSheet1.cell(row=row + 1, column=1).value = title_text
    row += 2

    # Write a section for each sales person in the list of data.
    # for sales_person, print_groups in data.iteritems():
    for sales_person in data:
        if sales_person[0]:
            docSheet1.cell(row=row + 1, column=1).value = "%s %s" % (
                sales_person[0].first_name,
                sales_person[0].last_name,
            )
        else:
            docSheet1.cell(row=row + 1, column=1).value = "No sales person"
        # Write a row for each print group
        for print_group in sales_person[1]:
            docSheet1.cell(row=row + 1, column=2).value = print_group[
                0
            ]  # Print group name
            docSheet1.cell(row=row + 1, column=3).value = print_group[1]  # Total
            # Move to the next row
            row += 1
        # Write a blank row between sales people.
        docSheet1.cell(row=row + 1, column=1).value = " "
        row += 1

    # Save the spreadsheet.
    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")
    response["Content-Disposition"] = (
        'attachment; filename="stale_charges_%s.xlsx"' % year_to
    )
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response


def stalecharges_by_date(date_from, date_to):
    """Totals charges that haven't been invoiced for a given time span and then breaks
    them out by sales person and print group.
    """
    items = Item.objects.filter(
        creation_date__range=(date_from, date_to),
        workflow__name="Foodservice",
        is_deleted=False,
    ).order_by("creation_date")

    # Stale charges (aka un-invoiced charges)
    stale_charges = Charge.objects.filter(item__in=items, invoice_date__isnull=True)

    """
    Since we need each sales person's charges grouped by print group we're going
    to store them in a dictionary. For each entry the key will be the sales person's
    user object and the value will be yet another dictionary. In this print group
    dictionary the key will be the print group's name and the value will be the
    total charges for that print group from the sales person's jobs.

    It will look like this:

        example_data = {
          <User: Mike_Turner>: {"Wendys": 123, "Starbucks":8394, "McDonalds": 87643},
          <User: Dana_Warf>: {"Dunkin": 976, "Buckys":4323, "Tims": 22987}
        }

    """

    # Master data dictionary.
    data = {}

    # Iterate through the stale charges and record them in the data dictionary.
    for charge in stale_charges:
        # Get the sales person for this charge.
        sales_person = charge.item.job.salesperson
        # Get this print group from the job this charge belongs to.
        if charge.item.job.printgroup:
            # Some don't have descriptions so use the name.
            print_group = str(charge.item.job.printgroup.name)
        else:
            print_group = "No print group"

        # Add the sales person to the master dictionary if needed.
        if sales_person not in data:
            # Include a "Total" print group for displaying the total charges.
            data[sales_person] = {"Total": 0}
        # Add the print group to this sales person's print group dictionary if needed.
        if print_group not in data[sales_person]:
            # Start with a total of zero.
            data[sales_person][print_group] = 0
        # Add the charge to the appropriate total in this sales persons print group dictionary.
        data[sales_person][print_group] += charge.amount
        # Also add it to that sales person's total
        data[sales_person]["Total"] += charge.amount

    """
    Convert the master data dictionary to a list of lists. Lists are easier to
    sort and we don't need dictionary functionality now that we're done adding to
    totals.

    Each list in the list will look like this:

        [<User: James_McCracken>, {'Total': 100.0, 'FAKEPG1': 40.0, 'FAKEPG2': 60.0}]

    The first element is the user object and the second is their print group
    dictionary.
    """
    # Start by converting the dictionary to a list of tuples.
    tuples_of_data = list(data.items())

    # A function that returns a sales person's total stale charges. Used for sorting.
    def TotalSort(val):
        return val[1]["Total"]

    # Sort the list of tuples by the total stale charges for each sales person.
    tuples_of_data.sort(key=TotalSort, reverse=True)

    # Now convert each tuple in the list to a list so we can change elements. This
    # results in a list of lists.
    list_of_data = []
    for tuple in tuples_of_data:
        # Convert the tuple to a list.
        list_of_data.append(list(tuple))

    """
    Now we'll sort all the print group dictionaries in a similar manner. Each list
    will look like this:

        [('Total', 100.0), ('FAKEPG1', 40.0), ('FAKEPG2', 60.0)]

    """

    # A function that returns a print group's total. Used for sorting.
    def PGTotalSort(val):
        return val[1]

    for row in list_of_data:
        # Convert the dictionary of printgroups to a list of tuples.
        list_of_printgroups = list(row[1].items())
        # Sort by totals
        list_of_printgroups.sort(key=PGTotalSort, reverse=True)
        # Make sure "Total" is the first item in the list.
        if list_of_printgroups[0][0] != "Total":
            list_of_printgroups.reverse()
        # Save the sorted list of tuples over the old dictionary of printgroups.
        row[1] = list_of_printgroups

    return list_of_data


def vacation(request):
    """Vacation Manager's Tool."""
    if request.POST:
        user = User.objects.get(id=request.POST["user_id"])
        profile = user.profile
        profile.total_vacation = request.POST["new_total"]
        profile.save()

    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        if (
            user.has_perms(["accounts.clemson_employee"])
            and user.username != "James_Baxter"
        ):
            clemson_employee_list.append(user)

    pagevars = {
        "page_title": "Vacation Tool",
        "clemson_employee_list": clemson_employee_list,
    }

    return render(request, "manager_tools/vacation.html", context=pagevars)


class QCForm(forms.Form):
    """Form used to display the qc manager tool for a given year."""

    # Year choice range.
    year_choices = []
    year_range = list(range(2009, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    year = forms.ChoiceField(choices=year_choices)


@login_required
def qc(request):
    """QC manager's report. Details errors per employee for the current year.
    User can also select a previous year.
    """
    # Default to this year.
    if request.POST:
        qcform = QCForm(request.POST)
        year = request.POST["year"]
    else:
        year = date.today().year
        qcform = QCForm(initial={"year": year})

    # Gather up all Clemson employees.
    clemson_perm = Permission.objects.get(codename="clemson_employee")
    users = (
        User.objects.filter(
            Q(groups__permissions=clemson_perm) | Q(user_permissions=clemson_perm),
            is_active=True,
        )
        .distinct()
        .order_by("last_name")
    )

    # Create a list for user metrics.
    qc_metrics_by_user = []

    # Gather up the current QC categories.
    qc_categories = QCCategory.objects.all().order_by("order")

    # Iterate through the users and gather data.
    for user in users:
        errors = QCWhoops.objects.filter(
            qc_response__qcdoc__job__artist=user, reported_date__year=year
        )
        num_errors = errors.count()
        num_unresolved = errors.filter(
            resolution_date__isnull=True, is_valid=True
        ).count()
        num_resolved = errors.filter(
            resolution_date__isnull=False, is_valid=True
        ).count()
        num_invalid = errors.filter(is_valid=False).count()
        # Pack this user's data up into a dictionary.
        dict = {
            "first_name": str(user.first_name),
            "last_name": str(user.last_name),
            "num_errors": num_errors,
            "num_unresolved": num_unresolved,
            "num_resolved": num_resolved,
            "num_invalid": num_invalid,
        }
        # Count the number of errors in each category for the user.
        errors_by_category = []
        for category in qc_categories:
            cat_errors = errors.filter(qc_response__category=category)
            errors_by_category.append(cat_errors.count())
        dict["errors_by_category"] = errors_by_category
        # Add the user's data dictionary to the list of user metrics.
        qc_metrics_by_user.append(dict)

    pagevars = {
        "page_title": "QC Report",
        "qcform": qcform,
        "qc_metrics_by_user": qc_metrics_by_user,
        "year": year,
        "qc_categories": qc_categories,
    }

    return render(request, "manager_tools/qc.html", context=pagevars)


class WorkflowModelChoiceField(forms.ModelChoiceField):
    """Used to make the beverage workflow display as evergreen."""

    def label_from_instance(self, obj):
        if obj.name == "Beverage":
            return "Evergreen"
        else:
            return obj.name


class StaleChargesForm(forms.Form):
    """Form used to select the time span in the stale charges report."""

    date_from = DateField()
    date_to = DateField()


class TurnTimeForm(forms.Form):
    """Form used to display the turn time manager tool for a given year and
    workflow. We only let the user go back to 2018 because there was a bug in
    the data before then so it's not valid.
    """

    # Year choice range.
    year_choices = []
    year_range = list(range(2018, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Workflow choices
    available = ["Foodservice", "Carton", "Beverage"]
    workflow_choices = Site.objects.filter(name__in=available).order_by("-name")

    # Fields
    year = forms.ChoiceField(choices=year_choices)
    workflow = WorkflowModelChoiceField(queryset=workflow_choices, empty_label=None)


class MonthlyBillingForm(forms.Form):
    """Form used to display the cost avoidance ammounts for a given year."""

    # Year choice range.
    year_choices = []
    year_range = list(range(2019, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    year = forms.ChoiceField(choices=year_choices)


class CostAvoidanceForm(forms.Form):
    """Form used to display the cost avoidance ammounts for a given year."""

    # Year choice range.
    year_choices = []
    year_range = list(range(2019, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    year = forms.ChoiceField(choices=year_choices)


@login_required
def turntimes_by_artist(request, month, year, workflow):
    """Breaks turn time averages out by artist for a given month and workflow.

    Returns a list of dictionaries. Each dictionary is an artist and the keys
    are "name" and "turn_time" which is the average turn time for the artist
    in days.
    """
    # Convert the month and year to a number if they aren't one already.
    month_nums = range(1, 12)
    if month not in month_nums:
        month_name_to_num = {
            name: num for num, name in enumerate(calendar.month_name) if num
        }
        month = month_name_to_num[month]
    year = int(year)

    # Get the workflow object based on the supplied name
    workflow = Site.objects.get(name=workflow)

    # Gather up all Clemson employees.
    clemson_perm = Permission.objects.get(codename="clemson_employee")
    users = (
        User.objects.filter(
            Q(groups__permissions=clemson_perm) | Q(user_permissions=clemson_perm),
            is_active=True,
        )
        .distinct()
        .order_by("last_name")
    )

    data = []

    # Make the first entry the average for all artists again.
    overall_data = turntimes_by_month(month, year, workflow)
    data.append(
        {
            "name": "All Artists",
            "id": None,
            "turn_time": overall_data["days"],
        }
    )

    # Exclude any test jobs
    test_jobs = [59300, 99999]

    # Calculate data for each artist.
    for artist in users:
        #         print(artist)
        # Gather the artist's items that proofed out this month.
        proof_joblogs = (
            JobLog.objects.filter(
                event_time__year=year,
                event_time__month=month,
                type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
                job__workflow__name=workflow.name,
                job__artist=artist,
            )
            .exclude(job__id__in=test_jobs)
            .distinct("item__id")
        )

        # Make sure the items didn't also proof out in a prior month. Items should count in the month they proofed out the first time.
        items = []
        for log in proof_joblogs:
            # Get the item
            item = log.item
            # Get the first proof out event for that item
            proof_joblog = (
                JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
                .order_by("event_time")
                .first()
            )
            # Make sure this items first proof out event was this month.
            if (
                proof_joblog.event_time.year == year
                and proof_joblog.event_time.month == month
            ):
                items.append(item)

        # Keep a list of the timespans between our start and finish events.
        timespans = []

        # Different workflows measure turn times from different events.
        if workflow.name == "Foodservice":
            # Go through each item and record the time from print location
            # assignment to proof if those two events occurred.
            for item in items:
                assignment_date = False
                proof_date = False
                timespan = False

                try:
                    # Assignment date
                    assignment_date = item.assignment_date
                    # First proof date
                    proof_joblog = (
                        JobLog.objects.filter(
                            item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT
                        )
                        .order_by("event_time")
                        .first()
                    )
                    proof_date = proof_joblog.event_time.date()
                except Exception:
                    pass

                if assignment_date and proof_date:
                    timespan = proof_date - assignment_date
                    # Ignore negative time spans.
                    if timespan.days >= 0:
                        #                         print("    %s-%s: %s" %(item.job.id, item.num_in_job, timespan))
                        timespans.append(timespan)

        elif workflow.name == "Beverage" or workflow.name == "Carton":
            # Go through each item and record the time from preflight to proof
            # if those two events occured.
            for item in items:
                preflight_date = False
                proof_date = False
                timespan = False

                try:
                    # Preflight date
                    preflight_date = item.preflight_date
                    # First proof date
                    proof_joblog = (
                        JobLog.objects.filter(
                            item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT
                        )
                        .order_by("event_time")
                        .first()
                    )
                    proof_date = proof_joblog.event_time.date()
                except Exception:
                    pass

                if preflight_date and proof_date:
                    timespan = proof_date - preflight_date
                    # Ignore negative time spans.
                    if timespan.days >= 0:
                        #                         print("    %s-%s: %s" %(item.job.id, item.num_in_job, timespan))
                        timespans.append(timespan)

        else:
            print("No turn time criteria available for %s jobs." % workflow.name)

        # Calculate the average timespan if there is one.
        if len(timespans) != 0:
            average_time = sum(timespans, timedelta()) / len(timespans)
            average_time = average_time.days
            # Create a dictionary for the artist with their name and average.
            full_name = "%s %s" % (artist.first_name, artist.last_name)
            dict = {
                "name": str(full_name),
                "id": artist.id,
                "turn_time": average_time,
            }
            # Append the data to the dictionary.
            data.append(dict)

    # Set up the page title.
    display_month = calendar.month_name[month]
    if workflow.name == "Beverage":
        display_workflow = "Evergreen"
    else:
        display_workflow = workflow.name
    page_title = "%s Turn Time Averages for %s %s" % (
        display_workflow,
        display_month,
        year,
    )

    pagevars = {
        "page_title": page_title,
        "month": month,
        "year": year,
        "workflow": workflow,
        "data": data,
    }

    return render(request, "manager_tools/turntime_artists.html", context=pagevars)


@login_required
def turntimes_by_item(request, month, year, workflow, artist_id):
    """Lists an artists turntimes by item for a given month and workflow.

    Returns a list of dictionaries. Each dictionary is an item and the keys
    are "name" and "turn_time".
    """
    month = int(month)
    year = int(year)

    # Get the workflow object based on the supplied name
    workflow = Site.objects.get(name=workflow)

    # Gather up all Clemson employees.
    clemson_perm = Permission.objects.get(codename="clemson_employee")
    artist = User.objects.get(id=artist_id)

    data = []

    # Exclude any test jobs
    test_jobs = [59300, 99999]

    # Gather the artist's items that proofed out this month.
    proof_joblogs = (
        JobLog.objects.filter(
            event_time__year=year,
            event_time__month=month,
            type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
            job__workflow__name=workflow.name,
            job__artist=artist,
        )
        .exclude(job__id__in=test_jobs)
        .distinct("item__id")
    )

    # Make sure the items didn't also proof out in a prior month. Items should count in the month they proofed out the first time.
    items = []
    for log in proof_joblogs:
        # Get the item
        item = log.item
        # Get the first proof out event for that item
        proof_joblog = (
            JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
            .order_by("event_time")
            .first()
        )
        # Make sure this items first proof out event was this month.
        if (
            proof_joblog.event_time.year == year
            and proof_joblog.event_time.month == month
        ):
            items.append(item)

    # Different workflows measure turn times from different events.
    if workflow.name == "Foodservice":
        # Go through each item and record the time from print location
        # assignment to proof if those two events occurred.
        for item in items:
            assignment_date = False
            proof_date = False
            timespan = False

            try:
                # Assignment date
                assignment_date = item.assignment_date
                # First proof date
                proof_joblog = (
                    JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
                    .order_by("event_time")
                    .first()
                )
                proof_date = proof_joblog.event_time.date()
            except Exception:
                pass

            if assignment_date and proof_date:
                timespan = proof_date - assignment_date
                # Ignore negative time spans.
                if timespan.days >= 0:
                    final_timespan = str(timespan)
                    data.append(
                        {
                            "name": "%s - %s" % (item.job.id, item.num_in_job),
                            "turn_time": final_timespan.split(", ")[0],
                        }
                    )

    elif workflow.name == "Beverage" or workflow.name == "Carton":
        # Go through each item and record the time from preflight to proof
        # if those two events occured.
        for item in items:
            preflight_date = False
            proof_date = False
            timespan = False

            try:
                # Preflight date
                preflight_date = item.preflight_date
                # First proof date
                proof_joblog = (
                    JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
                    .order_by("event_time")
                    .first()
                )
                proof_date = proof_joblog.event_time.date()
            except Exception:
                pass

            if preflight_date and proof_date:
                timespan = proof_date - preflight_date
                # Ignore negative time spans.
                if timespan.days >= 0:
                    final_timespan = str(timespan)
                    data.append(
                        {
                            "name": "%s - %s" % (item.job.id, item.num_in_job),
                            "turn_time": final_timespan.split(", ")[0],
                        }
                    )

    else:
        print("No turn time criteria available for %s jobs." % workflow.name)

    # Set up the page title.
    display_month = calendar.month_name[month]
    if workflow.name == "Beverage":
        display_workflow = "Evergreen"
    else:
        display_workflow = workflow.name
    page_title = "%s Turn Times for %s %s in %s %s" % (
        display_workflow,
        artist.first_name,
        artist.last_name,
        display_month,
        year,
    )

    pagevars = {
        "page_title": page_title,
        "data": data,
    }

    return render(request, "manager_tools/turntime_items.html", context=pagevars)


def billing_by_month(month, year):
    """Monthly Billing Report"""
    # For some reason the billing script always sets the invoice date
    # as the first day of the following month. We correct for that here.
    if month == 12:
        target_month = 1
        target_year = year + 1
    else:
        target_month = month + 1
        target_year = year

    workflow = "Foodservice"
    # Get billable charge qset for workflow.
    billable_charges = billing_funcs.get_billable_data(
        target_year, target_month, workflow
    )["charges"]
    # Use this qset if the charges have already been marked as invoiced, and the
    # spreadsheet needs to be recreated.
    # billable_charges = billing_funcs.get_  invoiced    _data(year_num, month_num, workflow)['charges']
    qty = billable_charges.count()
    # Gather Foodservice cost avoidance items
    items = Item.objects.filter(
        workflow__name="Foodservice", job__type="BillableSales", is_deleted=False
    )

    # Gather invoiced charges for cost avoidance items
    invoiced_charges = Charge.objects.filter(
        item__in=items, invoice_date__year=target_year, invoice_date__month=target_month
    ).order_by("invoice_date")

    # Total the charges
    total = 0
    for charge in invoiced_charges:
        total += charge.amount

    # Append data to master list
    dict = {
        "workflow": workflow,
        "year": year,
        "month": month,
        "charges": qty,
        "total": total,
    }

    return dict


def costavoidance_by_month(month, year):
    """Calculates the total charges for all the jobs marked as "cost avoidance" for
    a given month. Keys off a charge's invoice date and that doesn't get filled
    in until we run month end billing. So, if the current month is zero then
    billing probably hasn't been run yet.
    """
    # For some reason the billing script always sets the invoice date
    # as the first day of the following month. We correct for that here.
    if month == 12:
        target_month = 1
        target_year = year + 1
    else:
        target_month = month + 1
        target_year = year

    # Gather Foodservice cost avoidance items
    items = Item.objects.filter(
        workflow__name="Foodservice", job__type="CostAvoidance", is_deleted=False
    )

    # Gather invoiced charges for cost avoidance items
    invoiced_charges = Charge.objects.filter(
        item__in=items, invoice_date__year=target_year, invoice_date__month=target_month
    ).order_by("invoice_date")

    # Total the charges
    total = 0
    for charge in invoiced_charges:
        total += charge.amount

    # Append data to master list
    dict = {
        "name": calendar.month_name[month],
        "total": total,
    }

    return dict


def turntimes_by_month(month, year, workflow):
    """Calculates average turn time in days. Foodservice is measured printlocation
    assignment date to first proof date. Beverage and Carton are from first
    preflight date to first proof date.

    Returns a dictionary representing a month and the keys are "name" and
    "days". "Name" is the name of the month and "days"is the average turn time
    in days for that month.

    We only go back to 2018 because data from before then isn't valid due to a
    bug where item print assignment dates got copied over from old items when
    press changes were entered. This bug was fixed in Feb 2018.
    """
    # FSB data from Jan and Feb 2018 isn't valid. Do not display.
    if (workflow.name == "Foodservice") and (year == 2018):
        if month == 1 or month == 2:
            # Append data to master list
            dict = {
                "name": calendar.month_name[month],
                "days": "No data",
            }
            return dict  # We're done.

    # Exclude any test jobs
    test_jobs = [59300, 99999]

    # Gather the items that proofed out this month.
    proof_joblogs = (
        JobLog.objects.filter(
            event_time__year=year,
            event_time__month=month,
            type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
            job__workflow__name=workflow.name,
        )
        .exclude(job__id__in=test_jobs)
        .distinct("item__id")
    )

    # Make sure the items didn't also proof out in a prior month. Items should count in the month they proofed out the first time.
    items = []
    for log in proof_joblogs:
        # Get the item
        item = log.item
        # Get the first proof out event for that item
        proof_joblog = (
            JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
            .order_by("event_time")
            .first()
        )
        # Make sure this items first proof out event was this month.
        if (
            proof_joblog.event_time.year == year
            and proof_joblog.event_time.month == month
        ):
            items.append(item)

    # Keep a list of the timespans between our start and finish
    timespans = []

    # Different workflows measure turn times from different events.
    if workflow.name == "Foodservice":
        # Go through each item and record the time from print location
        # assignment to proof if those two events occurred.
        for item in items:
            assignment_date = False
            proof_date = False
            timespan = False

            try:
                # Assignment date
                assignment_date = item.assignment_date
                # First proof date
                proof_joblog = (
                    JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
                    .order_by("event_time")
                    .first()
                )
                proof_date = proof_joblog.event_time.date()
            except Exception:
                pass

            if assignment_date and proof_date:
                timespan = proof_date - assignment_date
                # Ignore negative time spans.
                if timespan.days >= 0:
                    timespans.append(timespan)

    elif workflow.name == "Beverage" or workflow.name == "Carton":
        # Go through each item and record the time from preflight to proof
        # if those two events occured.
        for item in items:
            preflight_date = False
            proof_date = False
            timespan = False

            try:
                # Preflight date.
                preflight_date = item.preflight_date
                # First proof date
                proof_joblog = (
                    JobLog.objects.filter(item=item, type=JOBLOG_TYPE_ITEM_PROOFED_OUT)
                    .order_by("event_time")
                    .first()
                )
                proof_date = proof_joblog.event_time.date()
            except Exception:
                pass

            if preflight_date and proof_date:
                timespan = proof_date - preflight_date
                # Ignore negative time spans.
                if timespan.days >= 0:
                    timespans.append(timespan)

    else:
        print("No turn time criteria available for %s jobs." % workflow.name)

    # Calculate the average timespan if there is one.
    if len(timespans) > 0:
        average_time = sum(timespans, timedelta()) / len(timespans)
        average_time = average_time.days
    else:
        average_time = "No data yet"
    # Append data to master list
    dict = {
        "name": calendar.month_name[month],
        "days": average_time,
    }

    return dict


@login_required
def turntime(request):
    """Turn time manager's report. We don't load the page with default data
    so the page will load quickly and the user can select their workflow and
    year before they incur the longer page load.
    """
    # Submitted form. Calculate data. (Slow)
    if request.POST:
        turnform = TurnTimeForm(request.POST)
        if turnform.is_valid():
            year = int(turnform.cleaned_data.get("year"))
            workflow = turnform.cleaned_data.get("workflow")

            # Only show the welcome message when the page first loads.
            welcome_message = None

            # Create a list for the data.
            turn_data = []

            # Need set where our month range ends.
            # End it at the current month if displaying the current year.
            if year == date.today().year:
                end_month = date.today().month
                end_month += 1
            # End it after December for all prior years.
            else:
                end_month = 13

            # Calculate the average time for each month in the year.
            for month in range(1, end_month):
                data = turntimes_by_month(month, year, workflow)
                turn_data.append(data)

    # Show a blank page without data at first.
    else:
        year = date.today().year
        workflow = "fsb"
        turnform = TurnTimeForm(
            initial={
                "year": year,
                "workflow": workflow,
            }
        )
        turn_data = None
        welcome_message = "Set the workflow and year then hit the green refresh arrow to see data. Be patient, it can take 30 seconds or more."

    # Pass the current year for some display stuff.
    current_year = date.today().year

    pagevars = {
        "page_title": "Turn Time Report",
        "turnform": turnform,
        "turn_data": turn_data,
        "year": year,
        "current_year": current_year,
        "workflow": workflow,
        "welcome_message": welcome_message,
    }

    return render(request, "manager_tools/turntime.html", context=pagevars)


def loading_complexity_average(job):
    """Used by artist_loading() to calculate how long it should take an artist to
    work on a given job based on how many items it has.
    """
    complexity_average = 0

    # Get the job complexity for this job.
    try:
        job_complexity = JobComplexity.objects.get(job=job)
    except Exception:
        job_complexity = None

    # If we have all the info we need, proceed.
    if job.type and job_complexity:
        job_average_hours = get_item_average_hours(
            job_complexity.category, job.type, job.artist
        )
        # Count the items
        items = Item.objects.filter(job=job)
        # Find the average for this complexity.
        for average in job_average_hours:
            if average[0] == job_complexity.complexity:
                complexity_average = average[1] * items.count()
                break

    return complexity_average


@login_required
def artist_loading(request):
    """Artist Loading Manager's Tool."""
    # Establish next 5 business days (plus extra to see past weekends weekends)
    day1 = date.today()  # Today
    day2 = day1 + timedelta(days=1)  # Tomorrow
    day3 = day2 + timedelta(days=1)
    day4 = day3 + timedelta(days=1)
    day5 = day4 + timedelta(days=1)
    day6 = day5 + timedelta(days=1)
    day7 = day6 + timedelta(days=1)
    day8 = day7 + timedelta(days=1)

    # Gather active Clemson employees.
    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        if user.has_perms(["accounts.clemson_employee"]):
            clemson_employee_list.append(user)

    # How many hours we think an artitst should be able to work in a given day.
    hours_per_day = 8

    # Store data as a lists of lists. Each entry will be an artist's data.
    daily_data = []
    timespan_data = []

    """
    Jobs due. Identiacal logic to search_views.job_todo_list().
    """
    # Search for jobs due with 8 day period
    # Query once, check each for completion, then sort.
    excluded_prepress_suppliers = [
        "Phototype",
        "PHT",
        "SGS",
        "SHK",
        "Schawk",
        "Southern Graphics",
    ]
    jobs_due_range = (
        Job.objects.filter(
            real_due_date__range=(day1, day8),
            status__in=["Active", "Pending"],
        )
        .exclude(prepress_supplier__in=excluded_prepress_suppliers)
        .order_by("workflow")
    )

    incomplete_jobs = []
    for job in jobs_due_range:
        complete = True
        for item in job.get_item_qset():
            if complete:
                if item.first_proof_date() or item.overdue_exempt:
                    # Yup, still complete (True)
                    pass
                else:
                    complete = False
                    incomplete_jobs.append(job)

    jobs_due_1 = []
    jobs_due_2 = []
    jobs_due_3 = []
    jobs_due_4 = []
    jobs_due_5 = []
    jobs_due_6 = []
    jobs_due_7 = []
    jobs_due_8 = []

    # Take query, append to appropriate list for each due_date. Overdue
    # jobs count towards today's total.
    for job in incomplete_jobs:
        if job.real_due_date == day1:
            jobs_due_1.append(job)
        elif job.real_due_date == day2:
            jobs_due_2.append(job)
        elif job.real_due_date == day3:
            jobs_due_3.append(job)
        elif job.real_due_date == day4:
            jobs_due_4.append(job)
        elif job.real_due_date == day5:
            jobs_due_5.append(job)
        elif job.real_due_date == day6:
            jobs_due_6.append(job)
        elif job.real_due_date == day7:
            jobs_due_7.append(job)
        elif job.real_due_date == day8:
            jobs_due_8.append(job)

    """
    Overdue jobs. Identiacal logic to search_views.job_todo_list(). Will be
    counted in day 1 totals.
    """
    start_range = day1 + timedelta(days=-60)
    yesterday = day1 + timedelta(days=-1)
    activejobs_pastdate = (
        Job.objects.filter(
            real_due_date__range=(start_range, yesterday),
            status__in=["Active", "Pending"],
        )
        .exclude(
            prepress_supplier__in=[
                "Phototype",
                "PHT",
                "SGS",
                "SHK",
                "Schawk",
                "Southern Graphics",
            ]
        )
        .order_by("-id")
        .select_related()
    )

    # Maybe query all items in created in last X days w/o final file date?
    overdue_jobs = []
    for job in activejobs_pastdate:
        overdue = False
        for item in job.get_item_qset():
            if not overdue:
                if item.first_proof_date() or item.overdue_exempt:
                    pass
                else:
                    overdue = True
                    overdue_jobs.append(job)

    """
    Revisions due. Identiacal logic to search_views.job_todo_list().
    """
    revisions = Revision.objects.filter(
        due_date__range=(day1, day8), complete_date__isnull=True
    ).exclude(item__job__status__in=["Hold", "Cancelled"])

    revisions_due_1 = []
    revisions_due_2 = []
    revisions_due_3 = []
    revisions_due_4 = []
    revisions_due_5 = []
    revisions_due_6 = []
    revisions_due_7 = []
    revisions_due_8 = []

    # Take query, append to appropriate tuple for each due_date. Overdue
    # revisions count towards today's total.
    for rev in revisions:
        if rev.due_date <= day1:
            if rev.item.job not in revisions_due_1:
                revisions_due_1.append(rev.item.job)
        elif rev.due_date == day2:
            if rev.item.job not in revisions_due_2:
                revisions_due_2.append(rev.item.job)
        elif rev.due_date == day3:
            if rev.item.job not in revisions_due_3:
                revisions_due_3.append(rev.item.job)
        elif rev.due_date == day4:
            if rev.item.job not in revisions_due_4:
                revisions_due_4.append(rev.item.job)
        elif rev.due_date == day5:
            if rev.item.job not in revisions_due_5:
                revisions_due_5.append(rev.item.job)
        elif rev.due_date == day6:
            if rev.item.job not in revisions_due_6:
                revisions_due_6.append(rev.item.job)
        elif rev.due_date == day7:
            if rev.item.job not in revisions_due_7:
                revisions_due_7.append(rev.item.job)
        elif rev.due_date == day8:
            if rev.item.job not in revisions_due_8:
                revisions_due_8.append(rev.item.job)

    """
    File outs due. Identiacal logic to search_views.job_todo_list().
    """
    nine_digit_start_range = day1 + timedelta(days=-180)
    items_to_file_out = (
        Item.objects.filter(item_status="File Out")
        .exclude(fsb_nine_digit_date__lte=nine_digit_start_range)
        .exclude(fsb_nine_digit="")
        .values("job__id")
        .query
    )
    jobs_needing_file_out = (
        Job.objects.filter(id__in=items_to_file_out)
        .exclude(status__in=("Hold", "Cancelled", "Complete"))
        .order_by("-id")
    )
    jobs_needing_file_out_day1 = []
    jobs_needing_file_out_day2 = []
    jobs_needing_file_out_day3 = []
    jobs_needing_file_out_day4 = []
    jobs_needing_file_out_day5 = []
    jobs_needing_file_out_day6 = []
    jobs_needing_file_out_day7 = []
    jobs_needing_file_out_day8 = []

    # Take query, append to appropriate tuple for each due_date. Overdue
    # file outs count towards today's total.
    for job in jobs_needing_file_out:
        if job.final_file_due_date():
            if job.final_file_due_date() <= day1:
                jobs_needing_file_out_day1.append(job)
            elif job.final_file_due_date() == day2:
                jobs_needing_file_out_day2.append(job)
            elif job.final_file_due_date() == day3:
                jobs_needing_file_out_day3.append(job)
            elif job.final_file_due_date() == day4:
                jobs_needing_file_out_day4.append(job)
            elif job.final_file_due_date() == day5:
                jobs_needing_file_out_day5.append(job)
            elif job.final_file_due_date() == day6:
                jobs_needing_file_out_day6.append(job)
            elif job.final_file_due_date() == day7:
                jobs_needing_file_out_day7.append(job)
            elif job.final_file_due_date() == day8:
                jobs_needing_file_out_day8.append(job)

    """
    Sort the jobs due, revisions due, and file outs due by artist.
    """
    for artist in clemson_employee_list:
        artist_daily_data = []
        artist_timespan_data = []
        name = str(artist.first_name + " " + artist.last_name)
        artist_daily_data.append(name)
        artist_timespan_data.append(name)
        # Calculate average loading for each day.

        # Used to total loading hours for the timespan data
        running_total = 0

        # Day 1
        day1_loading = 0
        # Jobs due
        for job in jobs_due_1:
            if job.artist == artist:
                day1_loading += loading_complexity_average(job)
        # Overdue jobs (Day 1 only)
        for job in overdue_jobs:
            if job.artist == artist:
                day1_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_1:
            if job.artist == artist:
                day1_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day1:
            if job.artist == artist:
                day1_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day1_loading, 2), hours_per_day))
        running_total += day1_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day)
        )

        # Day 2
        day2_loading = 0
        # Jobs due
        for job in jobs_due_2:
            if job.artist == artist:
                day2_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_2:
            if job.artist == artist:
                day2_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day2:
            if job.artist == artist:
                day2_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day2_loading, 2), hours_per_day))
        running_total += day2_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 2)
        )

        # Day 3
        day3_loading = 0
        # Jobs due
        for job in jobs_due_3:
            if job.artist == artist:
                day3_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_3:
            if job.artist == artist:
                day3_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day3:
            if job.artist == artist:
                day3_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day3_loading, 2), hours_per_day))
        running_total += day3_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 3)
        )

        # Day 4
        day4_loading = 0
        # Jobs due
        for job in jobs_due_4:
            if job.artist == artist:
                day4_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_4:
            if job.artist == artist:
                day4_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day4:
            if job.artist == artist:
                day4_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day4_loading, 2), hours_per_day))
        running_total += day4_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 4)
        )

        # Day 5
        day5_loading = 0
        # Jobs due
        for job in jobs_due_5:
            if job.artist == artist:
                day5_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_5:
            if job.artist == artist:
                day5_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day5:
            if job.artist == artist:
                day5_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day5_loading, 2), hours_per_day))
        running_total += day5_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 5)
        )

        # Day 6
        day6_loading = 0
        # Jobs due
        for job in jobs_due_6:
            if job.artist == artist:
                day6_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_6:
            if job.artist == artist:
                day6_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day6:
            if job.artist == artist:
                day6_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day6_loading, 2), hours_per_day))
        running_total += day6_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 6)
        )

        # Day 7
        day7_loading = 0
        # Jobs due
        for job in jobs_due_7:
            if job.artist == artist:
                day7_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_7:
            if job.artist == artist:
                day7_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day7:
            if job.artist == artist:
                day7_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day7_loading, 2), hours_per_day))
        running_total += day7_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 7)
        )

        # Day 8
        day8_loading = 0
        # Jobs due
        for job in jobs_due_8:
            if job.artist == artist:
                day8_loading += loading_complexity_average(job)
        # Revisions due
        for job in revisions_due_8:
            if job.artist == artist:
                day8_loading += loading_complexity_average(job)
        # File outs due
        for job in jobs_needing_file_out_day8:
            if job.artist == artist:
                day8_loading += job.avg_fileout_time()
        # Append the data to this artist's list.
        artist_daily_data.append("%s of %s" % (round(day8_loading, 2), hours_per_day))
        running_total += day8_loading
        artist_timespan_data.append(
            "%s of %s" % (round(running_total, 2), hours_per_day * 8)
        )

        # Append this artits data to the master data list.
        daily_data.append(artist_daily_data)
        timespan_data.append(artist_timespan_data)

    pagevars = {
        "page_title": "Artist Loading Report",
        "daily_data": daily_data,
        "timespan_data": timespan_data,
        "hours_per_day": hours_per_day,
        "day1": day1,
        "day2": day2,
        "day3": day3,
        "day4": day4,
        "day5": day5,
        "day6": day6,
        "day7": day7,
        "day8": day8,
    }

    return render(request, "manager_tools/loading.html", context=pagevars)


# @login_required
def monthly_billing(request):
    """Cost avoidance manager's report."""
    # Submitted form.
    if request.POST:
        billing_form = MonthlyBillingForm(request.POST)
        if billing_form.is_valid():
            year = int(billing_form.cleaned_data.get("year"))

    # Default to current year.
    else:
        year = date.today().year
        billing_form = MonthlyBillingForm(
            initial={
                "year": year,
            }
        )

    # Create a list for the data.
    billing_data = []

    # Need set where our month range ends.
    # End it at the current month if displaying the current year.
    if year == date.today().year:
        end_month = date.today().month
        end_month += 1
    # End it after December for all prior years.
    else:
        end_month = 13

    # Calculate the cost avoidance ammount for each month in the year.
    for month in range(1, end_month):
        data = billing_by_month(month, year)
        billing_data.append(data)
    report = billing_funcs.generate_monthly_billing_report_xlsx(month, year)
    pagevars = {
        "page_title": "Monthly Billing Report",
        "costform": billing_form,
        "costdata": billing_data,
        "report": report,
        "year": year,
    }

    return render(request, "manager_tools/monthly_billing.html", context=pagevars)


@login_required
def costavoidance(request):
    """Cost avoidance manager's report."""
    # Submitted form.
    if request.POST:
        costform = CostAvoidanceForm(request.POST)
        if costform.is_valid():
            year = int(costform.cleaned_data.get("year"))

    # Default to current year.
    else:
        year = date.today().year
        costform = CostAvoidanceForm(
            initial={
                "year": year,
            }
        )

    # Create a list for the data.
    costdata = []

    # Need set where our month range ends.
    # End it at the current month if displaying the current year.
    if year == date.today().year:
        end_month = date.today().month
        end_month += 1
    # End it after December for all prior years.
    else:
        end_month = 13

    # Calculate the cost avoidance ammount for each month in the year.
    for month in range(1, end_month):
        data = costavoidance_by_month(month, year)
        costdata.append(data)

    pagevars = {
        "page_title": "Cost Avoidance Report",
        "costform": costform,
        "costdata": costdata,
        "year": year,
    }

    return render(request, "manager_tools/costavoidance.html", context=pagevars)


@login_required
def jobcategory_all(request):
    """Displays the average amount of time artists put into a given type of job
    based on time sheet data. All the data is displayed at once.
    """
    # Each element in this list will be a row in the table.
    data_rows = []

    # First row in the table lists the types so it's separate.
    first_row = []
    first_row.append("BLANK")  # Used to nudge everything over one cell.
    for type in JOB_TYPES:
        first_row.append(type[1])

    # Iterate through each category and type. Each category is a row and each
    # type is a column in that row.
    for category in COMPLEXITY_CATEGORIES:
        data_row = []
        data_row.append(category[1])  # Used to label the rows.
        for type in JOB_TYPES:
            # Add the data to this column.
            data = get_job_average_hours(category[0], type[0])
            data_row.append(data)
        # Now that the row is filled out add it to the list.
        data_rows.append(data_row)

    pagevars = {
        "page_title": "Job Category Report",
        "first_row": first_row,
        "data_rows": data_rows,
    }

    return render(request, "manager_tools/jobcategory_all.html", context=pagevars)


@login_required
def jobcategory_artists(request, supplied_category, supplied_type):
    """This is a supplemental report for jobcategory_all that breaks things down
    by artist.
    """
    # Gather up all Clemson employees.
    clemson_perm = Permission.objects.get(codename="clemson_employee")
    users = (
        User.objects.filter(
            Q(groups__permissions=clemson_perm) | Q(user_permissions=clemson_perm),
            is_active=True,
        )
        .distinct()
        .order_by("last_name")
    )

    # Get the display names for the category and type
    display_category = None
    display_type = None
    for category in COMPLEXITY_CATEGORIES:
        if category[0] == supplied_category:
            display_category = category[1]
    for type in JOB_TYPES:
        if type[0] == supplied_type:
            display_type = type[1]

    # Create a list for user metrics.
    averages_by_artist = []

    # Make the first entry the average for all artists again.
    overall_data = get_job_average_hours(supplied_category, supplied_type)
    averages_by_artist.append(
        {
            "name": "All Artists",
            "data": overall_data,
        }
    )

    # Iterate through the users and gather data.
    for user in users:
        data = get_job_average_hours(supplied_category, supplied_type, user)
        # If the averages are all 0 then don't bother.
        average_total = 0
        for average in data:
            # Skip the partial URL at the start of the data.
            if average[0] == "URL":
                pass
            else:
                average_total += average[1]
        # Pack this user's data up into a dictionary and then append it to the list.
        if average_total > 0:
            averages_by_artist.append(
                {
                    "name": str(user.first_name + " " + user.last_name),
                    "data": data,
                }
            )

    pagevars = {
        "page_title": "%s / %s" % (display_category, display_type),
        "averages_by_artist": averages_by_artist,
        "display_category": display_category,
        "display_type": display_type,
    }

    return render(request, "manager_tools/jobcategory_artists.html", context=pagevars)


def overview(request):
    """Overview Manager's Tool.
    Gather information about employee's
    vacation usage, sick days, performance data, etc...
    """
    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        if (
            user.has_perms(["accounts.clemson_employee"])
            and user.username != "James_Baxter"
        ):
            clemson_employee_list.append(user)

    totalPercentage = group_percentage()
    errorPercentage = error_percentage()
    pagevars = {
        "page_title": "Overview Tool",
        "clemson_employee_list": clemson_employee_list,
        "totalPercentage": totalPercentage,
        "totalErrorPercentage": errorPercentage,
    }

    return render(request, "manager_tools/overview.html", context=pagevars)


class TimesheetsForm(forms.Form):
    """Form used to display the timesheets manager tool for a given month and year."""

    # Month choice range.
    month_choices = []
    for month_num in range(1, 13):
        month_choices.append((month_num, calendar.month_name[month_num]))

    # Year choice range.
    year_choices = []
    year_range = list(range(2018, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    month = forms.ChoiceField(choices=month_choices)
    year = forms.ChoiceField(choices=year_choices)


class MaterialsForm(forms.Form):
    """Form used to display the materials manager tool for a given year."""

    # Year choice range.
    year_choices = []
    year_range = list(range(2020, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    year = forms.ChoiceField(choices=year_choices)


@login_required
def timesheets(request):
    """Timesheet Manager's Tool.
    This is the default view that lists the employees and how many hours they
    spent on each type of activity this month. User can also select a different
    month and year.
    """
    # Default to this month and year.
    if request.POST:
        timesheetform = TimesheetsForm(request.POST)
        month = int(request.POST["month"])
        year = int(request.POST["year"])
    else:
        month = date.today().month
        year = date.today().year
        timesheetform = TimesheetsForm(initial={"month": month, "year": year})

    # Currently available time sheet activities.
    categories = TimeSheetCategory.objects.all().order_by("order")

    # Make a list of just the names for display.
    category_display_list = []
    for cat in categories:
        category_display_list.append(str(cat.name))

    # Add a "Total" category at the end.
    category_display_list.append("Total*")

    # Gather the monthly time sheet data.
    last_day_of_month = calendar.monthrange(year, month)[1]
    start_date = date(year=year, month=month, day=1)
    end_date = date(year=year, month=month, day=last_day_of_month)
    month_data = timesheets_by_date_span(start_date, end_date)

    # Gather the billable hours for the month.
    billable_data = timesheets_billable_hours(month, year)

    # Gather time sheet data for every week in this month.
    weekly_headers = []
    weekly_data = []
    week_num = 1
    month_as_weeks = get_list_of_weeks(month, year)
    for week in month_as_weeks:
        # Generate the table headers
        header = "Week %s (%s - %s)" % (week_num, week[0], week[-1])
        weekly_headers.append(header)
        week_num += 1
        # Gather the data.
        start_date = date(year=year, month=month, day=week[0])
        end_date = date(year=year, month=month, day=week[-1])
        data_this_week = timesheets_by_date_span(start_date, end_date)
        # Check if there was actually any data that week.
        if len(data_this_week) < 2:
            data_this_week = None
        weekly_data.append(data_this_week)

    # Combine the header list and the data list.
    weekly_data_w_headers = zip(weekly_headers, weekly_data)

    pagevars = {
        "page_title": "Timesheets Report",
        "timesheetform": timesheetform,
        "category_display_list": category_display_list,
        "month_data": month_data,
        "billable_data": billable_data,
        "weekly_data_w_headers": weekly_data_w_headers,
    }

    return render(request, "manager_tools/timesheets.html", context=pagevars)


def timesheets_by_date_span(start_date, end_date):
    """Generates the data for the monthly and weekly tables in the time sheets
    report. Could be used to get data for any span of time.
    """
    # Gather active Clemson employees.
    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        if user.has_perms(["accounts.clemson_employee"]):
            clemson_employee_list.append(user)

    # Create a list for user metrics.
    timesheets_by_user = []

    # Currently available time sheet activities.
    categories = TimeSheetCategory.objects.all().order_by("order")

    # Create a dictionary of categories to track total hours in each.
    total_hours_by_category = {}
    for cat in categories:
        total_hours_by_category[str(cat.name)] = 0

    # Some activities should not be counted in the grand total of hours.
    excluded_activities = ["Break", "Lunch"]

    # Create a master list of employees and accompanying data.
    # Sorry, this is going to be a confusing list-of-lists situation.
    employee_data_list = []

    # Gather data for each employee and append it to the list.
    for user in clemson_employee_list:
        # Create a list of this employee's hours in each category.
        # We'll add (category, data) tuples to this list.
        cat_hours = []

        # Used to track all the users hours across categories.
        grand_total_hours = 0

        # Add the user's name and user id to the start of the list.
        cat_hours.append(("Name", str(user.first_name + " " + user.last_name), user.id))
        # Go through all the timesheet categories and total this user's hours in each.
        for category in categories:
            hours_total = 0
            # Gather all the timesheet entries for the user in this category.
            entries = TimeSheet.objects.filter(
                artist=user, date__gte=start_date, date__lte=end_date, category=category
            )
            # Total up the hours.
            for entry in entries:
                hours_total += entry.hours
            # Add this category and the total hours to the list.
            cat_hours.append((str(category.name), hours_total))
            # Also increase this caterory's total in the total_hours_by_category dict.
            total_hours_by_category[str(category.name)] += hours_total
            # Add to the employees grand total.
            if category.name not in excluded_activities:
                grand_total_hours += hours_total

        # Add an extra tuple at the end for total hours in all categories.
        cat_hours.append((str("Total"), grand_total_hours))

        # Put this list of data into the main list.
        if grand_total_hours > 0:
            employee_data_list.append(cat_hours)

    # Add one more entry in the data list for the totals in each category.
    cat_hours = []
    # This isn't a user but let's ape the info we put in for users
    cat_hours.append(("Name", "Total", None))
    for category in categories:
        # Get the total hours for this category.
        hours_total = total_hours_by_category.get(category.name)
        cat_hours.append((str(category.name), hours_total))
    # Put this list of data into the main list.
    employee_data_list.append(cat_hours)

    # Return all the data we've collected.
    return employee_data_list


def timesheets_billable_hours(month, year):
    """Generates the data for the billable hours table in the timesheets report."""
    # Gather active Clemson employees.
    active_list = User.objects.filter(is_active=True).order_by("last_name")
    clemson_employee_list = []
    for user in active_list:
        if user.has_perms(["accounts.clemson_employee"]):
            clemson_employee_list.append(user)

    # Some activities should not be counted in the grand total of hours.
    excluded_activities = ["Break", "Lunch"]

    # Master list for tracking the employee billable hours.
    billable_hours_data = []

    # Gather data for each employee and append it to the list.
    for user in clemson_employee_list:
        # Used to track all the users hours.
        total_hours = 0
        # Used to track the users billable hours.
        billable_hours = 0

        # Gather all the timesheet entries for the user during the month.
        entries = TimeSheet.objects.filter(
            artist=user, date__month=month, date__year=year
        ).exclude(category__name__in=excluded_activities)

        # Total up all the hours.
        for entry in entries:
            total_hours += entry.hours

        # Total up the billable hours
        billable_entries = entries.filter(job__isnull=False)
        for entry in billable_entries:
            billable_hours += entry.hours

        # Now add the user's hour totals to the master list.
        if total_hours > 0:
            billable_percent = (billable_hours / total_hours) * 100.0
            billable_hours_data.append(
                [
                    str(user.first_name + " " + user.last_name),
                    billable_hours,
                    total_hours,
                    round(billable_percent, 2),
                ]
            )

    # Return all the data we've gathered.
    return billable_hours_data


@login_required
def materials(request):
    """Materials Manager Tool.
    This view is used to show supplies and consumables usage by month. Currently
    just tracks how many proofs we print in a month.
    """
    # Default to this month and year.
    if request.POST:
        materialsform = MaterialsForm(request.POST)
        year = int(request.POST["year"])
    else:
        year = date.today().year
        materialsform = MaterialsForm(initial={"year": year})

    # Set the month range.
    if year == date.today().year:
        month_range = list(range(1, date.today().month + 1))
    else:
        month_range = list(range(1, 12))

    # Gather trackers from the designated year.
    trackers = ProofTracker.objects.filter(creation_date__year=year)

    # Get all the proofers recorded by trackers in the designated year.
    proofers = (
        ProofTracker.objects.filter(creation_date__year=year)
        .values_list("proofer", flat=True)
        .distinct()
    )

    # Header for the data table.
    headers = ["Month"]
    for proofer in proofers:
        # Let's shorten some of the known proofers for display.
        if proofer == "Epson P7000 Cold Cup Proofer":
            proofer = "Cold Cup"
        elif proofer == "New Epson 7900 Hot Cup Proofer":
            proofer = "Hot Cup"
        elif proofer == "Shelbyville Colorkeys to Epson T3270":
            proofer = "Colorkeys"
        elif proofer == "Proofs to Epson 4900 on bare Kraft Board":
            proofer = "Kraft Board"
        headers.append(proofer)

    # Store the data for each on in this list.
    data_by_month = []

    for month in month_range:
        # Each month's data will be an additional list.
        data = []
        # The first entry will be the month abbreviation.
        data.append(calendar.month_abbr[month])
        trackers_this_month = trackers.filter(creation_date__month=month)
        # Each additional entry will be the totals for each proofer.
        for proofer in proofers:
            trackers_for_this_proofer = trackers_this_month.filter(proofer=proofer)
            data.append(trackers_for_this_proofer.count())
        # Append this month's data to the master list.
        data_by_month.append(data)

    pagevars = {
        "page_title": "Materials Report",
        "materialsform": materialsform,
        "data_by_month": data_by_month,
        "headers": headers,
    }

    return render(request, "manager_tools/materials.html", context=pagevars)


def get_list_of_weeks(month, year):
    """Returns a month as a list of weeks and days in that month. Used mostly
    to generate weekly reports
    """
    # Set the week to start on Sunday
    calendar.setfirstweekday(calendar.MONDAY)

    # Get a list of weeks in the month and days in the weeks.
    month_list = calendar.monthcalendar(year, month)

    # Remove the zeros that represent the days not in this month.
    month_list_cleaned = []
    for week in month_list:
        week_list_cleaned = []
        for day in week:
            if day != 0:
                week_list_cleaned.append(day)
        month_list_cleaned.append(week_list_cleaned)

    return month_list_cleaned


def get_items_total(user, date_from, date_to):
    """Used to count the number of items assigned to a given user during a period
    of time.
    """
    total = Item.objects.filter(
        creation_date__range=(date_from, date_to), job__artist=user
    ).count()
    return total


def get_amount_charged(user, date_from, date_to):
    """Used to count the number of errors for a given user during a period of
    time.
    """
    preAug2015_charges = (
        Charge.objects.filter(
            creation_date__range=(date_from, date_to),
            item__job__is_deleted=False,
            item__job__artist=user,
        )
        .exclude(item__job__id=99999)
        .exclude(artist=user)
    )
    preAug2015_charges = preAug2015_charges.exclude(
        description__type="Plates"
    ).aggregate(Sum("amount"))
    if preAug2015_charges["amount__sum"] is None:
        preAug2015_charges["amount__sum"] = 0

    postAug2015_charges = Charge.objects.filter(
        creation_date__range=(date_from, date_to),
        item__job__is_deleted=False,
        artist=user,
    ).exclude(item__job__id=99999)
    postAug2015_charges = postAug2015_charges.exclude(
        description__type="Plates"
    ).aggregate(Sum("amount"))
    if postAug2015_charges["amount__sum"] is None:
        postAug2015_charges["amount__sum"] = 0

    total_charges = (
        preAug2015_charges["amount__sum"] + postAug2015_charges["amount__sum"]
    )
    if total_charges:
        amt_charged = total_charges
    else:
        amt_charged = 0

    return amt_charged


def get_qc_reviews_total(user, date_from, date_to):
    """Used to count the number of QC reviews for a given user during a period of
    time.
    """
    num_qc_reviews = QCResponseDoc.objects.filter(
        reviewer=user, parent__isnull=False, review_date__range=(date_from, date_to)
    )

    # get all qc review items from a responseDOC and add them up
    total_num_qcs_reviews = 0
    for reviews in num_qc_reviews:
        total_num_qcs_reviews = total_num_qcs_reviews + reviews.items.count()

    return total_num_qcs_reviews


def get_proof_out_total(user, date_from, date_to):
    """Used to count the number of items proofed by a given user during a period of
    time.
    """
    proof_out_logs = JobLog.objects.filter(
        event_time__range=(date_from, date_to),
        type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
        user=user,
    )
    return proof_out_logs.count()


def get_preflight_total(user, date_from, date_to):
    """Used to count the number of items preflighted by a given user during a
    period of time.
    """
    preflight_logs = JobLog.objects.filter(
        event_time__range=(date_from, date_to),
        type=JOBLOG_TYPE_ITEM_PREFLIGHT,
        user=user,
    )
    return preflight_logs.count()


def get_revisions_proofed_total(user, date_from, date_to):
    """Used to count the number of revisions proofed out by a given user during a
    period of time.
    """
    revision_logs = JobLog.objects.filter(
        event_time__range=(date_from, date_to),
        type=JOBLOG_TYPE_ITEM_REVISION,
        user=user,
    )
    num_revisions_proofed = 0
    # Go through the revisions.
    for log in revision_logs:
        # Try to find a proof out for this item dated after it was revised.
        try:
            proof_out_after_revision = JobLog.objects.filter(
                event_time__gt=log.event_time,
                item=log.item,
                type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
            )
        except Exception:
            proof_out_after_revision = False
        # If you found one then count it.
        if proof_out_after_revision:
            num_revisions_proofed += 1

    return num_revisions_proofed


def get_final_file_total(user, date_from, date_to):
    """Used to count the number of final files by a given user during a period of
    time.
    """
    final_file_logs = JobLog.objects.filter(
        event_time__range=(date_from, date_to),
        type=JOBLOG_TYPE_ITEM_FILED_OUT,
        user=user,
    )
    return final_file_logs.count()


def metrics_data(date_from, date_to):
    """Gathers metrics data for artists over a given time period. Returns it as a
    list of dictionaries. Each dictionary will contain the user's data. We use
    dictionaries so that we can easily call out just the data we need.
    """
    # Gather up all Clemson employees.
    clemson_perm = Permission.objects.get(codename="clemson_employee")
    users = (
        User.objects.filter(
            Q(groups__permissions=clemson_perm) | Q(user_permissions=clemson_perm),
            is_active=True,
        )
        .distinct()
        .order_by("last_name")
    )

    # Create a list for user metrics.
    metrics_by_user = []

    # Iterate through the users and gather data.
    for user in users:
        num_items = get_items_total(user, date_from, date_to)
        amt_charged = get_amount_charged(user, date_from, date_to)
        total_num_qcs_reviews = get_qc_reviews_total(user, date_from, date_to)
        num_proof_outs = get_proof_out_total(user, date_from, date_to)
        num_preflights = get_preflight_total(user, date_from, date_to)
        num_revisions_proofed = get_revisions_proofed_total(user, date_from, date_to)
        num_final_file = get_final_file_total(user, date_from, date_to)
        # Pack this user's data up into a dictionary and then append it to the list.
        metrics_by_user.append(
            {
                "name": str(user.first_name + " " + user.last_name),
                "num_items": num_items,
                "amt_charged": amt_charged,
                "total_num_qcs_reviews": total_num_qcs_reviews,
                "num_proof_outs": num_proof_outs,
                "num_preflights": num_preflights,
                "num_revisions_proofed": num_revisions_proofed,
                "num_final_file": num_final_file,
            }
        )

    return metrics_by_user


def metrics_excel(
    request,
    year_from="0",
    month_from="0",
    day_from="0",
    year_to="0",
    month_to="0",
    day_to="0",
):
    """Download raw metrics data as an excel spreadsheet."""
    # Set the time period.
    date_from = date(int(year_from), int(month_from), int(day_from))
    date_to = date(int(year_to), int(month_to), int(day_to))

    # Gather up the data on the artists.
    user_data = metrics_data(date_from, date_to)

    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = "Data"

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Name"
    docSheet1.cell(row=1, column=2).value = "Items"
    docSheet1.cell(row=1, column=3).value = "Amt. Charged"
    docSheet1.cell(row=1, column=4).value = "QC Reviews"
    docSheet1.cell(row=1, column=5).value = "Items Proofed"
    docSheet1.cell(row=1, column=6).value = "Items Preflighted"
    docSheet1.cell(row=1, column=7).value = "Revisions Proofed"
    docSheet1.cell(row=1, column=8).value = "Items Final Filed"

    # Used to track which row we're on in the spreadsheet.
    row = 1

    # Write a row in the spreadsheet for each item the data dictionary.
    for data in user_data:
        docSheet1.cell(row=row + 1, column=1).value = data["name"]
        docSheet1.cell(row=row + 1, column=2).value = data["num_items"]
        docSheet1.cell(row=row + 1, column=3).value = data["amt_charged"]
        docSheet1.cell(row=row + 1, column=4).value = data["total_num_qcs_reviews"]
        docSheet1.cell(row=row + 1, column=5).value = data["num_proof_outs"]
        docSheet1.cell(row=row + 1, column=6).value = data["num_preflights"]
        docSheet1.cell(row=row + 1, column=7).value = data["num_revisions_proofed"]
        docSheet1.cell(row=row + 1, column=8).value = data["num_final_file"]

        # Move to the next row
        row += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    filename = date_to.strftime("%m_%d_%Y")
    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")
    response["Content-Disposition"] = (
        'attachment; filename="artist_metrics_%s.xlsx"' % filename
    )
    return response


def metrics_pdf(
    request,
    year_from="0",
    month_from="0",
    day_from="0",
    year_to="0",
    month_to="0",
    day_to="0",
):
    """Generate a PDF report with data and charts."""
    # PDF specific imports.
    from reportlab.graphics import renderPDF
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    from reportlab.graphics.shapes import Drawing
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.platypus.tables import Table, TableStyle
    from svglib.svglib import svg2rlg

    # Set the time period.
    date_from = date(int(year_from), int(month_from), int(day_from))
    date_to = date(int(year_to), int(month_to), int(day_to))

    # Create the HttpResponse object with the appropriate file name.
    response = HttpResponse(content_type="application/pdf")
    filename = date_to.strftime("%m_%d_%Y")
    response["Content-Disposition"] = (
        "attachment; filename=artist_metrics_%s.pdf" % filename
    )

    # Set up the document.
    pagesize = landscape(letter)
    canvas = canvas.Canvas(response, pagesize=pagesize)

    # Set up and draw the page title
    canvas.setFont("Helvetica-Bold", 16)
    title = "Artist Metrics Report from %s to %s" % (
        date_from.strftime("%m/%d/%Y"),
        date_to.strftime("%m/%d/%Y"),
    )
    canvas.drawString(1 * inch, 7 * inch, title)

    # Set up and draw the GPI logo.
    file_path = os.path.join(settings.MEDIA_ROOT, "img/GPI_Black_logo.svg")
    graphic = svg2rlg(file_path)
    logo_drawing = Drawing()
    logo_drawing.add(graphic)
    renderPDF.draw(logo_drawing, canvas, 7.5 * inch, 7 * inch)

    # Gather up the data on the artists.
    user_data = metrics_data(date_from, date_to)

    # Set up a data table for the user data starting with the headers.
    table_data = [
        [
            "Name",
            "Items",
            "Amt. Charged",
            "QC Reviews",
            "Items Proofed",
            "Items Preflighted",
            "Revisions Proofed",
            "Items Final Filed",
        ]
    ]

    # Iterate through the users and add their data to the table data.
    for data in user_data:
        table_data.append(
            [
                data["name"],
                data["num_items"],
                "${:,.2f}".format(data["amt_charged"]),  # Currency format
                data["total_num_qcs_reviews"],
                data["num_proof_outs"],
                data["num_preflights"],
                data["num_revisions_proofed"],
                data["num_final_file"],
            ]
        )

    # Draw the table
    table = Table(table_data)
    table.setStyle(
        TableStyle(
            [
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.black),
                ("BACKGROUND", (0, 0), (7, 0), colors.lightgrey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    table.wrapOn(canvas, pagesize[0], pagesize[1])
    table.drawOn(canvas, 1 * inch, 2.25 * inch)

    # Save this page and start a new one.
    canvas.showPage()

    # These are the charts we need to make. data:chart title
    chart_types = [
        ["num_items", "Items"],
        ["amt_charged", "Amt. Charged"],
        ["total_num_qcs_reviews", "QC Reviews"],
        ["num_proof_outs", "Items Proofed"],
        ["num_preflights", "Items Preflighted"],
        ["num_revisions_proofed", "Revisions Proofed"],
        ["num_final_file", "Items Final Filed"],
    ]

    # Make a chart on a new page for each chart type.
    for type in chart_types:
        # Set up chart data.
        chart_keys = []
        chart_data = []
        for data in reversed(user_data):  # Reversed to keep alphabetical order.
            chart_keys.append(str(data["name"]))
            chart_data.append(data[type[0]])

        # For some reason the chart data needs to be a tuple in a list.
        chart_data = [tuple(chart_data)]

        # Draw a chart.
        drawing = Drawing()
        barchart = HorizontalBarChart()
        barchart.data = chart_data
        barchart.categoryAxis.categoryNames = chart_keys
        # Adjust the chart height based on number of users reported.
        barchart.height = len(chart_keys) * 15
        barchart.width = 7 * inch
        # Change bar colors.
        barchart.bars[0].fillColor = colors.blue

        # Add some bar labels.
        if type[0] == "amt_charged":
            barchart.barLabels.nudge = 25
            barchart.barLabelFormat = "$%.2f"  # Currency formatting.
        else:
            barchart.barLabels.nudge = 10
            barchart.barLabelFormat = "%d"
        drawing.add(barchart)
        # Render the chart at a given x y location.
        y_center = (pagesize[1] / 2) - (barchart.height / 2)
        renderPDF.draw(drawing, canvas, 2 * inch, y_center)

        # Draw the page header again.
        renderPDF.draw(logo_drawing, canvas, 7.5 * inch, 7 * inch)
        canvas.setFont("Helvetica-Bold", 16)
        title = type[1]
        canvas.drawString(1 * inch, 7 * inch, title)

        # Save this page and start a new one.
        canvas.showPage()

    # Save the PDF. All done.
    canvas.save()

    return response


class MetricsForm(forms.Form):
    """Form used to gather employee metrics during a given time span."""

    date_from = DateField(label="Dates")
    date_to = DateField()


def metrics_form(request):
    """Used to gather employee metrics for the artists and then format it for
    display as tables and charts.
    """
    # Process a submitted form.
    if request.GET:
        form = MetricsForm(request.GET)
        if form.is_valid():
            # Read choices from the form.
            date_from = form.cleaned_data.get("date_from", None)
            date_to = form.cleaned_data.get("date_to", None)

            # Gather up the data on the artists.
            user_data = metrics_data(date_from, date_to)

            # We need to split the user data up into individual charts. Set up
            # data tables for each of the charts starting with the headers.
            items_table_data = [["Artist", "Items"]]
            charges_table_data = [["Artist", "Amount"]]
            qc_reviews_table_data = [["Artist", "QCs"]]
            items_proofed_table_data = [["Artist", "Items Proofed"]]
            items_preflighted_table_data = [["Artist", "Items Preflighted"]]
            revisions_proofed_table_data = [["Artist", "Revisions Proofed"]]
            final_filed_table_data = [["Artist", "Final Filed"]]

            # Iterate through the user data and append the appropriate data to
            # each chart's data table.
            for data in user_data:
                items_table_data.append([data["name"], data["num_items"]])
                charges_table_data.append([data["name"], data["amt_charged"]])
                qc_reviews_table_data.append(
                    [data["name"], data["total_num_qcs_reviews"]]
                )
                items_proofed_table_data.append([data["name"], data["num_proof_outs"]])
                items_preflighted_table_data.append(
                    [data["name"], data["num_preflights"]]
                )
                revisions_proofed_table_data.append(
                    [data["name"], data["num_revisions_proofed"]]
                )
                final_filed_table_data.append([data["name"], data["num_final_file"]])

            # Send the data to the results page.
            pagevars = {
                "page_title": "Artist Metrics Report",
                "user_data": user_data,
                "items_table_data": items_table_data,
                "charges_table_data": charges_table_data,
                "qc_reviews_table_data": qc_reviews_table_data,
                "items_proofed_table_data": items_proofed_table_data,
                "items_preflighted_table_data": items_preflighted_table_data,
                "revisions_proofed_table_data": revisions_proofed_table_data,
                "final_filed_table_data": final_filed_table_data,
                "date_from": date_from,
                "date_to": date_to,
            }

            return render(
                request, "manager_tools/metrics_results.html", context=pagevars
            )

    # Display an empty form.
    else:
        pagevars = {
            "page_title": "Employee Metrics Report",
            "form": MetricsForm(),
        }
        return render(request, "manager_tools/metrics_form.html", context=pagevars)


def group_percentage():
    """Used to calculate the current year to date percentage of On-time jobs"""
    Item = ContentType.objects.get(app_label="workflow", model="item").model_class()
    year_num = date.today().year
    num_assigned = Item.objects.filter(creation_date__year=year_num).count()

    num_overdue_items = (
        Item.objects.filter(
            creation_date__year=year_num,
            job__status__in=("Active", "Complete"),
            job__is_deleted=False,
        )
        .exclude(overdue_exempt=True)
        .exclude(job__id=99999)
    )

    overdue = []
    for i in num_overdue_items:
        proof = i.first_proof_date()
        # If it was never proofed, it was probably a cancelled item.
        if proof:
            # Foodservice jobs need to finish beofre the due date.
            if i.job.workflow.name == "Foodservice":
                # Proofed on or after job due date, and added before due date.
                if (
                    proof.date() >= i.job.due_date
                    and i.creation_date.date() < i.job.due_date
                ):
                    overdue.append(i)
            else:
                # Proofed after job due date, and added before due date.
                if (
                    proof.date() > i.job.due_date
                    and i.creation_date.date() < i.job.due_date
                ):
                    overdue.append(i)

    try:
        return float(num_assigned - len(overdue)) / num_assigned * 100.0
    except Exception:
        return 0.0


def error_percentage():
    """Used to calculate the current year to date percentage of errors committed on assigned jobs"""
    Item = ContentType.objects.get(app_label="workflow", model="item").model_class()
    Error = ContentType.objects.get(
        app_label="error_tracking", model="error"
    ).model_class()
    year_num = date.today().year

    num_items_assigned = Item.objects.filter(creation_date__year=year_num).count()
    num_errors = Error.objects.filter(reported_date__year=year_num).count()

    try:
        return (
            1.0 - (float(num_items_assigned - num_errors) / num_items_assigned)
        ) * 100.0
    except Exception:
        return 0.0


class ArtTrackingForm(forms.Form):
    """Form used to display the Art Tracking manager tool for a given month and year."""

    # Month choice range.
    month_choices = []
    for month_num in range(1, 13):
        month_choices.append((month_num, calendar.month_name[month_num]))

    # Year choice range.
    year_choices = []
    year_range = list(range(2018, date.today().year + 1))
    for year in year_range:
        year_choices.append((year, year))

    # Fields
    month = forms.ChoiceField(choices=month_choices, initial=date.today().month)
    year = forms.ChoiceField(choices=year_choices, initial=date.today().year)


@login_required
def artworktracking(request):
    """Used to display metrics about incoming artwork broken down by month and user"""
    # Process a submitted form.
    if request.POST:
        form = ArtTrackingForm(request.POST)
        if form.is_valid():
            # Read choices from the form.
            month = form.cleaned_data.get("month", None)
            year = form.cleaned_data.get("year", None)
            month = int(month)
            year = int(year)

            # Gather the monthly time sheet data.
            last_day_of_month = calendar.monthrange(year, month)[1]
            start_date = date(year=year, month=month, day=1)
            end_date = date(year=year, month=month, day=last_day_of_month)
            totals_by_salesperson, totals_by_type = artwork_tracking_by_date(
                start_date, end_date
            )

            welcome_message = None

    # Show a blank page without data at first.
    else:
        form = ArtTrackingForm()
        totals_by_salesperson = None
        totals_by_type = None
        welcome_message = (
            "Set the month and year then hit the green refresh arrow to see data."
        )

    pagevars = {
        "page_title": "Artwork Tracking Report",
        "welcome_message": welcome_message,
        "form": form,
        "totals_by_type": totals_by_type,
        "totals_by_salesperson": totals_by_salesperson,
    }
    return render(request, "manager_tools/artwork_tracking.html", context=pagevars)


def artwork_tracking_by_date(start_date, end_date):
    all_art_trackers = ItemTracker.objects.filter(
        item__is_deleted=False,
        item__workflow__name="Foodservice",
        item__creation_date__range=(start_date, end_date),
        type__category__name="Artwork",
        removal_date__isnull=True,
    )

    """
    Since we need each sales person's artwork grouped by category and amount we're going
    to store them in a dictionary. For each entry the key will be the sales person's
    user object and the value will be yet another dictionary. In this artwork categroy
    dictionary the key will be the status of the artwork and the number of occurences
    across all received artwork for the specified time period

    It will look like this:

        example_data = {
          <User: Mike_Turner>: {"Bad Art": 0, "On Template":2, "Not On Template": 6},
          <User: Dana_Warf>: {"Bad Art": 4, "On Template":0, "Not On Template": 3}
        }

    """

    # totals_by_type dictionary is for all trackers for ItemTracker Type "Artwork"
    totals_by_type = {}
    # totals_by_salesperson dictionary is for all Trackers by Salesperson
    totals_by_salesperson = {}

    # Iterate through the artwork trackers and total their amounts
    for art_tracker in all_art_trackers:
        # get the type of artwork tracker and start counting
        type = art_tracker.type
        # Add the art type and increment occurence
        try:
            # try adding 1
            totals_by_type[type] += 1
        except Exception:
            # create the dictionary key and value starting at 1
            totals_by_type[type] = 1

    # Iterate through the stale charges and record them in the data dictionary.
    for art_tracker in all_art_trackers:
        # Get the sales person for this charge.
        sales_person = art_tracker.item.job.salesperson
        # add the salesperson as a key to the top level dictionary
        try:
            totals_by_salesperson[sales_person]
        except Exception:
            # create the salesperson key in the main dict
            totals_by_salesperson[sales_person] = {}
            for type in totals_by_type.keys():
                # blank out the salespersons trackers to 0 so we can incremement only the ones we want down below, but
                # have a record of all types so the grid will displaty evenly in the template view
                totals_by_salesperson[sales_person][type] = 0

        # get the type of artwork tracker and start counting per salesperson
        type = art_tracker.type
        # Add the art type and increment occurence in the second level dictionary for that salesperson key
        try:
            # try adding 1
            totals_by_salesperson[sales_person][type] += 1
        except Exception:
            # create the dictionary key and value starting at 1
            totals_by_salesperson[sales_person][type] = 1

    return totals_by_salesperson, totals_by_type


def artwork_excel(request, month, year, spreadsheetType):
    """Download raw metrics data as an excel spreadsheet."""
    month = int(month)
    year = int(year)

    # Gather the monthly time sheet data.
    last_day_of_month = calendar.monthrange(year, month)[1]
    start_date = date(year=year, month=month, day=1)
    end_date = date(year=year, month=month, day=last_day_of_month)
    totals_by_salesperson, totals_by_type = artwork_tracking_by_date(
        start_date, end_date
    )

    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = "Artwork"

    if spreadsheetType == "sales":
        # Write the title at the top of the sheet.
        title_text = "Incoming artwork by salesperson from (%s/%s)" % (month, year)
        docSheet1.cell(row=1, column=2).value = title_text
        # create a blank space on row 2 column 1 for layout
        docSheet1.cell(row=2, column=1).value = ""
        col = 2
        # starting from row2 column 2, add all the titles for the artwork categories
        for type, total in totals_by_type.items():
            docSheet1.cell(row=2, column=col).value = type.name
            col += 1

        row = 2
        # Write a section for each sales person in the list of data.
        for sales_person, artwork in totals_by_salesperson.items():
            if sales_person:
                docSheet1.cell(row=row + 1, column=1).value = "%s %s" % (
                    sales_person.first_name,
                    sales_person.last_name,
                )
            else:
                docSheet1.cell(row=row + 1, column=1).value = "No sales person"
            # Write a row for each artwork total
            otherCol = 2
            for type, total in artwork.items():
                docSheet1.cell(row=row + 1, column=otherCol).value = total
                otherCol += 1
            row += 1
    else:
        # Write the title at the top of the sheet.
        title_text = "Incoming artwork by totals from (%s/%s)" % (month, year)
        docSheet1.cell(row=1, column=1).value = title_text
        # create headers for column 1 and 2
        docSheet1.cell(row=2, column=1).value = "Type"
        docSheet1.cell(row=2, column=2).value = "Totals"

        row = 2
        # Write our the type and total of each tracker
        for type, total in totals_by_type.items():
            docSheet1.cell(row=row + 1, column=1).value = type.name
            docSheet1.cell(row=row + 1, column=2).value = total
            row += 1

    # Save the spreadsheet.
    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")
    response["Content-Disposition"] = (
        'attachment; filename="Artwork_Tracking_%s_%s.xlsx"' % (month, year)
    )
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response
