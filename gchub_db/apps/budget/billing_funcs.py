"""General billing & budget functions"""

import calendar
import os
from datetime import date, timedelta

import openpyxl
from django.conf import settings
from django.core.mail import EmailMessage
from django.db.models import Sum
from django.template import loader
from openpyxl.styles import NamedStyle

from gchub_db.apps.budget import billing_funcs
from gchub_db.apps.budget.models import Budget
from gchub_db.apps.joblog import app_defs as joblog_defs
from gchub_db.apps.joblog.models import JobLog
from gchub_db.apps.workflow.models import Charge


def generate_monthly_billing_report_xlsx(month, year):
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()
    # Setup the first sheet to be the summary sheet

    # Set to False for initial run, then to True to update the billing charges
    # as being invoiced.
    UPDATE_INVOICING = False

    # Set your month and year here and only here: (as a number)
    # To be passed as a variable eventually.
    # FORCE_BILLING = 59390
    month_num = month
    year_num = year

    print("Begin Foodservice Invoicing.")
    print("Update as Invoiced: %s" % UPDATE_INVOICING)

    # Set the month name for use in the file name.
    month_name = calendar.month_abbr[month_num]

    date_style = NamedStyle(name="datetime", number_format="DD/MM/YYYY")

    if month_num == 12:
        next_month = 1
        next_year = year_num + 1
    else:
        next_month = month_num + 1
        next_year = year_num

    # This is needed for the invoice updating.
    # Set to the first of the next month.
    end_date = date(next_year, next_month, 1)
    workflow = "Foodservice"

    # Get billable charge qset for workflow.
    billable_charges = billing_funcs.get_billable_data(year_num, month_num, workflow)[
        "charges"
    ]
    # Use this qset if the charges have already been marked as invoiced, and the
    # spreadsheet needs to be recreated.
    # billable_charges = billing_funcs.get_invoiced_data(year_num, month_num, workflow)['charges']

    # memphis cost center variables
    memphisALL = "Memphis All 760279"
    memphisFSB = "Memphis FSB 117409"

    # QADClark = "QAD Migration - Clarksville"
    # QADPitt = "QAD Migration - Pittston"

    # Get list of plants to be invoiced
    plants = []
    items = []

    for charge in billable_charges:
        # Add to the list of items for billing checks.
        if charge.item not in items:
            items.append(charge.item)
        try:
            if charge.item.printlocation.plant.name not in plants:
                if charge.item.printlocation.plant.name == "Memphis":
                    if memphisALL not in plants:
                        plants.append(memphisALL)
                        plants.append(memphisFSB)
                else:
                    plants.append(charge.item.printlocation.plant.name)
        except AttributeError:
            # Use Other to catch items not assigned to a plant.
            if "Other" not in plants:
                plants.append("Other")

    print("---> BEGIN ITEM BILLING WARNINGS.")
    warnings = []
    warning_fields = (
        "Job",
        "Item",
        "#",
        "Artist",
        "Message",
        "Category",
    )
    category = None
    text = None

    for item in items:
        if item.check_too_few_charges():
            category = "Single Charge"
            text = "Warning: only one charge for item"
        if item.check_too_few_revision_charges():
            category = "Missing Revision"
            text = "Warning: missing revision charges for item"
        if item.check_fileout_post_production():
            category = "No Post"
            text = "Warning: no post production charge for item"
        if item.check_prepress_charges():
            category = "No Prepress"
            text = "Warning: no prepress charge for item"
        if item.check_color_keys():
            category = "Color Key"
            text = "Warning: no color keys for item"
        values = (
            item.job,
            item,
            item.num_in_job,
            item.job.artist,
            text,
            category,
        )
        warnings.append(values)
    plant_info = []
    plants.append("Avante-QAD")
    sheetCounter = 0
    for plant in plants:
        # Other is the category for items not assigned to a plant.
        if plant == "Other":
            charge_set = billable_charges.filter(item__printlocation__isnull=True)
        elif plant == memphisALL:
            charge_set = billable_charges.filter(
                item__printlocation__plant__name="Memphis",
                item__printlocation__press__name="All",
            )
        elif plant == memphisFSB:
            charge_set = billable_charges.filter(
                item__printlocation__plant__name="Memphis",
                item__printlocation__press__name="FSB",
            )
        elif plant == "Avante-QAD":
            charge_set = billable_charges.filter(
                item__job__name__startswith="Letica - QAD/Avante"
            )
        else:
            charge_set = billable_charges.filter(
                item__printlocation__plant__name=plant
            ).exclude(item__job__name__startswith="Letica - QAD/Avante")

        if len(charge_set) == 0:
            continue
        # Some feedback as to how many charges per plant there are.
        print(plant, charge_set.count())
        plant_info.append({"plant": plant, "count": charge_set.count()})

        if sheetCounter == 0:
            docSheet1 = workBookDocument.active
            docSheet1.title = "%s Billing" % plant
        else:
            # Create a new sheet for each plant.
            docSheet1 = workBookDocument.create_sheet("%s Billing" % plant)

        # Label column headings
        docSheet1.cell(row=1, column=1).value = "Date Billed"
        docSheet1.cell(row=1, column=2).value = "Salesperson"
        docSheet1.cell(row=1, column=3).value = "Job No."
        docSheet1.cell(row=1, column=4).value = "Job Name"
        docSheet1.cell(row=1, column=5).value = "Size"
        docSheet1.cell(row=1, column=6).value = "Filed Out Date"
        docSheet1.cell(row=1, column=7).value = "Charge Type"
        docSheet1.cell(row=1, column=8).value = "Rush Days"
        docSheet1.cell(row=1, column=9).value = "Plant"
        docSheet1.cell(row=1, column=10).value = "Press"
        docSheet1.cell(row=1, column=11).value = "Amount"

        sheetCounter = sheetCounter + 1

        for i in range(len(charge_set)):
            # Increment rows, write charge data.
            # docSheet1.write(row, column, value)creation_date
            docSheet1.cell(row=i + 2, column=1).value = charge_set[
                i
            ].creation_date.strftime("%m/%d/%y")
            try:
                docSheet1.cell(row=i + 2, column=2).value = str(
                    charge_set[i].item.job.salesperson.username
                )
            except Exception:
                pass
            docSheet1.cell(row=i + 2, column=3).value = str(charge_set[i].item.job.id)
            docSheet1.cell(row=i + 2, column=4).value = str(
                charge_set[i].item.job.name.encode("utf8", "replace")
            )
            docSheet1.cell(row=i + 2, column=5).value = str(charge_set[i].item.size)

            docSheet1.cell(row=i + 2, column=6).value = (
                charge_set[i].item.final_file_date().strftime("%m/%d/%y")
            )
            docSheet1.cell(row=i + 2, column=7).value = str(charge_set[i].description)
            docSheet1.cell(row=i + 2, column=8).value = str(charge_set[i].rush_days)
            try:
                docSheet1.cell(row=i + 2, column=9).value = str(
                    charge_set[i].item.printlocation.plant.name
                )
            except AttributeError:
                docSheet1.cell(row=i + 2, column=9).value = "----"
            try:
                docSheet1.cell(row=i + 2, column=10).value = str(
                    charge_set[i].item.printlocation.press.name
                )
            except AttributeError:
                docSheet1.cell(row=i + 2, column=10).value = "----"
            docSheet1.cell(row=i + 2, column=11).value = charge_set[i].amount

        # Freeze the top row of column headings.
        docSheet1.panes_frozen = docSheet1["B2"]

    if UPDATE_INVOICING:
        # Mark all charges as invoiced today. Set Invoice number to job number?
        for charge in billable_charges:
            # today = datetime.date.today()
            charge.invoice_date = end_date
            charge.invoice_number = str(charge.item.job.id)
            charge.save()

    # Save XLS document
    report_name = "FSB_%s_Billing.xlsx" % month_name
    workBookDocument.save(report_name)

    print("Exported.", report_name)
    mail_body = loader.get_template("emails/monthly_billing.txt")
    mail_context = {
        "report_name": report_name,
        "warnings": warnings,
        "warning_fields": warning_fields,
        "plant_info": plant_info,
    }
    email = EmailMessage(
        "Billing Report",
        mail_body.render(mail_context),
        settings.EMAIL_FROM_ADDRESS,
        ["jacey.r.harris@graphicpkg.com", "Shelly.Congdon@graphicpkg.com"],
    )
    # Attach the file and specify type.
    report_xlsx = os.path.join("gchub_db/media/files/", report_name)
    with open(report_xlsx, "rb") as f:
        data = f.read()
        # Attach the file and specify type.
        email.attach(report_name, data, "application/vnd.ms-excel")

    # Poof goes the mail.
    email.send(fail_silently=False)
    return {
        "filename": report_name,
        "warnings": warnings,
        "warning_fields": warning_fields,
        "charges_count": str(billable_charges.count()),
        "plant_info": plant_info,
    }


def get_items_filed_out(cycle_start, cycle_end, workflow):
    """Return item value list of items filed out in given cycle."""
    items_filed_out = (
        JobLog.objects.filter(
            type=joblog_defs.JOBLOG_TYPE_ITEM_FILED_OUT,
            item__job__workflow__name=workflow,
            event_time__range=(cycle_start, cycle_end),
        )
        .exclude(job__id=99999)
        .values("item__id")
        .query
    )

    return items_filed_out


def get_invoiced_data(year_num, month_num, workflow, plates=False):
    """Return qset and total of charges invoiced for given month."""
    # Set up ranges based on approximate time each workflow is invoiced.
    if workflow == "Foodservice" or workflow == "Container":
        if month_num == 12:
            next_month = 1
            next_year = year_num + 1
        else:
            next_month = month_num + 1
            next_year = year_num

        invoice_start = date(year_num, month_num, 28)
        invoice_end = date(next_year, next_month, 5)
        invoiced_charges = Charge.objects.filter(
            item__workflow__name=workflow,
            invoice_date__range=(invoice_start, invoice_end),
        ).order_by(
            "item__job__temp_printlocation__plant",
            "item__job__id",
            "item__id",
        )

    if workflow == "Beverage":
        invoiced_charges = Charge.objects.filter(
            item__workflow__name=workflow,
            bev_invoice__creation_date__month=month_num,
            bev_invoice__creation_date__year=year_num,
        )

    if not plates:
        invoiced_charges = invoiced_charges.exclude(
            description__type__in=("Plates", "Films")
        )

    # Set invoiced charges to 0 if there are none, rather than returning None
    if invoiced_charges:
        invoiced_total = invoiced_charges.aggregate(total=Sum("amount"))["total"]
    else:
        invoiced_total = 0

    # Build dictionary of billing data.
    monthly_billing_dict = {}
    monthly_billing_dict["charges"] = invoiced_charges
    monthly_billing_dict["total"] = invoiced_total

    return monthly_billing_dict


def get_billable_timeframe(startDate, endDate, workflow, plant=None, plates=False):
    """Return qset of billable charges and their total."""
    # Declare final dictionary here, as some workflows will have extra entries.
    monthly_billing_dict = {}

    if workflow == "Foodservice":
        # Set up dates needed. should come in format "3/14/18"
        dateStartArr = startDate.split("/")
        dateEndArr = endDate.split("/")

        #                            year               month                day
        fsb_cycle_end = date(int(dateEndArr[2]), int(dateEndArr[0]), int(dateEndArr[1]))
        fsb_cycle_start = date(
            int(dateStartArr[2]), int(dateStartArr[0]), int(dateStartArr[1])
        )
        # Charges for items with file out during month (charges to invoice at end of month)
        # Returns items with a File Out action between the specificed date range.
        """
        items_filed_out = get_items_filed_out(fsb_cycle_start,
                                              fsb_cycle_end,
                                              workflow)
        """
        # Using this rather than the standard method to account for jobs that were
        # filed out before the monthly cycle began, but had charges added to it
        # after it was invoiced the first time.
        items_filed_out = (
            JobLog.objects.filter(
                type=joblog_defs.JOBLOG_TYPE_ITEM_FILED_OUT,
                item__job__workflow__name=workflow,
                event_time__range=(fsb_cycle_start, fsb_cycle_end),
            )
            .exclude(job__id=99999)
            .values("item__id")
            .query
        )
        items_filed_out = (
            JobLog.objects.filter(
                type=joblog_defs.JOBLOG_TYPE_ITEM_FILED_OUT,
                item__job__workflow__name=workflow,
                event_time__range=(fsb_cycle_start, fsb_cycle_end),
            )
            .exclude(job__id=99999)
            .values("item__id")
        )

        # Return all uninvoiced charges associated with items in above query.
        """
        Looks like some items got marked as invoiced prematurely during the import.
        Include a invoice_date__isnull = True on the filter for future billings.
        """
        billable_charges = Charge.objects.filter(
            item__in=items_filed_out,
            item__is_deleted=False,
            invoice_date__isnull=True,
        ).order_by("item__printlocation__plant", "item__job__id", "item__id")

        # Total of all charges unbillable that were applied in the last 90 days.
        fsb_potential_start = fsb_cycle_end + timedelta(days=-90)
        potential_charge_total = (
            Charge.objects.filter(
                creation_date__range=(fsb_potential_start, fsb_cycle_end),
                item__is_deleted=False,
                item__job__workflow__name=workflow,
                invoice_date__isnull=True,
            )
            .exclude(item__in=items_filed_out)
            .aggregate(total=Sum("amount"))["total"]
        )
        if not potential_charge_total:
            potential_charge_total = 0
        monthly_billing_dict["potential"] = potential_charge_total

    if workflow == "Beverage":
        # Beverage File Out stuff.
        # Previous Month Data
        if month_num == 1:
            last_month_num = 12
            last_year = year_num - 1
        else:
            last_month_num = month_num - 1
            last_year = year_num

        dateStartArr = startDate.split("/")
        dateEndArr = endDate.split("/")

        # Accomodate the Beverage billing cycle.
        bev_cycle_end = date(year_num, month_num, 21)
        bev_cycle_start = date(last_year, last_month_num, 21)

        # Charges for items with file out during month (charges to invoice at end of month)
        items_filed_out = get_items_filed_out(bev_cycle_start, bev_cycle_end, workflow)

        billable_charges = (
            Charge.objects.filter(
                bev_invoice__isnull=True,
                item__in=items_filed_out,
                item__is_deleted=False,
                invoice_date__isnull=True,
                item__job__prepress_supplier__in=("OPT", "Optihue"),
            )
            .exclude(item__job__status__in=("Cancelled"))
            .order_by(
                "item__job__temp_printlocation__plant",
                "item__job__id",
                "item__id",
            )
        )
        if not plates:
            billable_charges = billable_charges.exclude(
                description__type__in=("Plates", "Films")
            )

        # Total of all charges unbillable that were applied in the last 90 days.
        bev_potential_start = bev_cycle_end + timedelta(days=-90)
        potential_charge_total = (
            Charge.objects.filter(
                creation_date__gte=bev_potential_start,
                item__is_deleted=False,
                bev_invoice__isnull=True,
                item__job__prepress_supplier__in=("OPT", "Optihue"),
                item__job__workflow__name=workflow,
            )
            .exclude(item__in=items_filed_out)
            .exclude(description__type__in=("Plates", "Films"))
            .aggregate(total=Sum("amount"))["total"]
        )
        if not potential_charge_total:
            potential_charge_total = 0
        monthly_billing_dict["potential"] = potential_charge_total

    if workflow == "Container":
        # Set up dates needed.
        if month_num == 12:
            next_month = 1
            next_year = year_num + 1
        else:
            next_month = month_num + 1
            next_year = year_num

        cycle_end = date(next_year, next_month, 1)
        cycle_start = date(year_num, h_num, 1)

        # Charges for items with file out during month (charges to invoice at end of month)
        # Returns items with a File Out action between the specificed date range.
        items_filed_out = get_items_filed_out(cycle_start, cycle_end, workflow)

        # Return all uninvoiced charges associated with items in above query.
        """
        Looks like some items got marked as invoiced prematurely during the import.
        Include a invoice_date__isnull = True on the filter for future billings.
        """
        billable_charges = Charge.objects.filter(
            item__in=items_filed_out,
            item__is_deleted=False,
            invoice_date__isnull=True,
        ).order_by("item__job__id", "item__id")

    # Total up billable charges.
    if billable_charges:
        # Filter further if a plant name is passed.
        if plant:
            billable_charges = billable_charges.filter(
                item__printlocation__plant__name=plant
            )
        billable_total = billable_charges.aggregate(total=Sum("amount"))["total"]
    else:
        billable_total = 0

    # Build dictionary of billing data.
    monthly_billing_dict["charges"] = billable_charges
    monthly_billing_dict["total"] = billable_total

    return monthly_billing_dict


def get_billable_data(year_num, month_num, workflow, plant=None, plates=False):
    """Return qset of billable charges and their total."""
    # Declare final dictionary here, as some workflows will have extra entries.
    monthly_billing_dict = {}

    if workflow == "Foodservice":
        # Set up dates needed.
        if month_num == 12:
            next_month = 1
            next_year = year_num + 1
        else:
            next_month = month_num + 1
            next_year = year_num

        fsb_cycle_end = date(next_year, next_month, 1)
        fsb_cycle_start = date(year_num, month_num, 1)

        # Charges for items with file out during month (charges to invoice at end of month)
        # Returns items with a File Out action between the specificed date range.
        """
        items_filed_out = get_items_filed_out(fsb_cycle_start,
                                              fsb_cycle_end,
                                              workflow)
        """
        # Using this rather than the standard method to account for jobs that were
        # filed out before the monthly cycle began, but had charges added to it
        # after it was invoiced the first time.
        items_filed_out = (
            JobLog.objects.filter(
                type=joblog_defs.JOBLOG_TYPE_ITEM_FILED_OUT,
                item__job__workflow__name=workflow,
                event_time__lte=fsb_cycle_end,
            )
            .exclude(job__id=99999)
            .values("item__id")
            .query
        )

        # Return all uninvoiced charges associated with items in above query.
        """
        Looks like some items got marked as invoiced prematurely during the import.
        Include a invoice_date__isnull = True on the filter for future billings.
        """
        billable_charges = Charge.objects.filter(
            item__in=items_filed_out,
            item__is_deleted=False,
            invoice_date__isnull=True,
        ).order_by("item__printlocation__plant", "item__job__id", "item__id")

        # Total of all charges unbillable that were applied in the last 90 days.
        fsb_potential_start = fsb_cycle_end + timedelta(days=-90)
        potential_charge_total = (
            Charge.objects.filter(
                creation_date__range=(fsb_potential_start, fsb_cycle_end),
                item__is_deleted=False,
                item__job__workflow__name=workflow,
                invoice_date__isnull=True,
            )
            .exclude(item__in=items_filed_out)
            .aggregate(total=Sum("amount"))["total"]
        )
        if not potential_charge_total:
            potential_charge_total = 0
        monthly_billing_dict["potential"] = potential_charge_total

    if workflow == "Beverage":
        # Beverage File Out stuff.
        # Previous Month Data
        if month_num == 1:
            last_month_num = 12
            last_year = year_num - 1
        else:
            last_month_num = month_num - 1
            last_year = year_num

        # Accomodate the Beverage billing cycle.
        bev_cycle_end = date(year_num, month_num, 21)
        bev_cycle_start = date(last_year, last_month_num, 21)

        # Charges for items with file out during month (charges to invoice at end of month)
        items_filed_out = get_items_filed_out(bev_cycle_start, bev_cycle_end, workflow)

        billable_charges = (
            Charge.objects.filter(
                bev_invoice__isnull=True,
                item__in=items_filed_out,
                item__is_deleted=False,
                invoice_date__isnull=True,
                item__job__prepress_supplier__in=("OPT", "Optihue"),
            )
            .exclude(item__job__status__in=("Cancelled"))
            .order_by(
                "item__job__temp_printlocation__plant",
                "item__job__id",
                "item__id",
            )
        )
        if not plates:
            billable_charges = billable_charges.exclude(
                description__type__in=("Plates", "Films")
            )

        # Total of all charges unbillable that were applied in the last 90 days.
        bev_potential_start = bev_cycle_end + timedelta(days=-90)
        potential_charge_total = (
            Charge.objects.filter(
                creation_date__gte=bev_potential_start,
                item__is_deleted=False,
                bev_invoice__isnull=True,
                item__job__prepress_supplier__in=("OPT", "Optihue"),
                item__job__workflow__name=workflow,
            )
            .exclude(item__in=items_filed_out)
            .exclude(description__type__in=("Plates", "Films"))
            .aggregate(total=Sum("amount"))["total"]
        )
        if not potential_charge_total:
            potential_charge_total = 0
        monthly_billing_dict["potential"] = potential_charge_total

    if workflow == "Container":
        # Set up dates needed.
        if month_num == 12:
            next_month = 1
            next_year = year_num + 1
        else:
            next_month = month_num + 1
            next_year = year_num

        cycle_end = date(next_year, next_month, 1)
        cycle_start = date(year_num, h_num, 1)

        # Charges for items with file out during month (charges to invoice at end of month)
        # Returns items with a File Out action between the specificed date range.
        items_filed_out = get_items_filed_out(cycle_start, cycle_end, workflow)

        # Return all uninvoiced charges associated with items in above query.
        """
        Looks like some items got marked as invoiced prematurely during the import.
        Include a invoice_date__isnull = True on the filter for future billings.
        """
        billable_charges = Charge.objects.filter(
            item__in=items_filed_out,
            item__is_deleted=False,
            invoice_date__isnull=True,
        ).order_by("item__job__id", "item__id")

    # Total up billable charges.
    if billable_charges:
        # Filter further if a plant name is passed.
        if plant:
            billable_charges = billable_charges.filter(
                item__printlocation__plant__name=plant
            )
        billable_total = billable_charges.aggregate(total=Sum("amount"))["total"]
    else:
        billable_total = 0

    # Build dictionary of billing data.
    monthly_billing_dict["charges"] = billable_charges
    monthly_billing_dict["total"] = billable_total

    return monthly_billing_dict


def get_billing_activity(year_num, month_num, workflow):
    """Return a total of billing activity (charges billed during this month.)
    This is a measure of workload, not invoicable amounts.
    """
    monthly_charges = Charge.objects.filter(
        item__workflow__name=workflow,
        creation_date__month=month_num,
        creation_date__year=year_num,
    ).exclude(description__type__in=("Plates", "Films"))

    if workflow == "Beverage":
        monthly_charges = monthly_charges.filter(
            item__job__prepress_supplier__in=("OPT", "Optihue")
        )

    billing_activity = {}
    if monthly_charges:
        billing_activity["total"] = monthly_charges.aggregate(total=Sum("amount"))[
            "total"
        ]
        billing_activity["charges"] = monthly_charges
    else:
        billing_activity["total"] = 0
    return billing_activity


def monthly_report(
    month_num,
    year_num,
    type="monthly",
    invoiced=True,
    billable=True,
    activity=True,
    budget=True,
):
    """Returns dictionary of budget reporting information for given month and year.
    In order to facilitate speedier reports, make each portion of the monthly
    optional by assigning True/False to each category. This should greatly reduce
    the number of queries per call.
    """
    monthname = calendar.month_name[month_num]
    report_dict = {
        "month": monthname,
        "year": year_num,
    }

    if invoiced:
        # Get amount invoiced for the month for each workflow.
        fsb_monthly_invoiced = get_invoiced_data(year_num, month_num, "Foodservice")[
            "total"
        ]
        bev_monthly_invoiced = get_invoiced_data(year_num, month_num, "Beverage")[
            "total"
        ]
        # con_monthly_invoiced = get_invoiced_data(year_num, month_num, "Container")['total']
        # Total of 3 workflows.
        hub_monthly_invoiced = fsb_monthly_invoiced + bev_monthly_invoiced

        report_dict["fsb_monthly_invoiced"] = fsb_monthly_invoiced
        report_dict["bev_monthly_invoiced"] = bev_monthly_invoiced
        report_dict["hub_monthly_invoiced"] = hub_monthly_invoiced

    # Set some variables outside of this if statement to make sure other
    # calculations don't fail.
    fsb_billable_total = 0
    bev_billable_total = 0
    hub_billable_total = 0
    if billable:
        # Get billable and potential billable charges for the month for each workflow..
        fsb_billable = get_billable_data(year_num, month_num, "Foodservice")
        fsb_billable_total = fsb_billable["total"]
        fsb_potential = fsb_billable["potential"]
        bev_billable = get_billable_data(year_num, month_num, "Beverage")
        bev_billable_total = bev_billable["total"]
        bev_potential = bev_billable["potential"]
        # con_billable = get_billable_data(year_num, month_num, "Container")
        # con_billable_total = con_billable['total']
        # Total of 3 workflows.
        hub_billable_total = fsb_billable_total + bev_billable_total
        hub_potential = fsb_potential + bev_potential

        report_dict["fsb_potential"] = fsb_potential
        report_dict["bev_potential"] = bev_potential
        report_dict["hub_potential"] = hub_potential
        report_dict["fsb_billable_total"] = fsb_billable_total
        report_dict["bev_billable_total"] = bev_billable_total
        report_dict["hub_billable_total"] = hub_billable_total

    if activity:
        # Retrieve billing activity for month by workflow.
        fsb_activity = get_billing_activity(year_num, month_num, "Foodservice")
        fsb_monthly_billing = fsb_activity["total"]
        bev_monthly_billing = get_billing_activity(year_num, month_num, "Beverage")[
            "total"
        ]
        # con_monthly_billing = get_billing_activity(year_num, month_num, "Container")['total']
        # Total of 3 workflows.
        hub_monthly_billing = fsb_monthly_billing + bev_monthly_billing

        report_dict["fsb_monthly_billing"] = fsb_monthly_billing
        report_dict["bev_monthly_billing"] = bev_monthly_billing
        report_dict["hub_monthly_billing"] = hub_monthly_billing

    if budget:
        # Get budgets for month by workflow.
        try:
            fsb_monthly_budget_grab = Budget.objects.get(
                year=year_num, workflow__name="Foodservice"
            )
            fsb_monthly_budget = getattr(fsb_monthly_budget_grab, monthname)
        except Budget.DoesNotExist:
            fsb_monthly_budget = 0

        try:
            bev_monthly_budget_grab = Budget.objects.get(
                year=year_num, workflow__name="Beverage"
            )
            bev_monthly_budget = getattr(bev_monthly_budget_grab, monthname)
        except Budget.DoesNotExist:
            bev_monthly_budget = 0

        """
        try:
            con_monthly_budget_grab = Budget.objects.get(year=year_num, workflow__name="Container")
            con_monthly_budget = getattr(con_monthly_budget_grab, monthname)
        except Budget.DoesNotExist:
            con_monthly_budget = 0
        """
        # Total hub budget.
        hub_monthly_budget = fsb_monthly_budget + bev_monthly_budget

        if type == "monthly":
            # Calculate budget progress as a percentage.
            if fsb_monthly_budget > 0:
                fsb_budget_prog = (
                    (fsb_monthly_invoiced + fsb_billable_total) / fsb_monthly_budget
                ) * 100
            else:
                fsb_budget_prog = 0

            if bev_monthly_budget > 0:
                bev_budget_prog = (
                    (bev_monthly_invoiced + bev_billable_total) / bev_monthly_budget
                ) * 100
            else:
                bev_budget_prog = 0

            """
            if con_monthly_budget > 0:
                con_budget_prog = ((con_monthly_invoiced + con_billable_total)/con_monthly_budget)*100
            else:
                con_budget_prog = 0
            """

            if hub_monthly_budget > 0:
                hub_budget_prog = (
                    (hub_monthly_invoiced + hub_billable_total) / hub_monthly_budget
                ) * 100
            else:
                hub_budget_prog = 0
        else:
            # Calculate budget progress as a percentage.
            if fsb_monthly_budget > 0:
                fsb_budget_prog = (fsb_monthly_invoiced / fsb_monthly_budget) * 100
            else:
                fsb_budget_prog = 0

            if bev_monthly_budget > 0:
                bev_budget_prog = (bev_monthly_invoiced / bev_monthly_budget) * 100
            else:
                bev_budget_prog = 0

            """
            if con_monthly_budget > 0:
                con_budget_prog = (con_monthly_invoiced/con_monthly_budget)*100
            else:
                con_budget_prog = 0
            """

            if hub_monthly_budget > 0:
                hub_budget_prog = (hub_monthly_invoiced / hub_monthly_budget) * 100
            else:
                hub_budget_prog = 0

        report_dict["fsb_monthly_budget"] = fsb_monthly_budget
        report_dict["bev_monthly_budget"] = bev_monthly_budget
        report_dict["hub_monthly_budget"] = hub_monthly_budget
        report_dict["fsb_budget_prog"] = fsb_budget_prog
        report_dict["bev_budget_prog"] = bev_budget_prog
        report_dict["hub_budget_prog"] = hub_budget_prog

    return report_dict
