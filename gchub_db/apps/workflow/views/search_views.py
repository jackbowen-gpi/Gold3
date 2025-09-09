"""Job and Item search and report views."""

import calendar
import time
from datetime import date, timedelta
from io import BytesIO
from tempfile import NamedTemporaryFile

# openpyxl is now used instead of pyExcelerator because it supports python3 and pyexcelerator is deprecated
import openpyxl
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission, User
from django.contrib.sites.models import Site
from django.db import models
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.template import RequestContext
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic.list import ListView

from gchub_db.apps.art_req.models import AdditionalInfo, ArtReq
from gchub_db.apps.error_tracking.models import Error
from gchub_db.apps.fedexsys.models import Shipment
from gchub_db.apps.joblog.app_defs import *
from gchub_db.apps.joblog.models import JobLog
from gchub_db.apps.manager_tools.manager_tool_funcs import get_item_average_hours
from gchub_db.apps.qad_data.models import QAD_PrintGroups
from gchub_db.apps.workflow import app_defs
from gchub_db.apps.workflow.models import (
    Item,
    ItemColor,
    ItemReview,
    ItemTracker,
    ItemTrackerType,
    Job,
    JobAddress,
    JobComplexity,
    Plant,
    Platemaker,
    Press,
    Revision,
    SpecialMfgConfiguration,
)
from gchub_db.includes import general_funcs
from gchub_db.includes.gold_json import JSMessage


def pending_jobs(request):
    """Simple list of Pending Jobs, to ease in Job Assignment/Distribution"""
    pending_list = Job.objects.filter(status="Pending").order_by("-id")

    pagevars = {"page_title": "Jobs with Status: Pending", "pending_list": pending_list}

    return render(request, "workflow/job/pending_jobs.html", context=pagevars)


class ReviewSearchForm(forms.Form):
    """A small form that that lets users search for old item reviews."""

    plants = ["Kenton", "Shelbyville", "Visalia", "Clarksville", "Pittston"]
    plant_choices = Plant.objects.filter(name__in=plants).order_by("name")
    plant = forms.ModelMultipleChoiceField(
        queryset=plant_choices, required=False, widget=forms.CheckboxSelectMultiple
    )
    status_choices = [
        ("approved", "Approved"),
        ("rejected", "Rejected"),
        ("expired", "Expired"),
    ]
    status = forms.MultipleChoiceField(
        choices=status_choices, required=False, widget=forms.CheckboxSelectMultiple
    )

    def __init__(self, request, *args, **kwargs):
        super(ReviewSearchForm, self).__init__(*args, **kwargs)


def review(request, category):
    """Displays item review objects to be review. Filters reviews into
    market, demand plan, and plant reviews.
    """
    form = ReviewSearchForm(request, request.GET)
    if request.GET and form.is_valid():
        # Call the result view directly for display.
        return ReviewSearchResults.as_view(form=form)(request)

    else:
        # Jobs to go to the main review list.
        review_list = (
            ItemReview.objects.filter(
                review_date__isnull=True,
                review_ok=False,
                review_catagory=category,
            )
            .exclude(item__job=99999)
            .order_by("-item__id")
        )

        # Items reviewed and currently rejected.
        rejected_list = (
            ItemReview.objects.filter(
                review_date__isnull=False,
                review_ok=False,
                review_catagory=category,
            )
            .exclude(comments="Resubmitted")
            .order_by("-item__id")
        )

        # Items reviewed, rejected, altered, then resubmitted.
        resub_list = ItemReview.objects.filter(
            comments="Resubmitted",
            review_date__isnull=False,
            review_ok=False,
            review_catagory=category,
        ).order_by("-item__id")
        # We'll populate this later if this is a plant review. The plants want
        # to see a list of expired reviews in case they missed something.
        expired_list = False

        # We need to calculate the date for three business days ago before we
        # can gather up the reviews that haven't expired.
        if timezone.now().isoweekday() == 7:
            # If it's Sunday skip Saturday.
            three_days_ago = timezone.now() - timedelta(4)
        elif timezone.now().isoweekday() >= 1 and timezone.now().isoweekday() <= 3:
            # If it's between Monday and Wednesday skip the whole weekend.
            three_days_ago = timezone.now() - timedelta(5)
        else:
            # Otherwise just go three days back.
            three_days_ago = timezone.now() - timedelta(3)

        if category == "plant":
            review_list = (
                review_list.filter(review_initiated_date__gt=three_days_ago)
                .exclude(item__printlocation__press__name="Other")
                .exclude(item__press_change=True)
            )
            resub_list = resub_list.filter(
                review_initiated_date__gt=three_days_ago
            ).exclude(item__printlocation__press__name="Other")
            expired_list = (
                ItemReview.objects.filter(
                    review_ok=False,
                    review_date__isnull=True,
                    resubmitted=False,
                    review_catagory=category,
                    review_initiated_date__range=(
                        timezone.now() - timedelta(days=14),
                        three_days_ago,
                    ),
                )
                .exclude(item__job=99999)
                .exclude(item__press_change=True)
                .order_by("-item__id")
            )
            # Figure out which plant reviews this user can see.
            user = request.user
            # Clemson employees see everything. Others get per plant access.
            if not user.has_perms(["accounts.clemson_employee"]):
                # Make a list of plants they have permission to view.
                permitted_plants = []
                if user.has_perm("accounts.kenton_art_approval"):
                    permitted_plants.append("Kenton")
                if user.has_perm("accounts.shelbyville_art_approval"):
                    permitted_plants.append("Shelbyville")
                if user.has_perm("accounts.visalia_art_approval"):
                    permitted_plants.append("Visalia")
                if user.has_perm("accounts.pittston_art_approval"):
                    permitted_plants.append("Pittston")
                if user.has_perm("accounts.clarksville_art_approval"):
                    permitted_plants.append("Clarksville")
                # Filter to only show reviews for the user's permitted plants.
                review_list = review_list.filter(
                    item__printlocation__plant__name__in=permitted_plants
                )
                rejected_list = rejected_list.filter(
                    item__printlocation__plant__name__in=permitted_plants
                )
                resub_list = resub_list.filter(
                    item__printlocation__plant__name__in=permitted_plants
                )
                expired_list = expired_list.filter(
                    item__printlocation__plant__name__in=permitted_plants
                )

        elif category == "demand" or category == "market":
            # These reviews don't use the search form.
            form = False
            if category == "demand":
                review_list = review_list.filter(
                    review_initiated_date__gt=three_days_ago
                )

        else:
            exit()

        # This section handles the Plant Report at the top of the page thats only visible to Clemson Employees
        # this array is for displaying the times in order on the template
        time_arr = ["3", "6", "9", "12"]
        # this array is for associating the times with their query timedelta values
        time_obj = {"3": 90, "6": 180, "9": 270, "12": 365}
        plant_obj = {}
        current_user = request.user
        # if the current user isnt from clemson, dont even do this so load times are not affected
        if current_user.has_perms(["accounts.clemson_employee"]):
            # get all reviews of this year in memory to use for later so we dont query again and again
            all = ItemReview.objects.filter(
                review_catagory="plant",
                review_initiated_date__range=[
                    str(three_days_ago - timedelta(365)),
                    str(three_days_ago),
                ],
            )
            for time in time_obj:
                plant_obj[time] = {}
                for plant in [
                    "Shelbyville",
                    "Kenton",
                    "Visalia",
                    "Clarksville",
                    "Pittston",
                ]:
                    # filter the all query by plant and time for each data point we want to display on the template
                    # All of these queries and filters have come from adjacent functions
                    all_review_list = all.filter(
                        item__printlocation__plant__name=plant,
                        review_catagory="plant",
                        review_initiated_date__range=[
                            str(three_days_ago - timedelta(time_obj[time])),
                            str(three_days_ago),
                        ],
                    ).exclude(
                        Q(item__press_change=True)
                        | Q(item__printlocation__press__name="Other")
                        | Q(item__printlocation__press__name="Corrugated")
                    )
                    review_approved_rejected_list = all_review_list.filter(
                        item__printlocation__plant__name=plant,
                        review_ok__in=[True, False],
                        review_date__isnull=False,
                    )
                    review_expired_list = (
                        all_review_list.filter(
                            item__printlocation__plant__name=plant,
                            review_ok=False,
                            review_date__isnull=True,
                            resubmitted=False,
                        )
                        .exclude(item__job=99999)
                        .exclude(item__press_change=True)
                        .order_by("-item__id")
                    )

                    # percentages are added by dividing the float values of the lengths of the arrays and rounding them to the first decimal
                    plant_obj[time][plant] = {}
                    plant_obj[time][plant]["accepted_rejected"] = len(
                        review_approved_rejected_list
                    )
                    try:
                        plant_obj[time][plant]["accepted_rejected_percent"] = round(
                            float(len(review_approved_rejected_list))
                            / float(len(all_review_list))
                            * 100,
                            1,
                        )
                    except Exception:
                        plant_obj[time][plant]["accepted_rejected_percent"] = 0
                    plant_obj[time][plant]["expired"] = len(review_expired_list)
                    try:
                        plant_obj[time][plant]["expired_percent"] = round(
                            float(len(review_expired_list))
                            / float(len(all_review_list))
                            * 100,
                            1,
                        )
                    except Exception:
                        plant_obj[time][plant]["expired_percent"] = 0

        pagevars = {
            "page_title": "%s Review Summary" % category.capitalize(),
            "review_list": review_list[:30],
            "rejected_list": rejected_list,
            "resub_list": resub_list[:30],
            "expired_list": expired_list,
            "plant_obj": plant_obj,
            "time_arr": time_arr,
            "category": category,
            "form": form,
        }

        return render(request, "workflow/plant_review/review.html", context=pagevars)


class ReviewSearchResults(ListView):
    """Displays plant review search results."""

    # Set up ListView stuff.
    paginate_by = 25
    template_name = "workflow/plant_review/review_search.html"
    # Extra parameter to handle a form object.
    form = None

    # Searching and filtering.
    def get_queryset(self):
        qset = ItemReview.objects.filter(review_catagory="plant")
        # Press reviews used to get generated for things the plant didn't need to
        # check. This will exclude those old reviews so their numbers aren't affected.
        qset = qset.exclude(
            Q(item__press_change=True)
            | Q(item__printlocation__press__name="Other")
            | Q(item__printlocation__press__name="Corrugated")
        )
        if self.form:
            # Filter by plant.
            plants = self.form.cleaned_data["plant"]
            if plants:
                qset = qset.filter(item__printlocation__plant__in=plants)

            # Filter by status.
            statuses = self.form.cleaned_data["status"]

            if statuses:
                # We need to calculate the date for three business days ago before we
                # can gather up the expired reviews. Use UTC-naive now.
                now_dt = general_funcs._utcnow_naive()
                if now_dt.isoweekday() == 7:
                    # If it's Sunday skip Saturday.
                    three_days_ago = now_dt - timedelta(4)
                elif 1 <= now_dt.isoweekday() <= 3:
                    # If it's between Monday and Wednesday skip the whole weekend.
                    three_days_ago = now_dt - timedelta(5)
                else:
                    # Otherwise just go three days back.
                    three_days_ago = now_dt - timedelta(3)

                qset = qset.exclude(
                    review_date=None, review_initiated_date__gt=three_days_ago
                )
                # If a status isn't checked don't return it in the results.
                if "approved" not in statuses:
                    qset = qset.exclude(review_ok=True)
                if "rejected" not in statuses:
                    qset = qset.exclude(review_ok=False, review_date__isnull=False)
                if "expired" not in statuses:
                    qset = qset.exclude(
                        review_initiated_date__lt=three_days_ago,
                        review_date__isnull=True,
                        review_ok=False,
                    )

        # Order by review start date and return.
        return qset.order_by("-review_initiated_date")

    # Set context data.
    def get_context_data(self, **kwargs):
        context = super(ReviewSearchResults, self).get_context_data(**kwargs)
        context["page_title"] = "Item Review Search Results"
        context["extra_link"] = general_funcs.paginate_get_request(self.request)
        return context


def mkt_review(request):
    """Experimental marketing review page. Might be a bad idea."""
    # Items that need to be reviewed by marketing.
    approval_list = (
        ItemReview.objects.filter(
            review_date__isnull=True,
            review_ok=False,
            review_catagory="market",
        )
        .exclude(item__job=99999)
        .order_by("-item__id")
    )

    # Items that were reviewed and rejected.
    rejected_list = (
        ItemReview.objects.filter(
            review_date__isnull=False, review_ok=False, review_catagory="market"
        )
        .exclude(comments="Resubmitted")
        .order_by("-item__id")
    )

    """
    Items reviewed, rejected, altered, then resubmitted. Will be listed under
    the same header as items in the approval list since both will be needing
    approval
    """
    resub_list = ItemReview.objects.filter(
        review_date__isnull=False,
        review_ok=False,
        review_catagory="market",
        comments="Resubmitted",
    ).order_by("-item__id")
    """
    Checks what marketing trackers are in existence. Used to create links to
    tracker specific reports at the top of the marketing review page.
    """
    tracker_types = ItemTrackerType.objects.filter(category__name="Marketing")

    pagevars = {
        "page_title": "Marketing Review Summary",
        "approval_list": approval_list,
        "rejected_list": rejected_list,
        "resub_list": resub_list,
        "tracker_types": tracker_types,
    }

    return render(request, "workflow/plant_review/mkt_review.html", context=pagevars)


class MktReviewReport(ListView):
    """Marketing sure loves their reports. Here's where we make them. Paginated for
    easy browsing.
    """

    paginate_by = 25
    template_name = "workflow/plant_review/mkt_review_lists.html"

    def get_queryset(self, **kwargs):
        type = self.kwargs["type"]
        tracked_art_type = self.kwargs["tracked_art_type"]
        qset = []
        # Lists the items that have tracked art types.
        if type == "trackedart":
            qset = ItemTracker.objects.filter(
                type__id=tracked_art_type, removal_date__isnull=True
            )

        # Lists reviews that have been approvaed but not filed out.
        elif type == "pending":
            qset = (
                ItemReview.objects.filter(
                    review_ok=True,
                    review_catagory="market",
                    item__item_status="Pending",
                )
                .exclude(item__job=99999)
                .order_by("-item__id")
            )
            form = "Dummy form! This is shameful..."

        # Lists reviews that have been approvaed and filed out.
        elif type == "completed":
            qset = (
                ItemReview.objects.filter(
                    review_ok=True,
                    review_catagory="market",
                    item__item_status="Complete",
                )
                .exclude(item__job=99999)
                .order_by("-item__id")
            )
            form = "Dummy form! This is shameful..."

        return qset

    def get_context_data(self, **kwargs):
        context = super(MktReviewReport, self).get_context_data(**kwargs)

        type = self.kwargs["type"]
        tracked_art_type = self.kwargs["tracked_art_type"]
        title = ""
        # Lists the items that have tracked art types.
        if type == "trackedart":
            title = ItemTrackerType.objects.get(id=tracked_art_type)
        # Lists reviews that have been approvaed but not filed out.
        elif type == "pending":
            title = "Pending"
        # Lists reviews that have been approvaed and filed out.
        elif type == "completed":
            title = "Completed"

        context["page_title"] = "%s Report for Marketing Review" % title
        context["type"] = type
        context["tracked_art_type"] = tracked_art_type
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context


def mkt_review_excel(request, tracked_art_type):
    """Creates a spreadsheet listing all the tracked art objects of a given type.
    In other words: if you want a spreadsheet listing all the items with SFI
    logos on them this is how you get it.
    """
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Get all the tracked art objects of the type we're looking for.
    qset = ItemTracker.objects.filter(
        type__id=tracked_art_type, removal_date__isnull=True
    )

    title = ItemTrackerType.objects.get(id=tracked_art_type)
    title = title.name

    docSheet1 = workBookDocument.active
    docSheet1.title = title

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Job Number"
    docSheet1.cell(row=1, column=2).value = "Name"
    docSheet1.cell(row=1, column=3).value = "Printgroup"
    docSheet1.cell(row=1, column=4).value = "Size"
    docSheet1.cell(row=1, column=5).value = "Comments"
    docSheet1.cell(row=1, column=6).value = "Added By"
    docSheet1.cell(row=1, column=7).value = "Date"

    # Used to track which row we're on in the spreadsheet.
    row = 1

    # Write a row in the spreadsheet for each object in the query set.
    for art in qset:
        docSheet1.cell(row=row + 1, column=1).value = art.item.job.id
        docSheet1.cell(row=row + 1, column=2).value = art.item.job.name
        docSheet1.cell(row=row + 1, column=3).value = "Printgroup"
        docSheet1.cell(row=row + 1, column=4).value = str(art.item.size)
        docSheet1.cell(row=row + 1, column=5).value = art.addition_comments
        docSheet1.cell(row=row + 1, column=6).value = str(art.edited_by)
        docSheet1.cell(row=row + 1, column=7).value = str(art.addition_date)

        # Move to the next row
        row += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")

    response["Content-Disposition"] = 'attachment; filename="%s Report.xlsx"' % title
    return response


def items_rejected(request):
    """Display item review objects rejected by the plant, demand planning, or marketing."""
    dp_rejected_list = (
        ItemReview.objects.filter(
            review_catagory="demand",
            review_date__isnull=False,
            review_ok=False,
        )
        .exclude(comments="Resubmitted")
        .exclude(resubmitted=True)
        .order_by("-item__id")
    )

    pr_rejected_list = (
        ItemReview.objects.filter(
            review_catagory="plant", review_date__isnull=False, review_ok=False
        )
        .exclude(comments="Resubmitted")
        .exclude(resubmitted=True)
        .order_by("-item__id")
    )

    mkt_rejected_list = (
        ItemReview.objects.filter(
            review_date__isnull=False,
            review_catagory="market",
            review_ok=False,
        )
        .exclude(comments="Resubmitted")
        .exclude(resubmitted=True)
        .order_by("-item__id")
    )

    pagevars = {
        "page_title": "Rejected Items",
        "dp_rejected_list": dp_rejected_list,
        "pr_rejected_list": pr_rejected_list,
        "mkt_rejected_list": mkt_rejected_list,
    }

    return render(
        request, "workflow/plant_review/items_rejected.html", context=pagevars
    )


def process_review(request, item_id, category, update_type):
    """Handles plant or marketing employee either accepting, rejecting,
    or resubmitting artwork based on PDF, assignments, etc...
    Update type will be either Accept or Reject.
    """
    # note: item_id is really item_review[n].id
    item_review = ItemReview.objects.get(id=item_id, review_catagory=category)
    comment = request.GET.get("comment", "")
    if comment == "null":
        comment = ""
    item_review.do_ok(update_type, comment)
    return HttpResponse(JSMessage("Saved."))


class WorkflowModelChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        if obj.name == "Beverage":
            return "Evergreen"
        else:
            return obj.name


class JobSearchForm(forms.Form):
    """Main form for Job editing. Displayed in upper left corner of the job
    detail page.
    """

    # Universal search field - searches across multiple fields
    search_all = forms.CharField(
        required=False,
        help_text="Search across job number, job name, brand, customer, PO#, comments, and instructions",
        widget=forms.TextInput(attrs={
            'placeholder': 'Search all fields...',
            'class': 'search-all-input',
            'style': 'width: 400px; padding: 8px; font-size: 14px;'
        })
    )

    job_num = forms.IntegerField(min_value=1, max_value=99999, required=False)
    name = forms.CharField(required=False)
    brand_name = forms.CharField(required=False)
    customer = forms.CharField(required=False)
    customer_id = forms.IntegerField(required=False)
    po_number = forms.CharField(required=False)
    e_tools_id = forms.IntegerField(required=False)
    
    # Phase 2: Additional key fields for enhanced search
    comments = forms.CharField(required=False, help_text="Search in job comments")
    instructions = forms.CharField(required=False, help_text="Search in job instructions")
    customer_email = forms.CharField(required=False, help_text="Search by customer email")
    graphic_supplier = forms.CharField(required=False, help_text="Search by graphic supplier")
    user_keywords = forms.CharField(required=False, help_text="Search in user-defined keywords")

    status_choices = []
    status_choices.append(("", "---------"))
    for choice in app_defs.JOB_STATUSES:
        status_choices.append(choice)

    status = forms.ChoiceField(choices=status_choices, required=False)
    date_in_start = forms.DateField(required=False)
    date_in_end = forms.DateField(required=False)
    date_needed_start = forms.DateField(required=False)
    date_needed_end = forms.DateField(required=False)
    # customer = forms.ModelChoiceField(queryset=Customer.objects.all(), required=False)

    # Limit the field to accounts with given permission.
    # This permission must be assigned via the Group.
    permission = Permission.objects.get(codename="in_artist_pulldown")
    # This is a list of people past and present who have worked at clemson at some point
    # This is mainly used for job and item searching.
    grouped_artist_users = group_members = User.objects.filter(
        groups__name="ClemsonPersonnel"
    ).order_by("username")
    artist = forms.ModelChoiceField(queryset=None, required=False)

    permission = Permission.objects.get(codename="salesperson")
    # Select users who are a member of the set of groups with the given permission.
    # we are changing this to include all salespeople and not just the active ones.
    grouped_sales_users = User.objects.filter(
        groups__in=permission.group_set.all()
    ).order_by("username")
    salesperson = forms.ModelChoiceField(queryset=None, required=False)

    permission = Permission.objects.get(codename="is_fsb_csr")
    # Select users who are a member of the set of groups with the given permission.
    grouped_csr_users = User.objects.filter(
        is_active=True, groups__in=permission.group_set.all()
    ).order_by("username")

    csr = forms.ModelChoiceField(queryset=grouped_csr_users, required=False)

    printgroup = forms.ModelChoiceField(
        queryset=QAD_PrintGroups.objects.all().order_by("name"), required=False
    )
    # Build list of prepress suppliers, an option search field.
    prepress_choices = []
    prepress_choices.append(("", "---------"))
    for choice in app_defs.PREPRESS_SUPPLIERS:
        prepress_choices.append(choice)
    prepress_supplier = forms.ChoiceField(choices=prepress_choices, required=False)
    keywords = forms.CharField(required=False)
    # Build list of carton types, an optional search field.
    carton_type_choices = []
    carton_type_choices.append(("", "---------"))
    for choice in app_defs.CARTON_JOB_TYPES:
        carton_type_choices.append(choice)
    carton_type = forms.ChoiceField(choices=carton_type_choices, required=False)
    workflow = WorkflowModelChoiceField(
        queryset=Site.objects.all().exclude(name="Container"), required=False
    )
    sort_by = forms.ChoiceField(choices=[("id", "Job #")], required=False)
    sort_order = forms.ChoiceField(
        choices=[("desc", "Descending"), ("asc", "Ascending")], required=False
    )

    def __init__(self, request, *args, **kwargs):
        """Populate some of the relational fields."""
        super(JobSearchForm, self).__init__(*args, **kwargs)
        artist_qset = general_funcs.filter_query_same_perms(
            request, self.grouped_artist_users
        )
        self.fields["artist"].queryset = artist_qset
        sales_qset = general_funcs.filter_query_same_perms(
            request, self.grouped_sales_users
        )
        self.fields["salesperson"].queryset = sales_qset


class JobSearchResultsView(ListView):
    paginate_by = 25
    template_name = "workflow/search/search_results.html"
    form = None

    def get_queryset(self):
        qset = Job.objects.all()
        if self.form is not None:
            # Universal search - searches across multiple fields
            search_all = self.form.cleaned_data.get("search_all", None)
            if search_all:
                search_words = search_all.split()
                q = Q()
                for word in search_words:
                    word_q = (
                        Q(id__icontains=word) |  # Job number search
                        Q(name__icontains=word) |
                        Q(brand_name__icontains=word) |
                        Q(customer_name__icontains=word) |
                        Q(po_number__icontains=word) |
                        Q(customer_po_number__icontains=word) |
                        Q(comments__icontains=word) |
                        Q(instructions__icontains=word) |
                        Q(e_tools_id__icontains=word) |
                        Q(customer_email__icontains=word) |
                        Q(graphic_supplier__icontains=word) |
                        Q(user_keywords__icontains=word)
                    )
                    q &= word_q  # AND all words together
                qset = qset.filter(q)

            s_job_num = self.form.cleaned_data.get("job_num", None)
            if s_job_num:
                qset = qset.filter(id__icontains=s_job_num)

            # Hell yeah. Filters everything containing each word in the search.
            s_job_name = self.form.cleaned_data.get("name", None)
            if s_job_name:
                # Split search string on spaces, return results that
                # contain each word. ( q |= Q(... ) would return this word OR that word.
                search_words = s_job_name.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(name__icontains=word)
                qset = qset.filter(q)

            s_job_status = self.form.cleaned_data.get("status", None)
            if s_job_status:
                qset = qset.filter(status=s_job_status)

            s_customer = self.form.cleaned_data.get("customer", None)
            if s_customer:
                # Split search string on spaces, return results that
                # contain each word. ( q |= Q ) would return this word OR that word.
                search_words = s_customer.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(customer_name__icontains=s_customer)
                qset = qset.filter(q)

            # Customer ID - only visible to Carton
            s_job_customer_id = self.form.cleaned_data.get("customer_id", None)
            if s_job_customer_id:
                qset = qset.filter(customer_identifier__icontains=s_job_customer_id)

            # Brand name - only visible to Beverage
            s_brand_name = self.form.cleaned_data.get("brand_name", None)
            if s_brand_name:
                qset = qset.filter(brand_name__icontains=s_brand_name.strip())

            # Brand name - only visible to Beverage
            s_po_number = self.form.cleaned_data.get("po_number", None)
            if s_po_number:
                qset = qset.filter(po_number__icontains=s_po_number.strip())

            # Etools ID - only visible to Foodservice
            s_job_e_tools_id = self.form.cleaned_data.get("e_tools_id", None)
            if s_job_e_tools_id:
                qset = qset.filter(e_tools_id__icontains=s_job_e_tools_id)

            # Print Group - only visible to Foodservice
            s_job_printgroup = self.request.GET.get("printgroup", None)
            if s_job_printgroup:
                qset = qset.filter(printgroup=s_job_printgroup)

            # If start and end date given, search on range, else search on just start.
            s_job_date_in_start = self.form.cleaned_data.get("date_in_start", None)
            s_job_date_in_end = self.form.cleaned_data.get("date_in_end", None)
            if s_job_date_in_end:
                if s_job_date_in_start:
                    qset = qset.filter(
                        creation_date__range=(s_job_date_in_start, s_job_date_in_end)
                    )
            elif s_job_date_in_start:
                qset = qset.filter(creation_date=s_job_date_in_start)

            # If start and end date given, search on range, else search on just start.
            s_job_date_needed_start = self.form.cleaned_data.get(
                "date_needed_start", None
            )
            s_job_date_needed_end = self.form.cleaned_data.get("date_needed_end", None)
            if s_job_date_needed_end:
                if s_job_date_needed_start:
                    qset = qset.filter(
                        due_date__range=(s_job_date_needed_start, s_job_date_needed_end)
                    )
            elif s_job_date_needed_start:
                qset = qset.filter(creation_date=s_job_date_needed_start)

            s_job_artist = self.form.cleaned_data.get("artist", None)
            if s_job_artist:
                qset = qset.filter(artist=s_job_artist)

            s_job_salesperson = self.form.cleaned_data.get("salesperson", None)
            if s_job_salesperson:
                qset = qset.filter(salesperson=s_job_salesperson)

            s_job_csr = self.form.cleaned_data.get("csr", None)
            if s_job_csr:
                qset = qset.filter(csr=s_job_csr)

            s_job_prepress_supplier = self.form.cleaned_data.get(
                "prepress_supplier", None
            )
            if s_job_prepress_supplier:
                qset = qset.filter(prepress_supplier=s_job_prepress_supplier)

            s_keywords = self.form.cleaned_data.get("keywords", None)
            if s_keywords:
                # Split search string on spaces, return results that
                # contain each word. ( q |= Q(... ) would return this word OR that word.
                search_words = s_keywords.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(generated_keywords__icontains=word)
                qset = qset.filter(q)

            s_job_carton_type = self.form.cleaned_data.get("carton_type", None)
            if s_job_carton_type:
                qset = qset.filter(carton_type=s_job_carton_type)

            s_job_workflow = self.form.cleaned_data.get("workflow", None)
            if s_job_workflow:
                qset = qset.filter(workflow=s_job_workflow)

            # Phase 2: Additional key field searches
            s_comments = self.form.cleaned_data.get("comments", None)
            if s_comments:
                search_words = s_comments.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(comments__icontains=word)
                qset = qset.filter(q)

            s_instructions = self.form.cleaned_data.get("instructions", None)
            if s_instructions:
                search_words = s_instructions.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(instructions__icontains=word)
                qset = qset.filter(q)

            s_customer_email = self.form.cleaned_data.get("customer_email", None)
            if s_customer_email:
                qset = qset.filter(customer_email__icontains=s_customer_email.strip())

            s_graphic_supplier = self.form.cleaned_data.get("graphic_supplier", None)
            if s_graphic_supplier:
                qset = qset.filter(graphic_supplier__icontains=s_graphic_supplier.strip())

            s_user_keywords = self.form.cleaned_data.get("user_keywords", None)
            if s_user_keywords:
                search_words = s_user_keywords.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(user_keywords__icontains=word)
                qset = qset.filter(q)

        # Sort records.
        sort = self.request.GET.get("sort_order", "")

        if sort == "asc":
            qset = qset.order_by("id")
        else:
            qset = qset.order_by("-id")

        # Start standard filters to qset that would be used in both browse and search.
        qset = qset.exclude(id=99999)

        # Filter based on user workflow access.
        qset = qset.filter(
            workflow__name__in=general_funcs.get_user_workflow_access(self.request)
        )

        return qset

    def get_context_data(self, **kwargs):
        context = super(JobSearchResultsView, self).get_context_data(**kwargs)

        context["page_title"] = "Job Search Results"
        context["type"] = "job"
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context

    # Require the user to be logged in to GOLD to view.
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        if self.get_queryset().count() == 1:
            job_detail_url = reverse("job_detail", args=[self.get_queryset()[0].id])
            return HttpResponseRedirect(job_detail_url)
        else:
            return super(JobSearchResultsView, self).dispatch(*args, **kwargs)


@login_required
def job_search(request):
    """Displays the job search form."""
    form = JobSearchForm(request, data=request.GET or None)

    if request.GET and form.is_valid() and not request.GET.get('legacy'):
        # Call the result view directly for display.
        return JobSearchResultsView.as_view(form=form)(request)
    else:
        # Check user preference for legacy interface (default to True)
        use_legacy = request.session.get('use_legacy_search', True)
        
        # URL parameter overrides user preference (for toggle functionality)
        if request.GET.get('legacy') is not None:
            use_legacy = request.GET.get('legacy') == '1'
        
        template_name = "workflow/search/search_form_legacy.html" if use_legacy else "workflow/search/search_form.html"
        
        # This is the search page to be re-displayed if there's a problem or no
        # POST data.
        pagevars = {
            "page_title": "Job Search",
            "form": form,
            "type": "job",
            "current_is_legacy": use_legacy,  # Pass this to template for toggle links
        }
        return render(request, template_name, context=pagevars)


# this is some shim sham to make the label for ItemTrackerType "Nutrition Facts" return as
# "New Nutrition Facts" rather than modify the current 114 objects this is attached to.
# Eventually we wont need the New so we can just delete this.
class TrackerModelChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return "New " + obj.name


class ItemSearchForm(forms.Form):
    """Item search form"""

    # Universal search field - searches across multiple fields
    search_all = forms.CharField(
        required=False,
        help_text="Search across job number, size, description, UPC, BOM, customer code, and more",
        widget=forms.TextInput(attrs={
            'placeholder': 'Search all fields...',
            'class': 'search-all-input',
            'style': 'width: 400px; padding: 8px; font-size: 14px;'
        })
    )

    job_num = forms.IntegerField(min_value=1, max_value=99999, required=False)
    name = forms.CharField(required=False)
    brand_name = forms.CharField(required=False)
    fsb_nine_digit = forms.IntegerField(required=False)
    size = forms.CharField(required=False)
    description = forms.CharField(required=False)
    upc_number = forms.CharField(required=False)
    bom_number = forms.IntegerField(required=False)
    wrin_number = forms.CharField(required=False)
    case_pack = forms.IntegerField(required=False)
    size_id = forms.IntegerField(required=False)
    
    # Phase 2: Additional key fields for enhanced search  
    customer_code = forms.CharField(required=False, help_text="Search by customer code")
    graphic_req_number = forms.CharField(required=False, help_text="Search by graphic request number")
    plant_comments = forms.CharField(required=False, help_text="Search in plant comments")
    mkt_review_comments = forms.CharField(required=False, help_text="Search in marketing review comments")
    
    plant = forms.ModelChoiceField(queryset=Plant.objects.none(), required=False)
    press = forms.ModelChoiceField(queryset=Press.objects.none(), required=False)
    platemaker = forms.ModelChoiceField(
        queryset=Platemaker.objects.none(), required=False
    )

    # New code --- Search for number of colors in job
    color_num_low = forms.IntegerField(min_value=0, max_value=9, required=False)
    color_num_high = forms.IntegerField(min_value=1, max_value=9, required=False)
    # End new code
    # color_prefix = forms.ChoiceField()
    itemcolor = forms.CharField(required=False)
    platetype_choices = []
    platetype_choices.append(("", "---"))
    for type in app_defs.PLATE_OPTIONS:
        platetype_choices.append(type)
    platetype = forms.ChoiceField(choices=platetype_choices, required=False)

    specialmfg = forms.ModelChoiceField(
        queryset=SpecialMfgConfiguration.objects.none(), required=False
    )

    # Build list of prepress suppliers, an option search field.
    prepress_choices = []
    prepress_choices.append(("", "---------"))
    for choice in app_defs.PREPRESS_SUPPLIERS:
        prepress_choices.append(choice)
    prepress_supplier = forms.ChoiceField(choices=prepress_choices, required=False)
    quality_choices = []
    quality_choices.append(("", "---"))
    for choice in app_defs.COMPLEXITY_OPTIONS:
        quality_choices.append(choice)
    quality = forms.ChoiceField(choices=quality_choices, required=False)

    current_proof_date_start = forms.DateField(required=False)
    current_proof_date_end = forms.DateField(required=False)
    date_in_start = forms.DateField(required=False)
    date_in_end = forms.DateField(required=False)

    # Adding proof out date
    all_proof_date_start = forms.DateField(required=False)
    all_proof_date_end = forms.DateField(required=False)

    approval_date_start = forms.DateField(required=False)
    approval_date_end = forms.DateField(required=False)
    final_file_date_start = forms.DateField(required=False)
    final_file_date_end = forms.DateField(required=False)
    # Build list of prepress suppliers, an option search field.
    item_situation_choices = []
    item_situation_choices.append(("", "---------"))
    for choice in app_defs.SITUATION_OPTIONS:
        item_situation_choices.append(choice)
    item_situation = forms.ChoiceField(choices=item_situation_choices, required=False)
    # This is a list of people past and present who have worked at clemson at some point
    # This is mainly used for job and item searching.
    grouped_artist_users = group_members = User.objects.filter(
        groups__name="ClemsonPersonnel"
    ).order_by("username")
    artist = forms.ModelChoiceField(queryset=grouped_artist_users, required=False)

    permission = Permission.objects.get(codename="salesperson")
    # Select users who are a member of the set of groups with the given permission.
    # we are changing this to include all salespeople and not just the active ones.
    grouped_sales_users = User.objects.filter(
        groups__in=permission.group_set.all()
    ).order_by("username")
    salesperson = forms.ModelChoiceField(queryset=grouped_sales_users, required=False)

    nut_trackers = ItemTrackerType.objects.filter(category__name="Beverage Nutrition")
    nutrition = TrackerModelChoiceField(queryset=nut_trackers, required=False)
    mkt_trackers = ItemTrackerType.objects.filter(category__name="Marketing")
    marketing = forms.ModelChoiceField(queryset=mkt_trackers, required=False)
    promo_trackers = ItemTrackerType.objects.filter(category__name="Promotional")
    promo = forms.ModelChoiceField(queryset=promo_trackers, required=False)
    workflow = WorkflowModelChoiceField(
        queryset=Site.objects.all().exclude(name="Container"), required=False
    )
    sort_by = forms.ChoiceField(choices=[("id", "Job #")], required=False)
    sort_order = forms.ChoiceField(
        choices=[("desc", "Descending"), ("asc", "Ascending")], required=False
    )
    # Carton job fields
    grn = forms.CharField(required=False)
    one_up_die = forms.CharField(required=False)
    step_die = forms.CharField(required=False)
    coating_pattern = forms.CharField(required=False)
    customer_code = forms.CharField(required=False)
    graphic_req_number = forms.CharField(required=False)

    def __init__(self, request, *args, **kwargs):
        super(ItemSearchForm, self).__init__(*args, **kwargs)
        # Only display artists linked to workflows the user has access to.
        artist_qset = general_funcs.filter_query_same_perms(
            request, self.grouped_artist_users
        )
        self.fields["artist"].queryset = artist_qset

        # Only display salespeople linked to workflows the user has access to.
        sales_qset = general_funcs.filter_query_same_perms(
            request, self.grouped_sales_users
        )
        self.fields["salesperson"].queryset = sales_qset

        # Get the workflows that this user has access to.
        user_workflows = general_funcs.get_user_workflow_access(request)
        # Only display plants linked to workflows the user has access to.
        self.fields["plant"].queryset = Plant.objects.filter(
            workflow__name__in=user_workflows
        ).order_by("name")
        # Only display presses linked to workflows the user has access to.
        self.fields["press"].queryset = Press.objects.filter(
            workflow__name__in=user_workflows
        ).order_by("name")
        # Only display platemakers linked to workflows the user has access to.
        self.fields["platemaker"].queryset = (
            Platemaker.objects.filter(workflow__name__in=user_workflows)
            .distinct()
            .order_by("name")
        )
        # Only display special mfg configs linked to the user workflows.
        self.fields["specialmfg"].queryset = (
            SpecialMfgConfiguration.objects.filter(workflow__name__in=user_workflows)
            .distinct()
            .order_by("name")
        )


class ItemSearchResultsView(ListView):
    paginate_by = 25
    template_name = "workflow/search/search_results.html"
    form = None

    def get_queryset(self):
        qset = Item.objects.all()
        if self.form is not None:
            # Universal search - searches across multiple fields
            search_all = self.form.cleaned_data.get("search_all", None)
            if search_all:
                search_words = search_all.split()
                q = Q()
                for word in search_words:
                    word_q = (
                        Q(job__id__icontains=word) |  # Job number search
                        Q(size__size__icontains=word) |
                        Q(bev_item_name__icontains=word) |
                        Q(description__icontains=word) |
                        Q(upc_number__icontains=word) |
                        Q(bom_number__icontains=word) |
                        Q(wrin_number__icontains=word) |
                        Q(customer_code__icontains=word) |
                        Q(coating_pattern__icontains=word) |
                        Q(graphic_req_number__icontains=word) |
                        Q(plant_comments__icontains=word) |
                        Q(mkt_review_comments__icontains=word) |
                        Q(job__name__icontains=word) |
                        Q(job__brand_name__icontains=word) |
                        Q(job__customer_name__icontains=word)
                    )
                    q &= word_q  # AND all words together
                qset = qset.filter(q)

            s_job_num = self.form.cleaned_data.get("job_num", None)
            if s_job_num:
                qset = qset.filter(job__id=s_job_num)

            # If a specific db-level ID is specified for size, use that.
            s_item_size_id = self.form.cleaned_data.get("size_id", None)
            if s_item_size_id:
                qset = qset.filter(size__id=s_item_size_id)
            else:
                # Otherwise, assume this is a text-based name search.
                s_item_size = self.form.cleaned_data.get("size", None)
                if s_item_size:
                    stripped_size = s_item_size.strip()
                    qset = qset.filter(
                        Q(size__size__icontains=stripped_size)
                        | Q(bev_item_name__icontains=stripped_size)
                    )

            # Hell yeah. Filters everything containing each word in the search.
            s_job_name = self.form.cleaned_data.get("name", None)
            if s_job_name:
                search_words = s_job_name.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(job__name__icontains=word)
                qset = qset.filter(q)

            # Brand name - only visible to Beverage
            s_brand_name = self.form.cleaned_data.get("brand_name", None)
            if s_brand_name:
                qset = qset.filter(job__brand_name__icontains=s_brand_name.strip())

            s_description = self.form.cleaned_data.get("description", None)
            if s_description:
                qset = qset.filter(description__icontains=s_description.strip())

            s_item_quality = self.form.cleaned_data.get("quality", None)
            if s_item_quality:
                qset = qset.filter(quality=s_item_quality)

            s_item_fsb_nine = self.form.cleaned_data.get("fsb_nine_digit", None)
            if s_item_fsb_nine:
                qset = qset.filter(fsb_nine_digit__icontains=s_item_fsb_nine)

            s_item_bom_number = self.form.cleaned_data.get("bom_number", None)
            if s_item_bom_number:
                qset = qset.filter(bom_number__icontains=s_item_bom_number)

            s_item_wrin_number = self.form.cleaned_data.get("wrin_number", None)
            if s_item_wrin_number:
                qset = qset.filter(wrin_number__icontains=s_item_wrin_number)

            s_item_case_pack = self.form.cleaned_data.get("case_pack", None)
            if s_item_case_pack:
                qset = qset.filter(case_pack=s_item_case_pack)

            s_item_upc_number = self.form.cleaned_data.get("upc_number", None)
            if s_item_upc_number:
                qset = qset.filter(upc_number__icontains=s_item_upc_number)

            s_item_plant = self.form.cleaned_data.get("plant", None)
            if s_item_plant:
                qset = qset.filter(printlocation__plant=s_item_plant)

            s_item_press = self.form.cleaned_data.get("press", None)
            if s_item_press:
                qset = qset.filter(printlocation__press=s_item_press)

            s_item_prepress_supplier = self.form.cleaned_data.get(
                "prepress_supplier", None
            )
            if s_item_prepress_supplier:
                qset = qset.filter(job__prepress_supplier=s_item_prepress_supplier)

            s_item_specialmfg = self.form.cleaned_data.get("specialmfg", None)
            if s_item_specialmfg:
                qset = qset.filter(special_mfg=s_item_specialmfg)

            s_item_platemaker = self.form.cleaned_data.get("platemaker", None)
            if s_item_platemaker:
                qset = qset.filter(platepackage__platemaker=s_item_platemaker)

            s_item_platetype = self.form.cleaned_data.get("platetype", None)
            if s_item_platetype:
                qset = qset.filter(platepackage__platetype=s_item_platetype)

            # If start and end date given, search on range, else search on just start.
            s_item_date_in_start = self.form.cleaned_data.get("date_in_start", None)
            s_item_date_in_end = self.form.cleaned_data.get("date_in_end", None)
            if s_item_date_in_end:
                if s_item_date_in_start:
                    qset = qset.filter(
                        creation_date__range=(s_item_date_in_start, s_item_date_in_end)
                    )
            elif s_item_date_in_start:
                qset = qset.filter(creation_date=s_item_date_in_start)

            # If start and end date given, search on range, else search on just start.
            # Get all the items proofed out between the dates.
            s_item_all_proof_date_start = self.form.cleaned_data.get(
                "all_proof_date_start", None
            )
            s_item_all_proof_date_end = self.form.cleaned_data.get(
                "all_proof_date_end", None
            )
            if s_item_all_proof_date_end:
                if s_item_all_proof_date_start:
                    qlog_set = (
                        JobLog.objects.filter(
                            type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
                            event_time__range=(
                                s_item_all_proof_date_start,
                                s_item_all_proof_date_end,
                            ),
                        )
                        .values("item")
                        .query
                    )
                    qset = qset.filter(id__in=qlog_set)
            elif s_item_all_proof_date_start:
                qlog_set = (
                    JobLog.objects.filter(
                        type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
                        event_time=s_item_all_proof_date_start,
                    )
                    .values("item")
                    .query
                )
                qset = qset.filter(id__in=qlog_set)

            # If start and end date given, search on range, else search on just start.
            s_item_apr_date_start = self.form.cleaned_data.get(
                "approval_date_start", None
            )
            s_item_apr_date_end = self.form.cleaned_data.get("approval_date_end", None)
            if s_item_apr_date_end:
                if s_item_apr_date_start:
                    qlog_set = (
                        JobLog.objects.filter(
                            type=JOBLOG_TYPE_ITEM_APPROVED,
                            event_time__range=(
                                s_item_apr_date_start,
                                s_item_apr_date_end,
                            ),
                        )
                        .values("item")
                        .query
                    )
                    qset = qset.filter(id__in=qlog_set)
            elif s_item_apr_date_start:
                qlog_set = (
                    JobLog.objects.filter(
                        type=JOBLOG_TYPE_ITEM_APPROVED, event_time=s_item_apr_date_start
                    )
                    .values("item")
                    .query
                )
                qset = qset.filter(id__in=qlog_set)

            # If start and end date given, search on range, else search on just start.
            s_item_ffo_date_start = self.form.cleaned_data.get(
                "final_file_date_start", None
            )
            s_item_ffo_date_end = self.form.cleaned_data.get(
                "final_file_date_end", None
            )
            if s_item_ffo_date_end:
                if s_item_ffo_date_start:
                    qlog_set = (
                        JobLog.objects.filter(
                            type=JOBLOG_TYPE_ITEM_FILED_OUT,
                            event_time__range=(
                                s_item_ffo_date_start,
                                s_item_ffo_date_end,
                            ),
                        )
                        .values("item")
                        .query
                    )
                qset = qset.filter(id__in=qlog_set)
            elif s_item_ffo_date_start:
                qlog_set = (
                    JobLog.objects.filter(
                        type=JOBLOG_TYPE_ITEM_FILED_OUT,
                        event_time=s_item_ffo_date_start,
                    )
                    .values("item")
                    .query
                )
                qset = qset.filter(id__in=qlog_set)

            s_item_situation = self.form.cleaned_data.get("item_situation", None)
            if s_item_situation:
                qset = qset.filter(item_situation=s_item_situation)

            s_job_artist = self.form.cleaned_data.get("artist", None)
            if s_job_artist:
                qset = qset.filter(job__artist=s_job_artist)

            s_job_salesperson = self.form.cleaned_data.get("salesperson", None)
            if s_job_salesperson:
                qset = qset.filter(job__salesperson=s_job_salesperson)

            s_item_marketing = self.form.cleaned_data.get("marketing", None)
            if s_item_marketing:
                qtracker_set = (
                    ItemTracker.objects.filter(type=s_item_marketing)
                    .values("item")
                    .query
                )
                qset = qset.filter(id__in=qtracker_set)

            s_item_nutrition = self.form.cleaned_data.get("nutrition", None)
            if s_item_nutrition:
                qtracker_set = (
                    ItemTracker.objects.filter(type=s_item_nutrition)
                    .values("item")
                    .query
                )
                qset = qset.filter(id__in=qtracker_set)

            s_item_promo = self.form.cleaned_data.get("promo", None)
            if s_item_promo:
                qtracker_set = (
                    ItemTracker.objects.filter(type=s_item_promo).values("item").query
                )
                qset = qset.filter(id__in=qtracker_set)

            s_itemcolor = self.form.cleaned_data.get("itemcolor", None)
            if s_itemcolor:
                thiscolor = ItemColor.objects.filter(
                    color__icontains=s_itemcolor
                ).values_list("item_id")
                qset = qset.filter(id__in=thiscolor)

            s_item_workflow = self.form.cleaned_data.get("workflow", None)
            if s_item_workflow:
                qset = qset.filter(job__workflow=s_item_workflow)

            # Carton job fields
            s_grn = self.form.cleaned_data.get("grn", None)
            if s_grn:
                qset = qset.filter(grn__icontains=s_grn.strip())
            s_one_up_die = self.form.cleaned_data.get("one_up_die", None)
            if s_one_up_die:
                qset = qset.filter(one_up_die__icontains=s_one_up_die.strip())
            s_step_die = self.form.cleaned_data.get("step_die", None)
            if s_step_die:
                qset = qset.filter(step_die__icontains=s_step_die.strip())
            s_coating_pattern = self.form.cleaned_data.get("coating_pattern", None)
            if s_coating_pattern:
                qset = qset.filter(coating_pattern__icontains=s_coating_pattern.strip())
            s_customer_code = self.form.cleaned_data.get("customer_code", None)
            if s_customer_code:
                qset = qset.filter(customer_code__icontains=s_customer_code.strip())
            s_graphic_req_number = self.form.cleaned_data.get(
                "graphic_req_number", None
            )
            if s_graphic_req_number:
                qset = qset.filter(
                    graphic_req_number__icontains=s_graphic_req_number.strip()
                )

            # Phase 2: Additional key field searches for items
            s_plant_comments = self.form.cleaned_data.get("plant_comments", None)
            if s_plant_comments:
                search_words = s_plant_comments.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(plant_comments__icontains=word)
                qset = qset.filter(q)

            s_mkt_review_comments = self.form.cleaned_data.get("mkt_review_comments", None)
            if s_mkt_review_comments:
                search_words = s_mkt_review_comments.split(" ")
                q = Q()
                for word in search_words:
                    q &= Q(mkt_review_comments__icontains=word)
                qset = qset.filter(q)

            # Filter based on user status.
            qset = qset.filter(
                workflow__name__in=general_funcs.get_user_workflow_access(self.request)
            )

            # Filter based on Number of Colors
            s_min_colors = self.form.cleaned_data.get("color_num_low", None)
            s_max_colors = self.form.cleaned_data.get("color_num_high", None)

            if s_min_colors or s_max_colors:
                result_list = []

                if s_min_colors and s_max_colors:
                    for item in qset:
                        num_colors = ItemColor.objects.filter(item=item).count()
                        if num_colors >= s_min_colors and num_colors <= s_max_colors:
                            result_list.append(item.id)
                elif s_min_colors:
                    for item in qset:
                        num_colors = ItemColor.objects.filter(item=item).count()
                        if num_colors >= s_min_colors:
                            result_list.append(item.id)
                elif s_max_colors:
                    for item in qset:
                        num_colors = ItemColor.objects.filter(item=item).count()
                        if num_colors <= s_max_colors:
                            result_list.append(item.id)

                qset = qset.filter(id__in=result_list)

            # Sort records.
            sort = self.form.cleaned_data.get("sort_order", "desc")
            if sort == "asc":
                qset = qset.order_by("job__id")
            else:
                qset = qset.order_by("-job__id")

            # Exclude the test job.
            qset = qset.exclude(job__id=99999)

        return qset

    def get_context_data(self, **kwargs):
        context = super(ItemSearchResultsView, self).get_context_data(**kwargs)

        context["page_title"] = "Item Search Results"
        context["type"] = "item"
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context

    # Require the user to be logged in to GOLD to view.
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        if self.get_queryset().count() == 1:
            job_detail_url = reverse("job_detail", args=[self.get_queryset()[0].job.id])
            return HttpResponseRedirect(job_detail_url)
        else:
            return super(ItemSearchResultsView, self).dispatch(*args, **kwargs)


@login_required
def item_search(request):
    """Displays the item search form."""
    form = ItemSearchForm(request, data=request.GET or None)
    if request.GET and form.is_valid() and not request.GET.get('legacy'):
        # Call the result view directly for display.
        return ItemSearchResultsView.as_view(form=form)(request)
    else:
        # Check user preference for legacy interface (default to True)
        use_legacy = request.session.get('use_legacy_search', True)
        
        # URL parameter overrides user preference (for toggle functionality)
        if request.GET.get('legacy') is not None:
            use_legacy = request.GET.get('legacy') == '1'
        
        template_name = "workflow/search/search_form_legacy.html" if use_legacy else "workflow/search/search_form.html"
        
        # This is the search page to be re-displayed if there's a problem or no
        # POST data.
        pagevars = {
            "page_title": "Item Search",
            "form": form,
            "type": "item",
            "current_is_legacy": use_legacy,  # Pass this to template for toggle links
        }
        return render(request, template_name, context=pagevars)


class ItemFindSame(ListView):
    """Do a quick search for all items of the same name."""

    paginate_by = 25
    template_name = "workflow/search/search_results.html"
    form = None

    def get_queryset(self, **kwargs):
        try:
            item_id = self.kwargs["item_id"]
        except Exception:
            item_id = None

        qset = Item.objects.all()
        item = Item.objects.get(id=item_id)

        q = Q(bev_item_name=item.bev_item_name)
        qset = qset.filter(q)

        # Limit to jobs of same workflow, include current job.
        qset = qset.filter(job__workflow=item.job.workflow).order_by("-id")

        return qset

    def get_context_data(self, **kwargs):
        context = super(ItemFindSame, self).get_context_data(**kwargs)

        context["page_title"] = "Item Search Results"
        context["type"] = "item"
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context


def job_todo_list(request, manager_tools=False):
    """Daily Report for the Clemson Graphics Hub
    -Lists upcoming due jobs.
    -Account for jobs due on weekends.
    -List both due jobs and revisions.
    -Be able to show FSB, BEV and CON, color coded?
    -Misc Hub stats, billing, #jobs, etc...

    The manager_tools flag is used to show extra info when requested. The
    manager may not always want that info visible.
    """
    # Establish next 5 business days (need to account for the weekends)
    day0 = date.today()
    day1 = day0
    # If report run on Saturday or Sunday, return Monday
    """
    if day0.isoweekday() == 6:
        day1 = day0 + timedelta(days=2)
    if day0.isoweekday() == 7:
        day1 = day0 + timedelta(days=1)
    #Again, skipping past the weekends
    day2 = day1 + timedelta(days=1)
    if day2.isoweekday() == 6:
        day2 = day2 + timedelta(days=1)
    if day2.isoweekday() == 7:
        day2 = day2 + timedelta(days=1)
    """
    day2 = day1 + timedelta(days=1)
    day3 = day2 + timedelta(days=1)
    day4 = day3 + timedelta(days=1)
    day5 = day4 + timedelta(days=1)
    day6 = day5 + timedelta(days=1)
    day7 = day6 + timedelta(days=1)
    day8 = day7 + timedelta(days=1)

    # Search for jobs due with 5 day period
    # Needs to use real_due_date
    view_workflows = general_funcs.get_user_workflow_access(request)
    """
    Jobs due in the next 6 days.
    """
    # Query once, check each for completion, then sort.
    excluded_prepress_suppliers = [
        "Phototype",
        "PHT",
        "SGS",
        "SHK",
        "Schawk",
        "Southern Graphics",
    ]
    jobs_due_range = (
        Job.objects.filter(
            real_due_date__range=(day1, day8),
            status__in=["Active", "Pending"],
            workflow__name__in=view_workflows,
        )
        .exclude(prepress_supplier__in=excluded_prepress_suppliers)
        .order_by("workflow")
    )

    incomplete_jobs = []
    for job in jobs_due_range:
        complete = True
        for item in job.get_item_qset():
            if complete:
                if item.first_proof_date() or item.overdue_exempt:
                    # Yup, still complete (True)
                    pass
                else:
                    complete = False
                    incomplete_jobs.append(job)

    jobs_due_1 = []
    jobs_due_2 = []
    jobs_due_3 = []
    jobs_due_4 = []
    jobs_due_5 = []
    jobs_due_6 = []
    jobs_due_7 = []
    jobs_due_8 = []
    # Take query, append to appropriate tuple for each due_date.
    for job in incomplete_jobs:
        if job.real_due_date == day1:
            jobs_due_1.append(job)
        elif job.real_due_date == day2:
            jobs_due_2.append(job)
        elif job.real_due_date == day3:
            jobs_due_3.append(job)
        elif job.real_due_date == day4:
            jobs_due_4.append(job)
        elif job.real_due_date == day5:
            jobs_due_5.append(job)
        elif job.real_due_date == day6:
            jobs_due_6.append(job)
        elif job.real_due_date == day7:
            jobs_due_7.append(job)
        elif job.real_due_date == day8:
            jobs_due_8.append(job)

    """
    We need to total up the artists' estimated times and the office wide averges
    for all jobs per day. Only if manager tools are turned on.
    """
    jobs_due_1_total_hours = 0
    jobs_due_1_avg_total_hours = 0
    jobs_due_2_total_hours = 0
    jobs_due_2_avg_total_hours = 0
    jobs_due_3_total_hours = 0
    jobs_due_3_avg_total_hours = 0
    jobs_due_4_total_hours = 0
    jobs_due_4_avg_total_hours = 0
    jobs_due_5_total_hours = 0
    jobs_due_5_avg_total_hours = 0
    jobs_due_6_total_hours = 0
    jobs_due_6_avg_total_hours = 0
    jobs_due_7_total_hours = 0
    jobs_due_7_avg_total_hours = 0
    jobs_due_8_total_hours = 0
    jobs_due_8_avg_total_hours = 0

    if manager_tools:
        for job in jobs_due_1:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_1_total_hours += averages[0]
            jobs_due_1_avg_total_hours += averages[1]
        for job in jobs_due_2:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_2_total_hours += averages[0]
            jobs_due_2_avg_total_hours += averages[1]
        for job in jobs_due_3:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_3_total_hours += averages[0]
            jobs_due_3_avg_total_hours += averages[1]
        for job in jobs_due_4:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_4_total_hours += averages[0]
            jobs_due_4_avg_total_hours += averages[1]
        for job in jobs_due_5:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_5_total_hours += averages[0]
            jobs_due_5_avg_total_hours += averages[1]
        for job in jobs_due_6:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_6_total_hours += averages[0]
            jobs_due_6_avg_total_hours += averages[1]
        for job in jobs_due_7:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_7_total_hours += averages[0]
            jobs_due_7_avg_total_hours += averages[1]
        for job in jobs_due_8:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_8_total_hours += averages[0]
            jobs_due_8_avg_total_hours += averages[1]

    """
    Revisions due in the next 6 days.
    """
    revisions = Revision.objects.filter(
        due_date__range=(day1, day8),
        complete_date__isnull=True,
        item__job__workflow__name__in=view_workflows,
    ).exclude(item__job__status__in=["Hold", "Cancelled"])
    # jobs_with_revisions = Job.objects.filter(id__in=revisons)
    revisions_due_1 = []
    revisions_due_2 = []
    revisions_due_3 = []
    revisions_due_4 = []
    revisions_due_5 = []
    revisions_due_6 = []
    revisions_due_7 = []
    revisions_due_8 = []
    overdue_revisions = []
    # Take query, append to appropriate tuple for each due_date.
    for rev in revisions:
        if rev.due_date == day1:
            if rev.item.job not in revisions_due_1:
                revisions_due_1.append(rev.item.job)
        elif rev.due_date == day2:
            if rev.item.job not in revisions_due_2:
                revisions_due_2.append(rev.item.job)
        elif rev.due_date == day3:
            if rev.item.job not in revisions_due_3:
                revisions_due_3.append(rev.item.job)
        elif rev.due_date == day4:
            if rev.item.job not in revisions_due_4:
                revisions_due_4.append(rev.item.job)
        elif rev.due_date == day5:
            if rev.item.job not in revisions_due_5:
                revisions_due_5.append(rev.item.job)
        elif rev.due_date == day6:
            if rev.item.job not in revisions_due_6:
                revisions_due_6.append(rev.item.job)
        elif rev.due_date == day7:
            if rev.item.job not in revisions_due_7:
                revisions_due_7.append(rev.item.job)
        elif rev.due_date == day8:
            if rev.item.job not in revisions_due_8:
                revisions_due_8.append(rev.item.job)
        elif rev.due_date < day1:
            if rev.item.job not in overdue_revisions:
                overdue_revisions.append(rev.item.job)

    """
    Add revisions to the artists' estimated times and the office wide averges
    for all jobs per day. Only if manager tools are turned on.
    """

    if manager_tools:
        for job in revisions_due_1:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_1_total_hours += averages[0]
            jobs_due_1_avg_total_hours += averages[1]
        for job in revisions_due_2:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_2_total_hours += averages[0]
            jobs_due_2_avg_total_hours += averages[1]
        for job in revisions_due_3:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_3_total_hours += averages[0]
            jobs_due_3_avg_total_hours += averages[1]
        for job in revisions_due_4:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_4_total_hours += averages[0]
            jobs_due_4_avg_total_hours += averages[1]
        for job in revisions_due_5:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_5_total_hours += averages[0]
            jobs_due_5_avg_total_hours += averages[1]
        for job in revisions_due_6:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_6_total_hours += averages[0]
            jobs_due_6_avg_total_hours += averages[1]
        for job in revisions_due_7:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_7_total_hours += averages[0]
            jobs_due_7_avg_total_hours += averages[1]
        for job in revisions_due_8:
            averages = job_todo_estimates(job)
            # Add the averages to the totals for this day.
            jobs_due_8_total_hours += averages[0]
            jobs_due_8_avg_total_hours += averages[1]

    """
    Jobs due in the last X days.
    """
    start_range = day0 + timedelta(days=-60)
    yesterday = day0 + timedelta(days=-1)
    activejobs_pastdate = (
        Job.objects.filter(
            real_due_date__range=(start_range, yesterday),
            workflow__name__in=view_workflows,
            status__in=["Active", "Pending"],
        )
        .exclude(
            prepress_supplier__in=[
                "Phototype",
                "PHT",
                "SGS",
                "SHK",
                "Schawk",
                "Southern Graphics",
            ]
        )
        .order_by("-id")
        .select_related()
    )
    # Maybe query all items in created in last X days w/o final file date?
    overdue_jobs = []
    for job in activejobs_pastdate:
        overdue = False
        for item in job.get_item_qset():
            if not overdue:
                if item.first_proof_date() or item.overdue_exempt:
                    pass
                else:
                    overdue = True
                    overdue_jobs.append(job)

    end_range = day0 + timedelta(days=60)
    # Jobs set to Hold, due in the last X days specified in limiting range.
    jobs_on_hold = Job.objects.filter(
        due_date__range=(start_range, end_range),
        workflow__name__in=view_workflows,
        status="Hold",
    ).order_by("-id")

    # Jobs with pending status.
    jobs_pending = Job.objects.filter(
        status="Pending", workflow__name__in=view_workflows
    ).exclude(
        prepress_supplier__in=(
            "PHT",
            "Phototype",
            "SGS",
            "Southern Graphics",
            "SHK",
            "Schawk",
        )
    )
    nine_digit_start_range = day0 + timedelta(days=-180)
    items_to_file_out = (
        Item.objects.filter(item_status="File Out")
        .exclude(fsb_nine_digit_date__lte=nine_digit_start_range)
        .exclude(fsb_nine_digit="")
        .values("job__id")
        .query
    )
    # Carton items don't have fsb nine digit numbers. Search for them separately.
    carton_items_to_file_out = (
        Item.objects.filter(job__workflow__name="Carton", item_status="File Out")
        .values("job__id")
        .query
    )
    jobs_needing_file_out = (
        Job.objects.filter(
            Q(id__in=items_to_file_out) | Q(id__in=carton_items_to_file_out)
        )
        .exclude(status__in=("Hold", "Cancelled", "Complete"))
        .order_by("-id")
    )

    jobs_needing_file_out_today = []
    jobs_needing_file_out_tomorrow = []
    jobs_needing_file_out_day3 = []
    jobs_needing_file_out_day4 = []
    jobs_needing_file_out_soon = []
    jobs_needing_file_out_overdue = []

    # We look 4 days out so we can see past weekends.
    for job in jobs_needing_file_out:
        if job.final_file_due_date():
            if job.final_file_due_date() == day1:
                jobs_needing_file_out_today.append(job)
            elif job.final_file_due_date() == day2:
                jobs_needing_file_out_tomorrow.append(job)
            elif job.final_file_due_date() == day3:
                jobs_needing_file_out_day3.append(job)
            elif job.final_file_due_date() == day4:
                jobs_needing_file_out_day4.append(job)
            elif job.final_file_due_date() > day4:
                jobs_needing_file_out_soon.append(job)
            elif job.final_file_due_date() < day1:
                jobs_needing_file_out_overdue.append(job)

    """
    Total up the up the estimated file out hours for each day.
    """
    jobs_needing_file_out_today_total = 0
    jobs_needing_file_out_tomorrow_total = 0
    jobs_needing_file_out_day3_total = 0
    jobs_needing_file_out_day4_total = 0
    jobs_needing_file_out_soon_total = 0
    jobs_needing_file_out_overdue_total = 0

    if manager_tools:
        for job in jobs_needing_file_out_today:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_today_total += estimate
        for job in jobs_needing_file_out_tomorrow:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_tomorrow_total += estimate
        for job in jobs_needing_file_out_day3:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_day3_total += estimate
        for job in jobs_needing_file_out_day4:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_day4_total += estimate
        for job in jobs_needing_file_out_soon:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_soon_total += estimate
        for job in jobs_needing_file_out_overdue:
            estimate = job.avg_fileout_time()
            # Add the estimate to the totals for this day.
            jobs_needing_file_out_overdue_total += estimate

    current_time = time.localtime()

    shipto_list = []
    shipto_list_dupes = []
    # Containers used while building duplicate shipment lists
    shipto_list_dupes_dict = {}
    # Prepare list of shipments going out in next 8 days, compare
    # addresses for any duplicates.
    for job in incomplete_jobs:
        for address in JobAddress.objects.filter(job=job):
            # Start a fresh job list for this address (include the current job)
            job_list = [job]
            # Check for duplicates.
            for existing_address in shipto_list:
                if (
                    existing_address.name.lower() == address.name.lower()
                    and existing_address.city.lower() == address.city.lower()
                ):
                    # artist averages may not be available in this scope; guard defensively
                    try:
                        # Use job_todo_estimates to compute artist and office averages
                        averages = job_todo_estimates(job)
                        artist_average, all_artist_average = averages
                    except Exception:
                        # If something goes wrong computing estimates, skip it
                        pass
            if (
                existing_address.name == address.name
                and existing_address.city == address.city
            ):
                job_list.append(existing_address.job)
            # End for existing_address
        # Record this address and the jobs associated with it
        shipto_list_dupes_dict[address.name] = {
            "address": address,
            "job_list": job_list,
        }
        # Also track addresses seen
        shipto_list.append(address)

    pagevars = {
        "page_title": "GCHUB Daily Report",
        "jobs_due_1": jobs_due_1,
        "jobs_due_2": jobs_due_2,
        "jobs_due_3": jobs_due_3,
        "jobs_due_4": jobs_due_4,
        "jobs_due_5": jobs_due_5,
        "jobs_due_6": jobs_due_6,
        "jobs_due_7": jobs_due_7,
        "jobs_due_8": jobs_due_8,
        "jobs_due_1_total_hours": jobs_due_1_total_hours,
        "jobs_due_1_avg_total_hours": jobs_due_1_avg_total_hours,
        "jobs_due_2_total_hours": jobs_due_2_total_hours,
        "jobs_due_2_avg_total_hours": jobs_due_2_avg_total_hours,
        "jobs_due_3_total_hours": jobs_due_3_total_hours,
        "jobs_due_3_avg_total_hours": jobs_due_3_avg_total_hours,
        "jobs_due_4_total_hours": jobs_due_4_total_hours,
        "jobs_due_4_avg_total_hours": jobs_due_4_avg_total_hours,
        "jobs_due_5_total_hours": jobs_due_5_total_hours,
        "jobs_due_5_avg_total_hours": jobs_due_5_avg_total_hours,
        "jobs_due_6_total_hours": jobs_due_6_total_hours,
        "jobs_due_6_avg_total_hours": jobs_due_6_avg_total_hours,
        "jobs_due_7_total_hours": jobs_due_7_total_hours,
        "jobs_due_7_avg_total_hours": jobs_due_7_avg_total_hours,
        "jobs_due_8_total_hours": jobs_due_8_total_hours,
        "jobs_due_8_avg_total_hours": jobs_due_8_avg_total_hours,
        "shipto_list": shipto_list,
        "shipto_list_dupes": shipto_list_dupes,
        "shipto_list_dupes_dict": shipto_list_dupes_dict,
        "revisions_due_1": revisions_due_1,
        "revisions_due_2": revisions_due_2,
        "revisions_due_3": revisions_due_3,
        "revisions_due_4": revisions_due_4,
        "revisions_due_5": revisions_due_5,
        "revisions_due_6": revisions_due_6,
        "revisions_due_7": revisions_due_7,
        "revisions_due_8": revisions_due_8,
        "jobs_needing_file_out": jobs_needing_file_out,
        "jobs_needing_file_out_today": jobs_needing_file_out_today,
        "jobs_needing_file_out_tomorrow": jobs_needing_file_out_tomorrow,
        "jobs_needing_file_out_day3": jobs_needing_file_out_day3,
        "jobs_needing_file_out_day4": jobs_needing_file_out_day4,
        "jobs_needing_file_out_soon": jobs_needing_file_out_soon,
        "jobs_needing_file_out_overdue": jobs_needing_file_out_overdue,
        "jobs_needing_file_out_today_total": jobs_needing_file_out_today_total,
        "jobs_needing_file_out_tomorrow_total": jobs_needing_file_out_tomorrow_total,
        "jobs_needing_file_out_day3_total": jobs_needing_file_out_day3_total,
        "jobs_needing_file_out_day4_total": jobs_needing_file_out_day4_total,
        "jobs_needing_file_out_soon_total": jobs_needing_file_out_soon_total,
        "jobs_needing_file_out_overdue_total": jobs_needing_file_out_overdue_total,
        "overdue_revisions": overdue_revisions,
        "jobs_overdue": overdue_jobs,
        "jobs_on_hold": jobs_on_hold,
        "jobs_pending": jobs_pending,
        "day0": day0,  # This is today.
        "day1": day1,  # First business day.
        "day2": day2,
        "day3": day3,
        "day4": day4,
        "day5": day5,
        "day6": day6,
        "day7": day7,
        "day8": day8,
        "month_name": calendar.month_name[current_time[1]],
        "year_num": current_time[0],
        "manager_tools": manager_tools,
    }

    context_instance = RequestContext(request)
    return render(request, "workflow/search/todo_list.html", context=pagevars)


def job_todo_estimates(job):
    """Estimate how long it should take an artitst to complete a given job based
    on the job's type and job complexity. Also provides a similar estimate for
    all artists on average. Use by job_todo_list() for manager tools.
    """
    artist_average = 0
    all_artist_average = 0

    if job.workflow.name != "Beverage":
        try:
            # Look up the job complexity
            job_complexity = JobComplexity.objects.get(job=job)

            # Use the get_item_average_hours() function to calculate average.
            artist_averages = get_item_average_hours(
                job_complexity.category, job.type, job.artist
            )
            all_artists_averages = get_item_average_hours(
                job_complexity.category, job.type
            )

            # Count the items
            items = Item.objects.filter(job=job)
            # get_item_average_hours() returns averages for each complexity.
            # Pick out the one that matches this job.
            for entry in artist_averages:
                if entry[0] == job_complexity.complexity:
                    artist_average = entry[1] * items.count()
            for entry in all_artists_averages:
                if entry[0] == job_complexity.complexity:
                    all_artist_average = entry[1] * items.count()
        except Exception:
            pass

    # Return the data as a tuple.
    return (artist_average, all_artist_average)


def list_reports(request):
    """Displays a list of available custom reports.
    Dukes mayo on a saltine cracker is the best!
    I always take a handful with me when
    I'm feedin my hog Myrtle. She loves it too. LOL
    """
    pagevars = {
        "page_title": "Custom Reports",
    }

    return render(request, "workflow/search/list_reports.html", context=pagevars)


class GetReport(ListView):
    """Displays reports."""

    paginate_by = 50
    template_name = "workflow/search/search_results.html"
    table = None
    search = None
    page_title = None

    def get_queryset(self, **kwargs):
        try:
            report_name = self.kwargs["report_name"]
            self.table = report_name.split("_")[0]
            self.search = report_name.split("_")[1]
        except Exception:
            self.table = None
            self.search = None

        this_day = date.today()
        begin_day = this_day + timedelta(days=-7)
        one_year_ago = this_day + timedelta(days=-365)
        two_years_ago = this_day + timedelta(days=-730)

        """
        Report_name should be formatted as table_report, ie. job_hold,
        item_revision, etc... Split on the underscore.
        This should limit repetition in the custom reports.
        """
        qset = []
        # Perform table-based query.
        if self.table == "job":
            qset = Job.objects.exclude(prepress_supplier="Phototype")
        if self.table == "item":
            qset = Item.objects.exclude(job__prepress_supplier="Phototype").exclude(
                job__status="Cancelled"
            )

        # Filter based on user status.
        qset = qset.filter(
            workflow__name__in=general_funcs.get_user_workflow_access(self.request)
        ).order_by("-id")

        # Filter using report_name, part 2.
        if self.search == "unassigned":
            qset = (
                qset.filter(artist__isnull=True)
                .exclude(
                    status__in=(
                        "Complete",
                        "Cancelled",
                    )
                )
                .order_by("-id")
            )
            self.page_title = "Jobs Unassigned"

        elif self.search == "hold":
            qset = qset.filter(status="Hold").order_by("-id")
            self.page_title = "Jobs on Hold"

        elif self.search == "active":
            # recent_log = JobLog.objects.filter(type=JOBLOG_TYPE_ITEM_FILED_OUT).values("item").query
            qset = qset.filter(status="Active", artist__isnull=False).order_by("-id")
            self.page_title = "Assigned Active Jobs"

        elif self.search == "revision":
            revisions = (
                Revision.objects.filter(complete_date__isnull=True).values("item").query
            )
            qset = (
                qset.filter(id__in=revisions)
                .exclude(
                    job__status__in=(
                        "Complete",
                        "Cancelled",
                        "Hold",
                    )
                )
                .order_by("-id")
            )
            self.page_title = "Items Needing Revision"

        elif self.search == "proofed":
            # IDs of items proofed in last X days.
            recent_proofed_log = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
                    event_time__range=(begin_day, this_day),
                )
                .order_by("-id")
                .values("item")
                .query
            )
            # IDs of items filed out in last X days (to exclude)
            recent_filedout_log = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_FILED_OUT,
                    event_time__range=(begin_day, this_day),
                )
                .order_by("-id")
                .values("item")
                .query
            )
            # IDs of items with pending revisions. (to exclude)
            revisions = (
                Revision.objects.filter(complete_date__isnull=True).values("item").query
            )
            qset = (
                qset.filter(id__in=recent_proofed_log)
                .exclude(id__in=recent_filedout_log)
                .exclude(id__in=revisions)
            )
            self.page_title = "Items Recently Proofed/Posted"

        elif self.search == "ninedigit":
            qset = qset.filter(
                fsb_nine_digit_date__range=(begin_day, this_day),
                job__workflow__name="Foodservice",
            ).order_by("-fsb_nine_digit_date")
            self.page_title = "Items with Recent Nine Digit Numbers"

        elif self.search == "fileout":
            recent_log = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_FILED_OUT,
                    event_time__range=(begin_day, this_day),
                )
                .values("item")
                .query
            )
            qset = qset.filter(id__in=recent_log).order_by("-job__id")
            self.page_title = "Items Recently Filed Out"

        elif self.search == "approved-noninedigit":
            recent_log = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_APPROVED,
                    item__job__workflow__name="Foodservice",
                )
                .values("item")
                .query
            )
            recent_log_fileout = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_FILED_OUT,
                    item__job__workflow__name="Foodservice",
                )
                .values("item")
                .query
            )
            qset = (
                qset.filter(id__in=recent_log, fsb_nine_digit="")
                .exclude(id__in=recent_log_fileout)
                .order_by("-job__id")
            )
            self.page_title = "Items Approved w/o Nine Digit Number"
        elif self.search == "new-nutrition":
            # TODO: speed this up.
            # Exclude items recently approved, or with that have been given a item_situation
            # (ie. Lost to Competitor, etc...)
            tracked_items = (
                ItemTracker.objects.filter(type__name="Nutrition Facts")
                .values("item")
                .query
            )

            qset = qset.filter(id__in=tracked_items).order_by("-job__id")

            self.page_title = "Items Using New Nutrition Facts Labels"
        elif self.search == "bioengineered":
            # All the bioengineered beverage labels like BE, BR, BS, etc.
            tracked_items = (
                ItemTracker.objects.filter(
                    type__category__name="Beverage Label", type__name__startswith="B"
                )
                .values("item")
                .query
            )

            qset = qset.filter(id__in=tracked_items).order_by("-job__id")

            self.page_title = "Items Using Bioengineered Labels"
        elif self.search == "proof-notapproved":
            # TODO: speed this up.
            recent_log = (
                JobLog.objects.filter(
                    type=JOBLOG_TYPE_ITEM_PROOFED_OUT,
                    item__job__workflow__name="Foodservice",
                    event_time__range=(two_years_ago, this_day),
                )
                .values("item")
                .query
            )
            recent_log_approved = (
                JobLog.objects.filter(
                    type__in=(JOBLOG_TYPE_ITEM_APPROVED, JOBLOG_TYPE_ITEM_FILED_OUT),
                    item__job__workflow__name="Foodservice",
                )
                .values("item")
                .query
            )
            # Exclude items recently approved, or with that have been given a item_situation
            # (ie. Lost to Competitor, etc...)
            qset = (
                qset.filter(id__in=recent_log, proof_reminder_email_sent__isnull=False)
                .exclude(
                    Q(id__in=recent_log_approved) | Q(item_situation__isnull=False)
                )
                .order_by("-job__id")
            )

            # Filter down further based on item activity since proof reminder.

            ### This section of code was filtering out jobs that Maddy was expection for her report.
            #        filtered_active_ids = []
            #        for item in qset:
            #            if item.is_proof_inactive():
            #                filtered_active_ids.append(item.id)
            #
            #        qset = qset.filter(id__in=filtered_active_ids)

            # This portion of code sends the Items Proofed... report to a seperate output HTML file
            # We are doing this so that the pagination/sorting is done by JavaScript so that the results
            # are sortable (ie by date).

            self.page_title = "Items Proofed, Reminder Sent, No Activity"

        elif self.search == "makeship":
            # MakeShip Jobs from Art Requests.
            artreq_infos = AdditionalInfo.objects.filter(forecast="makeandship")
            artreqs = ArtReq.objects.filter(id__in=artreq_infos.values("artreq"))
            artreq_makeship_jobs = Job.objects.filter(id__in=artreqs.values("job_num"))

            # MakeShip Jobs from eTools. If it has an eTools id we assume its a makeship job.
            etools_makeship_jobs = (
                Job.objects.filter()
                .exclude(e_tools_id__isnull=True)
                .exclude(e_tools_id__exact="")
                .order_by("-id")
            )

            # Merge the two querysets.
            makeship_jobs = artreq_makeship_jobs | etools_makeship_jobs

            # Gather all the items in these jobs.
            qset = qset.filter(job__in=makeship_jobs).order_by("-job__id")

            self.page_title = "Make and Ship Items"

        return qset

    def get_context_data(self, **kwargs):
        context = super(GetReport, self).get_context_data(**kwargs)

        context["page_title"] = self.page_title
        context["type"] = self.table
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        if self.search == "proof-notapproved":
            context["search_number"] = self.get_queryset().count()
            context["object_list"] = self.get_queryset()
            self.template_name = "workflow/search/items_proofed_results.html"
        elif self.search == "makeship":
            self.template_name = "workflow/search/makeship_results.html"
        elif self.search == "bioengineered":
            self.template_name = "workflow/search/bioengineered_results.html"

        return context


@method_decorator(login_required)
def advanced_reports(request):
    """Advanced reporting for internal users only."""
    pagevars = {
        "page_title": "Advanced Reports",
    }

    return render(request, "workflow/search/advanced_reports.html", context=pagevars)


@method_decorator(login_required)
def otifne_report(request, year):
    """Main OTIFNE report view.
    OTIFNE = On time, In full, No errors.
    Report calculates error percentages and late items for both proofing and
    file out the plant.
    """
    jobs = (
        Job.objects.filter(creation_date__year=year)
        .exclude(status__in=("Cancelled", "Hold", "Hold for Art"))
        .exclude(prepress_supplier__in=("PHT", "SGS", "SHK"))
        .values("id")
        .query
    )
    items = Item.objects.filter(job__id__in=jobs, is_deleted=False)

    file_outs = JobLog.objects.filter(
        type=JOBLOG_TYPE_ITEM_FILED_OUT, event_time__year=year
    )
    num_file_outs = file_outs.count()

    error_list = Error.objects.filter(reported_date__year=year).order_by("-id")
    errors_reported = error_list.count()

    # Calculate error percentage.
    if errors_reported and num_file_outs:
        error_percentage = 100 - (float(errors_reported) / num_file_outs) * 100
    else:
        error_percentage = 100

    # Set up blank list.
    overdue_item_set = []

    # Sort eligable items to see how they are.
    for i in items:
        if i.is_overdue():
            overdue_item_set.append(i)

    num_overdue_items = len(overdue_item_set)

    # Calculate proof on time percentage.
    if items and num_overdue_items:
        proof_percentage = 100 - (float(num_overdue_items) / items.count()) * 100
    else:
        proof_percentage = 100

    pagevars = {
        "page_title": "OTIFNE",
        "year": year,
        "error_percentage": error_percentage,
        "proof_percentage": proof_percentage,
        "num_items": items.count(),
    }

    return render(request, "workflow/search/otifne_report.html", context=pagevars)


def get_fedex_shipment_pdf(request, start_date, end_date):
    ship_list = Shipment.objects.filter(
        date_shipped__range=(start_date, end_date)
    ).order_by("-date_shipped")

    workBookDocument = openpyxl.Workbook()
    docSheet1 = workBookDocument.active
    docSheet1.title = "Shipments"

    title = "fedex_shipments_%s_%s" % (start_date, end_date)

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Created"
    docSheet1.cell(row=1, column=2).value = "Recipient"
    docSheet1.cell(row=1, column=3).value = "Job"
    docSheet1.cell(row=1, column=4).value = "Tracking"

    # Used to track which row we're on in the spreadsheet.
    row = 1

    # Write a row in the spreadsheet for each object in the query set.
    for shipment in ship_list:
        if shipment.address:
            if shipment.address.name:
                docSheet1.cell(row=row + 1, column=1).value = str(
                    shipment.address.name.encode("utf-8")
                )
            else:
                docSheet1.cell(row=row + 1, column=1).value = str(
                    shipment.address.company
                )
        else:
            continue
        docSheet1.cell(row=row + 1, column=2).value = str(
            shipment.date_shipped.strftime("%Y-%m-%d")
        )
        docSheet1.cell(row=row + 1, column=3).value = str(shipment.job)
        docSheet1.cell(row=row + 1, column=4).value = shipment.tracking_num

        # Move to the next row
        row += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")

    response["Content-Disposition"] = 'attachment; filename="%s.xlsx"' % title
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response


class fedexShipForm(forms.Form):
    start_date = forms.DateField(required=False)
    end_date = forms.DateField(required=False)

    def __init__(self, *args, **kwargs):
        super(fedexShipForm, self).__init__(*args, **kwargs)
        self.fields["start_date"].label = "Start Date (Older)"
        self.fields["end_date"].label = "End Date (Newer)"


class FedexShipments(ListView):
    """This page is for cindy mueller to query recent fedex shipments so that she can
    compare them to the list of questionable fedex people for review
    """

    paginate_by = 25
    template_name = "workflow/search/fedex_shipments.html"
    # default search is on week
    start_date = None
    end_date = None
    starting_date = None
    ending_date = None

    page_title = "Fedex Shipments"

    def get_queryset(self, **kwargs):
        qset = []
        if self.request.GET:
            dateform = fedexShipForm(self.request.GET)
            if dateform.is_valid():
                self.start_date = dateform.cleaned_data["start_date"]
                self.end_date = dateform.cleaned_data["end_date"]

        if self.start_date is None or self.end_date is None:
            # default search is on week; use UTC-naive current date
            self.start_date = general_funcs._utcnow_naive().date() - timedelta(days=7)
            self.end_date = general_funcs._utcnow_naive().date()

        self.starting_date = self.start_date.strftime("%d-%m-%Y")
        self.ending_date = self.end_date.strftime("%d-%m-%Y")

        qset = Shipment.objects.filter(
            date_shipped__range=(self.start_date, self.end_date)
        ).order_by("-date_shipped")
        return qset

    def get_context_data(self, **kwargs):
        context = super(FedexShipments, self).get_context_data(**kwargs)

        context["page_title"] = self.page_title
        context["starting_date"] = self.starting_date
        context["ending_date"] = self.ending_date
        context["dateform"] = fedexShipForm()
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context

    # Require the user to be logged in to GOLD to view.
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(FedexShipments, self).dispatch(*args, **kwargs)


class InkCovReport(ListView):
    """Used by Evergreen to generate ink coverage reports by plant. Returns color
    info for items that have filed out. Checks for duplicated by bev_item_name
    and returns only the most recent item with that name according to database
    ID.
    """

    paginate_by = 50
    template_name = "workflow/search/inkcov_report.html"
    plant = "plant"

    def get_queryset(self, **kwargs):
        try:
            plant = self.kwargs["plant"]
        except Exception:
            plant = None
        # Get items that have filed out for the given plant.
        fileout_logs = JobLog.objects.filter(
            type=JOBLOG_TYPE_ITEM_FILED_OUT, item__printlocation__plant__name=plant
        ).values("item")
        fileout_items = Item.objects.filter(id__in=fileout_logs).exclude(job__id=99999)

        # Here's how we'll remove duplicates

        # Make a list/dictionary of each unique bev_item_name in our filed-out items
        # and note the most recent item that uses that name.
        unique_names = (
            fileout_items.values("bev_item_name")
            .order_by()
            .annotate(max_id=models.Max("id"))
        )
        # Make a list of just the item IDs.
        unique_items = [str(item["max_id"]) for item in unique_names]
        # Filter filed-out items leaving only the most recent item to use that name.
        ready_items = fileout_items.filter(id__in=unique_items)

        # Gather the colors for the items specified.
        colors = ItemColor.objects.filter(item__in=ready_items).order_by(
            "-item__job__id"
        )

        return colors

    def get_context_data(self, **kwargs):
        context = super(InkCovReport, self).get_context_data(**kwargs)

        try:
            plant = self.kwargs["plant"]
        except Exception:
            plant = "Plant"
        # Gather up the plants involved in this report. Exclude closed plants.
        exclude_plants = ["Clinton", "Framingham", "Raleigh", "Plant City"]
        bev_plants = (
            Plant.objects.filter(workflow__name="Beverage")
            .order_by("name")
            .exclude(name__in=exclude_plants)
        )

        context["page_title"] = "Ink Coverage by Plant"
        context["bev_plants"] = bev_plants
        context["plant"] = plant
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context

    # Require the user to be logged in to GOLD to view.
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(InkCovReport, self).dispatch(*args, **kwargs)


def inkcov_excel(request, plant):
    """Used by Evergreen to generate ink coverage reports by plant. Basically uses
    the same logic as inkcov_report but spits it out as a spreadsheet for
    download.
    """
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Gather up the plants involved in this report. Exclude closed plants.
    exclude_plants = ["Clinton", "Framingham", "Raleigh", "Plant City"]
    bev_plants = (
        Plant.objects.filter(workflow__name="Beverage")
        .order_by("name")
        .exclude(name__in=exclude_plants)
    )

    # Get items that have filed out for the given plant.
    fileout_logs = JobLog.objects.filter(
        type=JOBLOG_TYPE_ITEM_FILED_OUT, item__printlocation__plant__name=plant
    ).values("item")
    fileout_items = Item.objects.filter(id__in=fileout_logs).exclude(job__id=99999)

    # Here's how we'll remove duplicates

    # Make a list/dictionary of each unique bev_item_name in our filed-out items
    # and note the most recent item that uses that name.
    unique_names = (
        fileout_items.values("bev_item_name")
        .order_by()
        .annotate(max_id=models.Max("id"))
    )
    # Make a list of just the item IDs.
    unique_items = [str(item["max_id"]) for item in unique_names]
    # Filter filed-out items leaving only the most recent item to use that name.
    ready_items = fileout_items.filter(id__in=unique_items)

    # Gather the colors for the items specified.
    qset = ItemColor.objects.filter(item__in=ready_items).order_by("-item__job__id")

    title = "Ink Coverage %s" % plant

    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = plant

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Job"
    docSheet1.cell(row=1, column=2).value = "Item"
    docSheet1.cell(row=1, column=3).value = "Color"
    docSheet1.cell(row=1, column=4).value = "Coverage (sq in)"

    # Used to track which row we're on in the spreadsheet.
    row = 1

    # Write a row in the spreadsheet for each object in the query set.
    for color in qset:
        docSheet1.cell(row=row + 1, column=1).value = int(color.item.job.id)
        docSheet1.cell(row=row + 1, column=2).value = str(color.item)
        docSheet1.cell(row=row + 1, column=3).value = str(color.fsb_display_name())
        # Leave this cell blank if the coverage is 'None'
        if color.coverage_sqin:
            docSheet1.cell(row=row + 1, column=4).value = float(color.coverage_sqin)

        # Move to the next row
        row += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")

    response["Content-Disposition"] = 'attachment; filename="%s.xlsx"' % title
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response


class FSB_InkCovReport_Form(forms.Form):
    """Form used to select the time span in the stale charges report."""

    start_date = forms.DateField(required=True)
    end_date = forms.DateField(required=True)
    plants = ["Kenton", "Shelbyville", "Visalia", "Clarksville", "Pittston"]
    plant_choices = Plant.objects.filter(name__in=plants).order_by("name")
    plant = forms.ModelChoiceField(queryset=plant_choices, required=True)

    def __init__(self, *args, **kwargs):
        super(FSB_InkCovReport_Form, self).__init__(*args, **kwargs)
        self.fields["start_date"].label = "Start Date (Older)"
        self.fields["end_date"].label = "End Date (Newer)"
        self.fields["plant"].label = "Plant"


class FSB_InkCovReport(ListView):
    """Used by FSB to generate ink coverage reports by plant. Returns color
    info for items that have filed out. Checks for duplicated by bev_item_name
    and returns only the most recent item with that name according to database
    ID.
    """

    paginate_by = 50
    template_name = "workflow/search/fsb_inkcov_report.html"
    plant = ""
    start_date = None
    end_date = None
    starting_date = None
    ending_date = None

    def get_queryset(self, **kwargs):
        if self.request.GET:
            inkform = FSB_InkCovReport_Form(self.request.GET)
            if inkform.is_valid():
                self.start_date = inkform.cleaned_data["start_date"]
                self.end_date = inkform.cleaned_data["end_date"]
                self.plant = inkform.cleaned_data["plant"]

        if self.start_date is None or self.end_date is None:
            # default search is on week; use UTC-naive current date
            self.start_date = general_funcs._utcnow_naive().date() - timedelta(days=7)
            self.end_date = general_funcs._utcnow_naive().date()

        self.starting_date = self.start_date.strftime("%Y-%m-%d")
        self.ending_date = self.end_date.strftime("%Y-%m-%d")

        # Get items that have filed out for the given plant.
        fileout_logs = JobLog.objects.filter(
            type=JOBLOG_TYPE_ITEM_FILED_OUT,
            event_time__range=(self.start_date, self.end_date),
            item__printlocation__plant__name=self.plant,
        ).values("item")
        fileout_items = Item.objects.filter(id__in=fileout_logs).exclude(job__id=59300)

        # Here's how we'll remove duplicates
        # Make a list/dictionary of each unique bev_item_name in our filed-out items
        # and note the most recent item that uses that name.
        unique_names = (
            fileout_items.values("id").order_by().annotate(max_id=models.Max("id"))
        )
        # Make a list of just the item IDs.
        unique_items = [str(item["max_id"]) for item in unique_names]
        # Filter filed-out items leaving only the most recent item to use that name.
        ready_items = fileout_items.filter(id__in=unique_items)

        # Gather the colors for the items specified.
        colors = ItemColor.objects.filter(item__in=ready_items).order_by(
            "-item__job__id", "item__num_in_job"
        )

        return colors

    def get_context_data(self, **kwargs):
        context = super(FSB_InkCovReport, self).get_context_data(**kwargs)

        context["inkform"] = FSB_InkCovReport_Form()
        context["page_title"] = "Ink Coverage by FSB Plant"
        context["starting_date"] = self.starting_date
        context["ending_date"] = self.ending_date
        context["plant"] = self.plant
        context["extra_link"] = general_funcs.paginate_get_request(self.request)

        return context

    # Require the user to be logged in to GOLD to view.
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(FSB_InkCovReport, self).dispatch(*args, **kwargs)


def fsb_inkcov_excel(request, start_date, end_date, plant):
    """Used by Evergreen to generate ink coverage reports by plant. Basically uses
    the same logic as inkcov_report but spits it out as a spreadsheet for
    download.
    """
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()

    # Get items that have filed out for the given plant.
    fileout_logs = JobLog.objects.filter(
        type=JOBLOG_TYPE_ITEM_FILED_OUT,
        event_time__range=(start_date, end_date),
        item__printlocation__plant__name=plant,
    ).values("item")
    fileout_items = Item.objects.filter(id__in=fileout_logs).exclude(job__id=59300)

    # Here's how we'll remove duplicates
    # Make a list/dictionary of each unique bev_item_name in our filed-out items
    # and note the most recent item that uses that name.
    unique_names = (
        fileout_items.values("id").order_by().annotate(max_id=models.Max("id"))
    )
    # Make a list of just the item IDs.
    unique_items = [str(item["max_id"]) for item in unique_names]
    # Filter filed-out items leaving only the most recent item to use that name.
    ready_items = fileout_items.filter(id__in=unique_items)

    # Gather the colors for the items specified.
    qset = ItemColor.objects.filter(item__in=ready_items).order_by(
        "-item__job__id", "item__num_in_job"
    )

    title = "FSB Ink Coverage %s" % plant

    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = plant

    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Job"
    docSheet1.cell(row=1, column=2).value = "Item Num"
    docSheet1.cell(row=1, column=3).value = "Nine Digit"
    docSheet1.cell(row=1, column=4).value = "Color"
    docSheet1.cell(row=1, column=5).value = "Coverage (sq in)"

    # Used to track which row we're on in the spreadsheet.
    row = 1

    # Write a row in the spreadsheet for each object in the query set.
    for color in qset:
        docSheet1.cell(row=row + 1, column=1).value = int(color.item.job.id)
        docSheet1.cell(row=row + 1, column=2).value = int(color.item.num_in_job)
        docSheet1.cell(row=row + 1, column=3).value = str(color.item.fsb_nine_digit)
        docSheet1.cell(row=row + 1, column=4).value = str(color.fsb_display_name())
        # Leave this cell blank if the coverage is 'None'
        if color.coverage_sqin:
            docSheet1.cell(row=row + 1, column=5).value = float(color.coverage_sqin)

        # Move to the next row
        row += 1

    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]

    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")

    response["Content-Disposition"] = 'attachment; filename="%s.xlsx"' % title
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response


def makeship_excel(request):
    """Create an excel spreadsheet listing all make and ship items. Basically uses
    the same logic as the make and ship item report but spits it out as a
    spreadsheet for download.
    """
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()
    # MakeShip Jobs from Art Requests.
    artreq_infos = AdditionalInfo.objects.filter(forecast="makeandship")
    artreqs = ArtReq.objects.filter(id__in=artreq_infos.values("artreq"))
    artreq_makeship_jobs = Job.objects.filter(id__in=artreqs.values("job_num"))

    # MakeShip Jobs from eTools. If it has an eTools id we assume its a makeship job.
    etools_makeship_jobs = (
        Job.objects.filter()
        .exclude(e_tools_id__isnull=True)
        .exclude(e_tools_id__exact="")
        .order_by("-id")
    )
    # Merge the two querysets.
    makeship_jobs = artreq_makeship_jobs | etools_makeship_jobs

    # Gather all the items in these jobs.
    qset = Item.objects.filter(job__in=makeship_jobs).order_by("-job__id")

    title = "Make and Ship Items"
    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = "Make and Ship"
    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Job"
    docSheet1.cell(row=1, column=2).value = "Size"
    docSheet1.cell(row=1, column=3).value = "Plant"
    docSheet1.cell(row=1, column=4).value = "Press"
    docSheet1.cell(row=1, column=5).value = "Approval Date"
    docSheet1.cell(row=1, column=6).value = "Item Num. Date"
    docSheet1.cell(row=1, column=7).value = "Final File Date"
    docSheet1.cell(row=1, column=8).value = "Artist"
    docSheet1.cell(row=1, column=9).value = "Salesperson"
    # Used to track which row we're on in the spreadsheet.
    row = 1
    # Write a row in the spreadsheet for each object in the query set.
    for item in qset:
        docSheet1.cell(row=row + 1, column=1).value = int(item.job.id)
        docSheet1.cell(row=row + 1, column=2).value = str(item)
        if item.printlocation:
            if item.printlocation.plant:
                docSheet1.cell(row=row + 1, column=3).value = str(
                    item.printlocation.plant
                )
            if item.printlocation.press:
                docSheet1.cell(row=row + 1, column=4).value = str(
                    item.printlocation.press
                )
        if item.approval_date():
            docSheet1.cell(row=row + 1, column=5).value = str(
                item.approval_date().strftime("%Y-%m-%d")
            )
        if item.fsb_nine_digit_date:
            docSheet1.cell(row=row + 1, column=6).value = str(item.fsb_nine_digit_date)
        if item.final_file_date():
            docSheet1.cell(row=row + 1, column=7).value = str(
                item.final_file_date().strftime("%Y-%m-%d")
            )
        docSheet1.cell(row=row + 1, column=8).value = str(item.job.artist)
        docSheet1.cell(row=row + 1, column=9).value = str(item.job.salesperson)

        # Move to the next row
        row += 1
    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]
    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="%s.xlsx"' % title
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response


def bioengineered_excel(request):
    """Create an excel spreadsheet listing all bioengineered items. Basically uses
    the same logic as the bioengineered item report but spits it out as a
    spreadsheet for download.
    """
    # Setup the Worksheet
    workBookDocument = openpyxl.Workbook()
    # All the bioengineered beverage labels like BE, BR, BS, etc.
    tracked_items = (
        ItemTracker.objects.filter(
            type__category__name="Beverage Label", type__name__startswith="B"
        )
        .values("item")
        .query
    )
    qset = (
        Item.objects.filter(id__in=tracked_items)
        .order_by("-job__id")
        .exclude(job__status="Cancelled")
    )

    title = "Bioengineered Items"
    # Add a page in the spreadsheet
    docSheet1 = workBookDocument.active
    docSheet1.title = "Bioengineered Items"
    # Label column headings
    docSheet1.cell(row=1, column=1).value = "Job"
    docSheet1.cell(row=1, column=2).value = "Size"
    docSheet1.cell(row=1, column=3).value = "Plant"
    docSheet1.cell(row=1, column=4).value = "Press"
    docSheet1.cell(row=1, column=5).value = "Approval Date"
    docSheet1.cell(row=1, column=6).value = "Item Num. Date"
    docSheet1.cell(row=1, column=7).value = "Final File Date"
    docSheet1.cell(row=1, column=8).value = "Artist"
    docSheet1.cell(row=1, column=9).value = "Salesperson"
    # Used to track which row we're on in the spreadsheet.
    row = 1
    # Write a row in the spreadsheet for each object in the query set.
    for item in qset:
        docSheet1.cell(row=row + 1, column=1).value = int(item.job.id)
        docSheet1.cell(row=row + 1, column=2).value = str(item)
        if item.printlocation:
            if item.printlocation.plant:
                docSheet1.cell(row=row + 1, column=3).value = str(
                    item.printlocation.plant
                )
            if item.printlocation.press:
                docSheet1.cell(row=row + 1, column=4).value = str(
                    item.printlocation.press
                )
        if item.approval_date():
            docSheet1.cell(row=row + 1, column=5).value = str(
                item.approval_date().strftime("%Y-%m-%d")
            )
        if item.fsb_nine_digit_date:
            docSheet1.cell(row=row + 1, column=6).value = str(item.fsb_nine_digit_date)
        if item.final_file_date():
            docSheet1.cell(row=row + 1, column=7).value = str(
                item.final_file_date().strftime("%Y-%m-%d")
            )
        docSheet1.cell(row=row + 1, column=8).value = str(item.job.artist)
        docSheet1.cell(row=row + 1, column=9).value = str(item.job.salesperson)

        # Move to the next row
        row += 1
    # Freeze the top row of column headings.
    docSheet1.panes_frozen = docSheet1["B2"]
    with NamedTemporaryFile() as tmp:
        workBookDocument.save(tmp.name)
        output = BytesIO(tmp.read())
    response = HttpResponse(content=output, content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="%s.xlsx"' % title
    # This cookie triggers the "on successful download" from jquery which triggers the modal closing
    response.set_cookie(key="fileDownload", value="true", path="/")
    return response
